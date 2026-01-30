from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError, MissingError
from odoo.tools import float_compare
# Decimal precision is now a standard field attribute or accessed via tools
# from odoo.addons.base.models.res_rounding import decimal_precision as dp

import io
import base64
import csv
import logging
import openpyxl
import xlsxwriter
import xlwt  # Note: xlwt is for .xls; openpyxl/xlsxwriter are preferred for .xlsx
from datetime import datetime, timedelta, date
from dateutil.relativedelta import relativedelta
from lxml import etree

_logger = logging.getLogger(__name__)


class SdfRegister(models.Model):
    _name = 'sdf.register'
    _inherit = ['mail.thread', 'mail.activity.mixin']  # Added activity mixin for modern chatter
    _description = 'SDF Registration'

    name = fields.Char(string="Name", required=True)
    # Ensure you have a 'name' or '_rec_name' for modern Odoo views


class ProductTemplate(models.Model):
    _inherit = 'product.template'

    mark_asset = fields.Boolean(string='Mark as Asset', default=False)


class StockPicking(models.Model):
    _inherit = 'stock.picking'

    mark_asset = fields.Boolean(string='Declare as Asset', default=False)
    # In Odoo 18, journal_id is often linked through moves, but we can keep it for custom logic
    journal_id = fields.Many2one('account.journal', string='Related Journal')
    asset_acc_id = fields.Many2one('account.account', string='Asset Account')


class AccountMove(models.Model):
    _inherit = 'account.move'

    is_expense_voucher_created = fields.Boolean(string="Expense Voucher", default=False)
    batch_no = fields.Char(string='Batch No')
    is_lump_sum = fields.Boolean(string='Is a DHET Entry?')
    scheme_year_id = fields.Many2one('scheme.year', string='Scheme Year')
    # account_analytic_id is now managed via analytic_distribution in Odoo 18
    account_analytic_id = fields.Many2one('account.analytic.account', string='Analytic Account')

    # API change: Odoo 18 handles name_get via _compute_display_name or automatic name field
    def name_get(self):
        return [(move.id, move.name) for move in self]

    @api.onchange('is_lump_sum')
    def _onchange_lump_sum(self):
        if self.is_lump_sum:
            dhet_journal = self.env['account.journal'].search([('name', '=', 'DHET Journal')], limit=1)
            if dhet_journal:
                self.journal_id = dhet_journal.id

    @api.onchange('journal_id')
    def _onchange_journal_id(self):
        if self.journal_id and hasattr(self.journal_id, 'account_analytic_id'):
            if self.journal_id.account_analytic_id:
                self.account_analytic_id = self.journal_id.account_analytic_id.id

    def action_post(self):
        # Legacy button_validate is now action_post in Odoo 18
        # We perform checks before calling super()

        cfo_group = self.env.ref('hwseta_finance.group_cfo')  # Using XML ID is safer
        fm_config = self.env['leavy.income.config'].search([], limit=1)

        if not fm_config or fm_config.fm_approval_limit == 0:
            raise UserError(_('Please configure Financial Manager amount approval limit inside Admin Configuration!'))

        fm_max_limit = fm_config.fm_approval_limit
        total_amount = sum(self.line_ids.mapped('debit'))  # Simpler way to get total

        if total_amount > fm_max_limit:
            if not self.env.user.has_group('hwseta_finance.group_cfo'):
                raise UserError(_('Only CFO can post this entry. Amount exceeds R %s!') % (fm_max_limit))

        res = super(AccountMove, self).action_post()

        # Expense Organisation Invoice Creation Logic
        for rec in self:
            invoice_line_vals = []
            for line in rec.line_ids:
                if line.name and line.name.startswith("Mandatory Grant") and line.move_line_type == 'income':
                    product = self.env['product.product'].search([('name', '=', 'Mandatory Grant')], limit=1)

                    # Modern ORM instead of self._cr.execute
                    wsp_submission = self.env['wsp.submission.track'].search([
                        ('employer_id', '=', line.partner_id.id),
                        ('status', '=', 'accepted'),
                        ('scheme_year_id', '=', line.scheme_year_id.id)
                    ], limit=1)

                    if line.scheme_year_id.state == 'closed' and wsp_submission:
                        income_account = product.property_account_income_id or product.categ_id.property_account_income_categ_id

                        invoice_line_vals.append((0, 0, {
                            'product_id': product.id,
                            'name': 'Mandatory Grant',
                            'quantity': 1,
                            'price_unit': line.credit or line.debit,
                            'account_id': income_account.id,
                        }))

            if invoice_line_vals:
                # Odoo 18 uses 'account.move' for invoices with move_type 'in_invoice'
                expense_bill = self.env['account.move'].create({
                    'move_type': 'in_invoice',
                    'partner_id': rec.partner_id.id or line.partner_id.id,
                    'scheme_year_id': rec.scheme_year_id.id,
                    'batch_no': rec.batch_no,
                    'invoice_date': fields.Date.today(),
                    'invoice_line_ids': invoice_line_vals,
                })
                rec.is_expense_voucher_created = True

        return res


class AccountMoveLine(models.Model):
    _inherit = 'account.move.line'

    scheme_year_id = fields.Many2one('scheme.year', string='Scheme Year')
    project_id = fields.Many2one('project.project', string="Project")
    move_line_type = fields.Selection([
        ('income', 'Income'),
        ('expense', 'Expense')
    ], string='Type')


class WspSubmissionTrack(models.Model):
    _name = 'wsp.submission.track'
    _description = 'WSP Submission Tracking'
    _inherit = ['mail.thread', 'mail.activity.mixin']

    name = fields.Char(string='Name', tracking=True)
    # fiscal_year: account.fiscalyear is removed in standard Odoo 18.
    # Usually replaced by account.fiscal.year (Enterprise) or a custom date range.
    fiscal_year_id = fields.Many2one('account.fiscal.year', string='Year')
    status = fields.Selection([
        ('draft', 'Draft'),
        ('submitted', 'Submitted'),
        ('evaluated', 'Evaluated'),
        ('rejected', 'Rejected'),
        ('accepted', 'Accepted'),
        ('query', 'Query')
    ], string='Status', default='draft', tracking=True)

    employer_id = fields.Many2one('res.partner', string='Related Employer', tracking=True)
    scheme_year_id = fields.Many2one('scheme.year', string='Scheme Year')
    date_created = fields.Date("Date Created", default=fields.Date.context_today)
    wsp_date_submitted = fields.Date("WSP Date Submitted")
    last_user_evaluated_updated = fields.Char("Submitted by")
    approved_date = fields.Date("Approved Date")
    approved_by = fields.Char("Approved By")
    rejected_date = fields.Date("Rejected Date")
    rejected_by = fields.Char("Rejected By")


class StockPicking(models.Model):
    _inherit = 'stock.picking'

    def button_validate(self):
        """
        In Odoo 18, we override button_validate on the picking itself
        instead of the transfer wizard.
        """
        res = super(StockPicking, self).button_validate()

        for picking in self:
            if picking.state == 'done' and picking.mark_asset:
                # 1. Verify Purchase Order Connection
                # Odoo 18 uses 'origin' or 'purchase_id' directly on picking
                purchase_order = self.env['purchase.order'].search([
                    ('name', '=', picking.origin)
                ], limit=1)

                if not purchase_order:
                    # Optional: Log a message instead of raising error if picking is already validated
                    continue

                # 2. Get or Create Asset Model (formerly Asset Category)
                asset_model = self.env['account.asset'].search([
                    ('name', '=', 'HWSETA Assets'),
                    ('state', '=', 'model')  # In v18, Categories are Models
                ], limit=1)

                if not asset_model:
                    asset_model = self.env['account.asset'].create({
                        'name': 'HWSETA Assets',
                        'state': 'model',
                        'account_asset_id': picking.asset_acc_id.id,
                        'account_depreciation_id': picking.asset_acc_id.id,
                        'account_depreciation_expense_id': picking.asset_acc_id.id,
                        'journal_id': picking.journal_id.id or purchase_order.journal_id.id,
                    })

                # 3. Create Asset for each relevant move line
                for move in picking.move_ids:
                    if not move.product_id.mark_asset:
                        continue

                    # Find matching price from PO
                    po_line = purchase_order.order_line.filtered(lambda l: l.product_id == move.product_id)
                    unit_price = po_line[0].price_unit if po_line else 0.0

                    # Create the actual Asset record
                    self.env['account.asset'].create({
                        'name': move.product_id.name,
                        'model_id': asset_model.id,
                        'original_value': unit_price * move.quantity_done,
                        'acquisition_date': fields.Date.context_today(self),
                        'asset_type': 'purchase',
                        'state': 'draft',
                        # Custom field mapping from your original code
                        'product_id': move.product_id.id,
                        'location_id': picking.location_dest_id.id
                    })
        return res

# class ir_attachment(models.Model):
#     _inherit = 'ir.attachment'
#
#     @api.model
#     def default_get(self, fields_list):
#         res = super(ir_attachment, self).default_get(fields_list)
#         context = self._context.copy()
#         if context and context.get('emp_doc',False):
#             res.update({'name':'Document'})
#         if context and context.get('doc_name',False):
#             res.update({'name':context['doc_name']})
#         return res
#
#     @api.model
#     def create(self, vals):
#         context = self._context.copy()
#         if context and context.get('emp_doc',False) and vals.get('datas_fname',False):
#             vals.update({'name' : vals.get('datas_fname',False)})
#         return super(ir_attachment,self).create(vals)
#
# ir_attachment()


class EmployerDocumentUpload(models.Model):
    _name = 'employer.document.upload'
    _description = 'Employer Document Upload'

    name = fields.Char(string='Document Name')
    # In Odoo 18, it is standard to use fields.Binary for file uploads
    # Adding filename storage allows Odoo to keep the original file extension
    document = fields.Binary(string='Document File', attachment=True)
    document_filename = fields.Char("Filename")

    employer_id = fields.Many2one(
        'res.partner',
        string='Employer',
        domain=[('is_company', '=', True)]  # Usually employers are companies in v18
    )


class ResPartnerChilds(models.Model):
    _name = 'res.partner.childs'
    _description = 'Related Employer Contacts'

    name = fields.Char(string='Name', required=True)
    employer_id = fields.Many2one('res.partner', string='Parent Employer')
    emp_child_id = fields.Many2one('res.partner', string='Related Employer Contact')
    sdl_number = fields.Char(string='SDL Number')
    seta_id = fields.Many2one('seta.branches', string='SETA ID')
    sic_code = fields.Many2one('hwseta.sic.master', string='SIC Code')
    email = fields.Char(string='Email')
    phone = fields.Char(string='Phone')
    mobile = fields.Char(string='Mobile')

    @api.onchange('employer_id')
    def _onchange_employer_id(self):
        """
        In Odoo 18, we don't return a dictionary.
        We simply assign values to self.
        """
        if self.employer_id:
            emp = self.employer_id
            self.sdl_number = emp.employer_sdl_no
            self.seta_id = emp.employer_seta_id.id if emp.employer_seta_id else False
            self.sic_code = emp.empl_sic_code.id if emp.empl_sic_code else False
            self.email = emp.email
            self.phone = emp.phone
            self.mobile = emp.mobile


class EmployerSdlNo(models.Model):
    _name = 'employer.sdl.no'
    _description = 'Employer SDL Numbers'

    name = fields.Char(string='SDL Number', required=True)
    employer_id = fields.Many2one('res.partner', string='Employer')

## Employer Related fields can be accessed here from hwseta_person module.
class ResPartner(models.Model):
    _inherit = 'res.partner'

    # Core Fields
    update_disclaimer = fields.Boolean()
    emp_walfare = fields.Boolean(string='Welfare')
    disclaimer = fields.Boolean()
    sdf_id = fields.Many2one('sdf.register', string='SDF')
    wsp_submitted = fields.Boolean(string='WSP Submitted')
    leavy_exempted = fields.Boolean(string='Levy Exempted')  # Note: 'levy' is usually spelled with one 'a'

    # Relationships
    leavy_history_ids = fields.One2many('leavy.history', 'employer_id', string='Levy History')
    empl_sic_code = fields.Many2one('hwseta.sic.master', string='SIC Code')
    empl_sic_code_id = fields.Char(string='SIC Code ID')
    npo_ngo = fields.Boolean(string='NPO/NGO')
    other = fields.Boolean(string='Other')
    empl_status = fields.Char(string="Employer Status")

    participated_project_ids = fields.One2many('project.participated', 'participated_employer_id',
                                               string='Participated in Projects')
    doc_upload_ids = fields.One2many('employer.document.upload', 'employer_id', string='Document Uploads')

    parent_employer_id = fields.Many2one('res.partner', string='Parent Employer', domain=[('is_company', '=', True)])
    child_emp_ids = fields.One2many('res.partner', 'parent_employer_id', string='Child Organisations')
    wsp_submission_ids = fields.One2many('wsp.submission.track', 'employer_id', string='WSP Submissions')

    # DHET / SETA Registration Fields
    ext_emp_reg_number_type = fields.Selection([
        ('cipro_number', 'Cipro Number'),
        ('comp_reg_no', 'Company Registration Number')
    ], string='Extra Registration Number Type')

    # In Odoo 18, use tracking=True instead of track_visibility
    ext_employer_registration_number = fields.Char(string='Extra Registration Number', tracking=True)
    sars_number = fields.Char(string='SARS Number', tracking=True)
    ext_total_annual_payroll = fields.Float(string='Extra Total Annual Payroll')

    # Address & Municipality (Modernizing the structure)
    physical_municipality = fields.Many2one('res.municipality', string='Physical Municipality')
    ext_physical_address_1 = fields.Char(string='Extra Physical Address1', tracking=True)
    ext_physical_code = fields.Char(string='Extra Physical Zip', tracking=True)
    ext_province_code_physical = fields.Many2one('res.country.state', string='Extra Physical Province Code',
                                                 tracking=True)

    _sql_constraints = [
        ('sdl_uniq', 'unique(employer_sdl_no)', 'SDL Number must be unique per Employer!'),
    ]

    # Modern Onchange Pattern
    @api.onchange('empl_sic_code')
    def _onchange_sic_code(self):
        if self.empl_sic_code:
            self.empl_sic_code_id = self.empl_sic_code.name

    @api.onchange('ext_province_code_physical')
    def _onchange_province_physical(self):
        if self.ext_province_code_physical:
            # Assumes country_for_province is a helper method in your class
            self.emp_country_code_physical = self.ext_province_code_physical.country_id.id

    # Modern Create Pattern
    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            # Set Employer Status based on Booleans
            if vals.get('npo_ngo'):
                vals['empl_status'] = 'NPO / NGO'
            elif vals.get('university'):
                vals['empl_status'] = 'University'
            elif vals.get('college'):
                vals['empl_status'] = 'TVET College'
            elif vals.get('employer_department'):
                vals['empl_status'] = 'Government'
            elif vals.get('other'):
                vals['empl_status'] = 'Other'

        partners = super(ResPartner, self).create(vals_list)

        for partner in partners:
            # Sync SDL table
            if partner.employer_sdl_no:
                self.env['employer.sdl.no'].create({
                    'name': partner.employer_sdl_no,
                    'employer_id': partner.id
                })

            # User Creation logic (Portal Access)
            if partner.employer:
                self._create_employer_portal_user(partner)

        return partners

    def _create_employer_portal_user(self, partner):
        user_obj = self.env['res.users']
        # Fetch groups by XML ID for stability
        portal_group = self.env.ref('base.group_portal')
        # Custom group XML ID would go here

        user_obj.create({
            'name': partner.name,
            'login': str(partner.employer_sdl_no),
            'partner_id': partner.id,
            'email': partner.email,
            'groups_id': [(6, 0, [portal_group.id])],
        })

    ## Method for checking whether the string is in Uppercase or not.
    def _check_uppercase_v18(self, value, field_name):
        """Standardized uppercase check using regex."""
        if value and not value.isupper():
            raise ValidationError(_("%s must be in all uppercase characters!") % field_name)

    def _is_forbidden(self, value):
        """Check for forbidden placeholder strings."""
        forbidden = {'%UNKNOWN%', '%AS ABOVE%', '%SOOSBO%', '%DELETE%', '%N/A%',
                     'NA', 'U', 'NONE', 'GEEN', '0', 'TEST', '%ONTBREEK%', 'NILL'}
        if value and value.upper() in forbidden:
            raise ValidationError(_("The value '%s' is not allowed in this field!") % value)

    def validate_employer_v18(self, vals):
        """Comprehensive validation suite for HWSETA Employer data."""

        # 1. SDL Number (Main and Branch)
        for field in ['employer_sdl_no', 'employer_main_sdl_no']:
            if vals.get(field):
                val = vals[field]
                if val.startswith(' '):
                    raise ValidationError(
                        _('%s should not start with a blank space!') % field.replace('_', ' ').title())
                if not re.match(r"^[LN]\d{9}$", val):
                    raise ValidationError(
                        _('%s must start with L or N followed by 9 digits!') % field.replace('_', ' ').title())

        # 2. Site and Registration Numbers
        for field in ['employer_site_no', 'employer_registration_number', 'employer_approval_status_num']:
            if vals.get(field):
                val = vals[field]
                if val.startswith(' '):
                    raise ValidationError(
                        _('%s should not start with a blank space!') % field.replace('_', ' ').title())
                self._check_uppercase_v18(val, field.replace('_', ' ').title())

        # 3. Numeric Strings (Phone, Fax, Lat/Long, Address Codes)
        numeric_fields = [
            'phone', 'fax', 'employer_physical_address_code', 'employer_postal_address_code',
            'employer_latitude_degree', 'employer_latitude_minutes', 'employer_latitude_seconds',
            'employer_longitude_degree', 'employer_longitude_minutes', 'employer_longitude_seconds'
        ]
        for field in numeric_fields:
            if vals.get(field):
                val = vals[field]
                if not re.match(r"^[0-9\+\-\s\(\)]+$", val):
                    raise ValidationError(
                        _("%s should contain only numeric characters!") % field.replace('_', ' ').title())
                self._is_forbidden(val)

        # 4. Dates
        today = date.today()
        if vals.get('employer_approval_status_start_date'):
            start_dt = fields.Date.to_date(vals['employer_approval_status_start_date'])
            if start_dt > today:
                raise ValidationError(_('Approval Start Date cannot be in the future!'))

        if vals.get('employer_date_stamp'):
            stamp_dt = fields.Date.to_date(vals['employer_date_stamp'])
            if stamp_dt > today or stamp_dt.year < 1900:
                raise ValidationError(_('Date Stamp must be between 1900 and today!'))

    # --- CORE LOGIC ---

    def write(self, vals):
        # 1. Update Employer Status based on Booleans
        status_map = {
            'npo_ngo': 'NPO / NGO',
            'university': 'University',
            'college': 'TVET College',
            'employer_department': 'Government',
            'other': 'Other'
        }
        for key, label in status_map.items():
            if vals.get(key):
                vals['empl_status'] = label

        # 2. Run Validations (Uncomment if needed)
        # if self.employer or vals.get('employer'):
        #     self.validate_employer_v18(vals)

        res = super(ResPartner, self).write(vals)

        # 3. Handle Automated Group Assignments (Enrollment Group)
        for record in self:
            if record.employer and record.user_ids:
                user = record.user_ids[0]  # Primary user linked to partner

                # Check for active fiscal year
                today = date.today()
                # In Odoo 18, we check account.fiscal.year (Enterprise)
                # or a custom date range on the WSP submission
                active_year = self.env['account.fiscal.year'].search([
                    ('date_from', '<=', today),
                    ('date_to', '>=', today)
                ], limit=1)

                enrollment_group = self.env.ref('your_module.group_enrollment', raise_if_not_found=False)

                if active_year and enrollment_group:
                    has_wsp = record.wsp_submission_ids.filtered(
                        lambda w: w.fiscal_year_id == active_year and w.status == 'accepted'
                    )

                    if has_wsp:
                        user.write({'groups_id': [(4, enrollment_group.id)]})
                    else:
                        user.write({'groups_id': [(3, enrollment_group.id)]})

        return res


class ProjectParticipated(models.Model):
    _name = 'project.participated'
    _description = 'Project Participation History'

    project_id = fields.Many2one('project.project', string='Project', ondelete='cascade')
    project_type_id = fields.Many2one('hwseta.project.types', string='Project Types')
    participated_employer_id = fields.Many2one('res.partner', string='Employer', domain=[('is_company', '=', True)])
    participated_provider_id = fields.Many2one('res.partner', string='Provider')
    participate_date = fields.Date(string='Date', default=fields.Date.context_today)


class ProjectProject(models.Model):
    _inherit = 'project.project'

    def add_partners(self):
        """Logic for adding specific selected partners to participation history."""
        res = super(ProjectProject, self).add_partners()

        # Search for selected relations
        emp_proj_relations = self.env['partner.project.rel'].search([
            ('select_emp', '=', True),
            ('emp_project_id', '=', self.id)
        ])
        self._update_participation_history(emp_proj_relations.mapped('employer_id'))
        return res

    def add_all_partners(self):
        """Logic for adding all linked partners to participation history."""
        res = super(ProjectProject, self).add_all_partners()

        emp_proj_relations = self.env['partner.project.rel'].search([
            ('emp_project_id', '=', self.id)
        ])
        self._update_participation_history(emp_proj_relations.mapped('employer_id'))
        return res

    def clear_partners(self):
        """Logic to remove partners from history when cleared from project."""
        res = super(ProjectProject, self).clear_partners()

        emp_proj_relations = self.env['partner.project.rel'].search([
            ('emp_project_id', '=', self.id)
        ])
        employer_ids = emp_proj_relations.mapped('employer_id')

        # Find and remove existing history entries for these employers on this project
        history_entries = self.env['project.participated'].search([
            ('project_id', '=', self.id),
            ('participated_employer_id', 'in', employer_ids.ids)
        ])
        if history_entries:
            # Command.delete(id) is the modern equivalent of (2, id)
            for employer in employer_ids:
                entries_to_rem = history_entries.filtered(lambda x: x.participated_employer_id == employer)
                if entries_to_rem:
                    employer.write({
                        'participated_project_ids': [Command.delete(e.id) for e in entries_to_rem]
                    })
        return res

    def _update_participation_history(self, employers):
        """Private helper to refresh history records."""
        proj_part_obj = self.env['project.participated']

        for employer in employers:
            # 1. Remove old entries for this project to avoid duplicates
            existing = proj_part_obj.search([
                ('project_id', '=', self.id),
                ('participated_employer_id', '=', employer.id)
            ])
            if existing:
                employer.write({'participated_project_ids': [Command.delete(e.id) for e in existing]})

            # 2. Create new history entry
            proj_part_obj.create({
                'project_id': self.id,
                'project_type_id': getattr(self, 'project_types', False) and self.project_types.id,
                'participated_employer_id': employer.id,
                'participate_date': date.today(),
            })


class AccountJournal(models.Model):
    _inherit = 'account.journal'

    # Standardizing 'leavy' to 'levy' is recommended,
    # but I've kept your spelling for database consistency.
    is_leavy_journal = fields.Boolean(string='Leavy Journal', default=False)

    # In Odoo 18, analytic accounts are often handled via 'account.analytic.plan'
    # but a direct Many2one is still valid for specific business logic.
    account_analytic_id = fields.Many2one(
        'account.analytic.account',
        string='Analytic Account',
        help="Default analytic account for moves in this journal."
    )

## This class is used to maintain Leavy Paid History for perticular Employer.
class LevyHistory(models.Model):
    _name = 'leavy.history'
    
    period = fields.Char(string='Period')
    leavy_month = fields.Char(string='Levy Month')
    scheme_year_id = fields.Many2one('scheme.year','Scheme Year')
    mand_grant_amt = fields.Float(string='Mand. Grant Amt')
    desc_grant_amt = fields.Float(string='Desc. Grant Amt')
    admn_grant_amt = fields.Float(string='Adm. Grant Amt')
    penalties = fields.Float(string='Penalty')
    interest = fields.Float(string='Interest')
    total_amt = fields.Float(string='Total Amt')
    employer_id = fields.Many2one('res.partner', string='Employer', domain=[('employer','=',True)])


class AccountMove(models.Model):
    _inherit = 'account.move'

    employer = fields.Boolean(string='Employer')
    employer_department = fields.Boolean(string='Department')
    # Odoo 18 uses Date ranges; account.period was removed years ago.
    scheme_year_id = fields.Many2one('scheme.year', string='Scheme Year')
    batch_no = fields.Char('Batch No.')
    journal_entry_ref = fields.Char('Reference')

    # Adding 'Suspend' to the state selection
    state = fields.Selection(selection_add=[('suspend', 'Suspend')], ondelete={'suspend': 'cascade'})

    def action_post(self):
        """
        In Odoo 18, action_post replaces action_move_create.
        We override it to ensure our custom fields (scheme_year_id, batch_no)
        propagate from the Invoice to the Journal Items (line_ids).
        """
        # Custom Validation before posting
        for move in self:
            if move.is_invoice(include_receipts=True):
                if not move.invoice_line_ids:
                    raise UserError(_('No Invoice Lines! Please create some invoice lines.'))

                # Propagate scheme_year_id to all journal items
                if move.scheme_year_id:
                    move.line_ids.write({
                        'scheme_year_id': move.scheme_year_id.id,
                        'move_line_type': 'expense'  # As per your original logic
                    })

                # Use journal_entry_ref as the move reference if provided
                if move.journal_entry_ref:
                    move.ref = move.journal_entry_ref

        return super(AccountMove, self).action_post()


class AccountMoveLine(models.Model):
    _inherit = 'account.move.line'

    scheme_year_id = fields.Many2one('scheme.year', string='Scheme Year')
    move_line_type = fields.Selection([
        ('income', 'Income'),
        ('expense', 'Expense')
    ], string='Type')


class AccountPayment(models.Model):
    _inherit = 'account.payment'

    # Custom Fields
    employer = fields.Boolean(string='Employer')
    organisation_id = fields.Many2one('res.partner', string='Organisation')
    bulk_payment = fields.Boolean(string='Organisation Bulk Payment')
    from_date = fields.Date(string='From Date')
    to_date = fields.Date(string='To Date')
    scheme_year_id = fields.Many2one('scheme.year', string='Scheme Year')

    # --- Print Report ---
    def action_print_payment_report(self):
        """ Modern report triggering in Odoo 18 """
        return self.env.ref('hwseta_finance.organisation_payment_report_action').report_action(self)

    # --- Logic: Batch Payment Computation ---
    def compute_bulk_matching_lines(self):
        """
        Replaces recompute_voucher_lines_datewise.
        In Odoo 18, we search for 'move_line_ids' that are open and match criteria.
        """
        self.ensure_one()
        if not self.bulk_payment:
            return

        # Get the Provision Account from your custom config
        admin_config = self.env['leavy.income.config'].search([], limit=1)
        if not admin_config or not admin_config.provision_acc:
            raise UserError(_("Please configure the Provision Account in Levy Income Config."))

        provision_account = admin_config.provision_acc

        # Determine account type (Payable/Receivable) based on payment type
        target_type = 'payable' if self.payment_type == 'outbound' else 'receivable'

        # Build Domain for Move Lines (Unreconciled items)
        domain = [
            ('parent_state', '=', 'posted'),
            ('account_id', '=', provision_account.id),
            ('account_id.account_type', '=', target_type),
            ('reconciled', '=', False),
        ]

        if self.scheme_year_id:
            domain.append(('scheme_year_id', '=', self.scheme_year_id.id))
        if self.from_date:
            domain.append(('date', '>=', self.from_date))
        if self.to_date:
            domain.append(('date', '<=', self.to_date))

        move_lines = self.env['account.move.line'].search(domain)

        if not move_lines:
            raise UserError(_("No matching open items found for the selected criteria."))

        # In Odoo 18, we don't 'write' lines to a voucher.
        # We perform reconciliation after posting or use the 'Batch Payment' feature.
        # For this custom implementation, we can link these lines to the payment.
        return move_lines

    # --- Validate / Post Payment ---
    def action_post(self):
        """ Replaces proforma_voucher and includes email logic """
        res = super(AccountPayment, self).action_post()

        # Email Notification Logic
        template = self.env.ref('hwseta_finance.email_template_org_payment', raise_if_not_found=False)
        if template:
            for payment in self:
                # Odoo 18 uses .send_mail() on the template object
                template.send_mail(payment.id, force_send=True)

        # If Bulk Payment, you would trigger the reconciliation logic here
        for payment in self:
            if payment.bulk_payment:
                lines_to_reconcile = payment.compute_bulk_matching_lines()
                # Logic to reconcile payment lines with move_lines...
        return res


class AccountMoveLine(models.Model):
    _inherit = 'account.move.line'

    # Keeping employer_id on move lines for tracking purposes in Odoo 18
    employer_id = fields.Many2one('res.partner', string='Employer')


### Inherited this class for implementing bulk payment condition.
# class account_move_reconcile(models.Model):
#     _inherit = 'account.move.reconcile'
#
#
#     @api.v7
#     def _check_same_partner(self, cr, uid, ids, context=None):
#         ''' Overriden this method so as to add bulk payment condition. For normal payment
#         it will check for the move line partner with voucher partner. in bulk payment this condition will not be considered.'''
#         for reconcile in self.browse(cr, uid, ids, context=context):
#             move_lines = []
#             if not reconcile.opening_reconciliation:
#                 if reconcile.line_id:
#                     first_partner = reconcile.line_id[0].partner_id.id
#                     move_lines = reconcile.line_id
#                 elif reconcile.line_partial_ids:
#                     first_partner = reconcile.line_partial_ids[0].partner_id.id
#                     move_lines = reconcile.line_partial_ids
#                 ### Condition for Bulk Payment
#                 for move_line in move_lines :
#                     voucher_line_id = self.pool.get('account.voucher.line').search(cr, uid,[('move_line_id','=',move_line.id)])
#                     if voucher_line_id :
#                         voucher_line_data = self.pool.get('account.voucher.line').browse(cr,uid,voucher_line_id[0])
#                         if voucher_line_data.voucher_id.bulk_payment :
#                             return True
#                 if any([(line.account_id.type in ('receivable', 'payable') and line.partner_id.id != first_partner) for line in move_lines]):
#                     return False
#         return True
#
#     _constraints = [
#         (_check_same_partner, 'You can only reconcile journal items with the same partner.', ['line_id', 'line_partial_ids']),
#     ]
#
# account_move_reconcile()
#
# class account_invoice_line(models.Model):
#     _inherit = 'account.invoice.line'
#
#     period_id = fields.Many2one('account.period', string='Leavy Period')
#     scheme_year_id = fields.Many2one('scheme.year','Scheme Year')
#
# account_invoice_line()

class LevyIncome(models.Model):
    _name = 'leavy.income'
    _description = "Levy Income Processing"
    _inherit = ['mail.thread', 'mail.activity.mixin']

    leavy_file_upload_ids = fields.One2many('leavy.file.upload', 'leavy_income_id', string='Levy Files Upload')
    desc = fields.Text(
        string='Description',
        default='This will import all uploaded Levy Files and will create:\n'
                '- Journal Entries for Discretionary and Admin Grants.\n'
                '- Levy History for Individual Employers.',
        readonly=True
    )

    month = fields.Char(string='Month')
    year = fields.Char(string='Year')
    # account.period is removed in Odoo 18; using scheme_year_id as the primary filter
    scheme_year_id = fields.Many2one('scheme.year', string='Scheme Year')
    seta_code = fields.Char(string='SETA Code')
    total_leavy = fields.Integer(string='Total Levy Imported')
    progress = fields.Float(string='Progress')

    # ---------------------------------------------------------
    # Helper: Generate Journal Item (Move Line) Dictionary
    # ---------------------------------------------------------
    def get_line_vals(self, name, debit, credit, account_id, partner_id, analytic_account_id, scheme_year_id):
        """
        Modernized for Odoo 18. Note: analytic_account_id is now often handled
        via analytic_distribution as a JSON field.
        """
        vals = {
            'name': name,
            'account_id': account_id,
            'partner_id': partner_id,
            'debit': debit,
            'credit': credit,
            'date_maturity': fields.Date.today() + relativedelta(months=+1),
            'scheme_year_id': scheme_year_id,
            'move_line_type': 'income',  # Custom field from previous steps
        }
        # Odoo 18 Analytic Distribution format: {str(id): 100.0}
        if analytic_account_id:
            vals['analytic_distribution'] = {str(analytic_account_id): 100.0}
        return vals

    # ---------------------------------------------------------
    # Helper: Invoice/Move Data Generation
    # ---------------------------------------------------------
    def get_move_dict(self, partner_id, name, line_ids, move_type, journal_id, scheme_year_id, due_date):
        """
        In Odoo 18, account.invoice is replaced by account.move with move_type.
        move_type options: 'entry', 'out_invoice', 'in_invoice', etc.
        """
        return {
            'partner_id': partner_id,
            'ref': name,
            'move_type': move_type,
            'journal_id': journal_id,
            'scheme_year_id': scheme_year_id,
            'invoice_date': fields.Date.today(),
            'invoice_date_due': due_date,
            'invoice_line_ids': line_ids,  # (0, 0, vals) format
        }

    def get_invoice_line_dict(self, product_id, name, account_id, qty, price_unit, scheme_year_id):
        return {
            'product_id': product_id,
            'name': name,
            'account_id': account_id,
            'quantity': qty,
            'price_unit': price_unit,
            'scheme_year_id': scheme_year_id,
        }

    # ---------------------------------------------------------
    # File Processing
    # ---------------------------------------------------------
    def get_file_read(self, file_content):
        """
        Refactored for Python 3 (standard in Odoo 18).
        Decodes binary file content and parses CSV.
        """
        decoded_file = file_content.decode('utf-8')
        io_string = io.StringIO(decoded_file)
        # Using '|' separator as per your previous SETA SDL file logic
        reader = csv.DictReader(io_string, delimiter='|')
        return reader

    def check_year_count(self, filename):
        current_year = fields.Date.today().year
        if filename and filename.upper().endswith('.SDL'):
            return filename.count(str(current_year))
        return 0
    

    ## Optimised code for levy income import
    def import_file(self):
        """
        Migrated HWSETA Levy Import for Odoo 18.
        Parses SDL files, validates DHET JVs, and creates Employer Entries.
        """
        self.ensure_one()
        partner_obj = self.env['res.partner']
        move_obj = self.env['account.move']

        # 1. File Validation
        if not self.leavy_file_upload_ids:
            raise UserError(_('Please upload Levy Files!'))

        valid_files = self.leavy_file_upload_ids.filtered(
            lambda l: l.attachment_file_id.name.endswith('.SDL') and 'Employers' not in l.attachment_file_id.name
        )
        employer_files = self.leavy_file_upload_ids.filtered(
            lambda l: l.attachment_file_id.name.endswith('.SDL') and 'Employers' in l.attachment_file_id.name
        )

        # 2. Extract Records and Check for Missing Employers
        levy_records = []
        missing_sdl = []
        total_levy_amount = 0.0
        levy_month, levy_year, file_year_str = 0, 0, ""

        for leavy_file in valid_files:
            content = base64.b64decode(leavy_file.attachment_file_id.datas).decode('utf-8')
            filename = leavy_file.attachment_file_id.name

            for line in content.split('\n'):
                if not line.strip(): continue
                row = line.split('|')
                try:
                    sdl_no = row[2]
                    amount = float(row[9])

                    # Store data for processing
                    if amount != 0.0:
                        levy_records.append({
                            'period_raw': row[0],
                            'sdl_no': sdl_no,
                            'amount': amount,
                            'filename': filename,
                            'scheme_year_name': row[12].strip(),
                            'mand': float(row[4]),
                            'desc': float(row[5]),
                            'admin': float(row[6]),
                            'penalties': float(row[7]),
                            'interest': float(row[8]),
                        })

                    # Check partner existence
                    partner = partner_obj.search([('employer_sdl_no', '=', sdl_no)], limit=1)
                    if not partner:
                        missing_sdl.append(sdl_no)

                    total_levy_amount += amount
                    levy_month = int(row[0][4:6])
                    levy_year = int(row[0][0:4])
                    file_year_str = row[12].strip()
                except (IndexError, ValueError):
                    continue

        # 3. Handle Missing Employers (Wizard Trigger)
        if missing_sdl:
            return self._handle_missing_employers(list(set(missing_sdl)), employer_files)

        # 4. Validate against DHET Lump Sum Journal Entry
        self._validate_dhet_jv(total_levy_amount, levy_month, levy_year)

        # 5. Levy Configuration and Processing
        levy_config = self.env['leavy.income.config'].search([], limit=1)
        if not levy_config:
            raise UserError(_('Please configure Levy accounts in Admin Configuration!'))

        # 6. Create Journal Entries and History
        processed_count = 0
        month_name = calendar.month_name[levy_month]

        for record in levy_records:
            employer = partner_obj.search([('employer_sdl_no', '=', record['sdl_no'])], limit=1)
            scheme_year = self.env['scheme.year'].search([('name', '=', record['scheme_year_name'])], limit=1)

            if not employer or not scheme_year:
                continue

            # Calculate Dates (Posting is usually Month + 1)
            raw_date = datetime.strptime(record['period_raw'] + "01", "%Y%m%d").date()
            posting_date = raw_date + relativedelta(months=1)

            # Create the Journal Entry using Command interface
            move_vals = self._prepare_employer_move_vals(record, employer, scheme_year, posting_date, levy_config)
            move_obj.create(move_vals)

            # Create Levy History
            employer.write({
                'leavy_history_ids': [Command.create({
                    'period': record['period_raw'],
                    'leavy_month': month_name,
                    'scheme_year_id': scheme_year.id,
                    'mand_grant_amt': record['mand'],
                    'desc_grant_amt': record['desc'],
                    'admn_grant_amt': record['admin'],
                    'penalties': record['penalties'],
                    'interest': record['interest'],
                    'total_amt': record['amount'],
                })]
            })
            processed_count += 1

        self.write({
            'month': month_name,
            'year': str(levy_year),
            'total_leavy': processed_count,
            'seta_code': levy_records[0]['sdl_no'][:1] if levy_records else ''
        })

        return {
            'name': _('Levy Import Log'),
            'type': 'ir.actions.act_window',
            'res_model': 'levy.import.log',
            'view_mode': 'form',
            'target': 'new',
            'context': {'default_message': _('Levy Files Imported Successfully!')}
        }

    def _validate_dhet_jv(self, total_levy_amount, levy_month, levy_year):
        """ Validates the Lump Sum entry from DHET exists and matches the SDL file total. """
        dhet_journal = self.env['account.journal'].search([('name', '=', 'DHET Journal')], limit=1)
        if not dhet_journal:
            raise UserError(_("DHET Journal not found. Please create it."))

        # Search for moves where analytic account matches mm/yyyy
        period_str = f"{str(levy_month).zfill(2)}/{levy_year}"
        moves = self.env['account.move'].search([
            ('journal_id', '=', dhet_journal.id),
            ('state', 'in', ['draft', 'posted'])
        ])

        valid_move = moves.filtered(lambda m: any(period_str in (l.analytic_distribution or {}) for l in m.line_ids))

        if not valid_move:
            raise UserError(_("There is no DHET JV for period %s!") % period_str)

        dhet_total = sum(valid_move[0].line_ids.mapped('debit'))
        if round(total_levy_amount, 2) != round(dhet_total, 2):
            raise UserError(_("DHET entry variance!\nFile Total: %s\nJV Total: %s") % (total_levy_amount, dhet_total))

    def _prepare_employer_move_vals(self, record, employer, scheme_year, posting_date, config):
        """ Prepares the Command list for move lines. """
        line_ids = []
        control_acc = config.control_acc.id

        # Determine Grant Allocation Logic (HWSeta Specific)
        # Assuming simplified version of your grant_config_new logic here
        # Debit Control Account, Credit Specific Grant Accounts

        # Control Line
        line_ids.append(Command.create({
            'name': f"Levy for {employer.name} - {record['period_raw']}",
            'partner_id': employer.id,
            'account_id': control_acc,
            'debit': record['amount'] if record['amount'] > 0 else 0.0,
            'credit': abs(record['amount']) if record['amount'] < 0 else 0.0,
            'scheme_year_id': scheme_year.id,
        }))

        # Grant Line (Simplified)
        line_ids.append(Command.create({
            'name': f"Mandatory Grant Allocation",
            'partner_id': employer.id,
            'account_id': config.mandatory_credit_acc.id,
            'debit': abs(record['amount']) if record['amount'] < 0 else 0.0,
            'credit': record['amount'] if record['amount'] > 0 else 0.0,
            'scheme_year_id': scheme_year.id,
        }))

        return {
            'move_type': 'entry',
            'journal_id': self.env['account.journal'].search([('code', '=', 'GEN')], limit=1).id,
            'date': posting_date,
            'ref': record['filename'],
            'line_ids': line_ids,
        }


class LevyFileUpload(models.Model):
    _name = 'leavy.file.upload'
    _description = 'Levy File Upload'

    name = fields.Char(string='Name')
    attachment_file_id = fields.Many2one('ir.attachment', string='File')
    leavy_income_id = fields.Many2one('leavy.income', string='Levy Income', ondelete='cascade')


class LevyIncomeConfig(models.Model):
    _name = 'leavy.income.config'
    _description = 'Levy Income Configuration'
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _rec_name = 'scheme_year_id'

    active = fields.Boolean(string='Active', default=True)

    # Account Fields with modern tracking
    mandatory_credit_acc = fields.Many2one('account.account', string='Mandatory Cr Acc', tracking=True)
    discretionary_credit_acc = fields.Many2one('account.account', string='Discretionary Cr Acc', tracking=True)
    admins_credit_acc = fields.Many2one('account.account', string='Admins Cr Acc', tracking=True)
    penalty_acc = fields.Many2one('account.account', string='Penalty Account', tracking=True)
    interest_acc = fields.Many2one('account.account', string='Interest Account', tracking=True)
    control_acc = fields.Many2one('account.account', string='DHET Clearing Account', tracking=True)
    expense_acc = fields.Many2one('account.account', string='Mandatory Grant Expense Account', tracking=True)
    provision_acc = fields.Many2one('account.account', string='Mandatory Grant Provision Account', tracking=True)
    disc_provision_acc = fields.Many2one('account.account', string='Discretionary Grant Provision Account',
                                         tracking=True)

    # Budgeting
    project_budget_acc = fields.Many2one('account.account', string='Project Budget Acc', tracking=True)
    total_budget = fields.Float(string='Total Budget', compute='_compute_budget_balance', store=True)
    rem_budget = fields.Float(string='Remaining Budget')
    state = fields.Selection([('new', 'New'), ('close', 'Close')], string="State", default='new')

    # WSP Configuration Dates
    wsp_extension_request_start_date = fields.Date(string='WSP Extension Request Start Date', tracking=True)
    wsp_extension_request_end_date = fields.Date(string='WSP Extension Request End Date', tracking=True)
    wsp_extension_approval_end_date = fields.Date(string='WSP Extension Approval End Date', tracking=True)
    wsp_start_date = fields.Date(string='WSP Start Date', tracking=True)
    wsp_end_date = fields.Date(string='WSP End Date', tracking=True)
    wsp_extension_date = fields.Date(string='WSP End Extension Date', tracking=True)

    # Fiscal Year handling (Fiscalyear model is usually in community/enterprise addons now)
    # Using many2one to scheme.year as primary link
    scheme_year_id = fields.Many2one('scheme.year', string='Scheme Year', tracking=True)

    # Financial Configuration
    discretionary_reserve_acc = fields.Many2one('account.account', string='Discretionary Reserve Acc', tracking=True)
    submission_end_date = fields.Date(string='Submission End Date')

    percent_one = fields.Float(string='Percent One')
    percent_two = fields.Float(string='Percent Two')
    percent_three = fields.Float(string='Percent Three')

    period_one = fields.Float(string='Period One')
    period_two = fields.Float(string='Period Two')
    period_three = fields.Float(string='Period Three')
    period_four = fields.Float(string='Period Four')

    petty_limit = fields.Float(string='Petty Cash Limit')
    fm_approval_limit = fields.Float(string='FM Approval Limit')

    grand_acc = fields.One2many('grant.account', 'grant1', string='Grant')

    _sql_constraints = [
        ('unique_active_config', 'unique(active)', 'There should be only one active record in Admin Configuration!'),
    ]

    @api.depends('project_budget_acc')
    def _compute_budget_balance(self):
        """ Replaces onchange_budget_acc with a robust compute method """
        for record in self:
            if record.project_budget_acc:
                # In Odoo 18, we fetch balance from the account's computed fields
                record.total_budget = record.project_budget_acc.current_balance
            else:
                record.total_budget = 0.0


# class AccountAsset(models.Model):
#     _inherit = 'account.asset'
#
#     product_id = fields.Many2one('product.product', string='Product')
#     asset_location = fields.Many2one('stock.location', string='Asset Location')


# class AccountAssetDepreciationLine(models.Model):
#     # In Odoo 18, depreciation lines are often managed via account.move.line,
#     # but the asset specific line remains for scheduling.
#     _inherit = 'account.asset.depreciation.line'
#
#     comment = fields.Text(string='Comment')


# --- Transient Model: Employer Import Utility ---

class EmployerInSystem(models.TransientModel):
    _name = 'employer.in.system'
    _description = 'Check Employers in System'

    employer_file = fields.Binary(string='Employer Excel File')

    def employer_not_in_system(self):
        """
        Migrated from xlrd/xlwt to openpyxl for Python 3/Odoo 18 compatibility.
        Identifies SDL numbers missing from the database and returns an Excel file.
        """
        if not self.employer_file:
            raise UserError(_("Please upload an Excel file!"))

        # Load Workbook
        file_data = base64.b64decode(self.employer_file)
        try:
            input_book = openpyxl.load_workbook(filename=io.BytesIO(file_data), data_only=True)
        except Exception as e:
            raise UserError(_("Incorrect File Format! %s") % str(e))

        output_book = openpyxl.Workbook()
        output_sheet = output_book.active
        output_sheet.title = "Employers to Import"

        partner_obj = self.env['res.partner']
        row_to_write = 1

        for sheet_name in input_book.sheetnames:
            input_sheet = input_book[sheet_name]

            for row in input_sheet.iter_rows(values_only=True):
                if not row or not row[0]:
                    continue

                sdl_no = str(row[0])
                # Search for employer
                exists = partner_obj.search_count([
                    ('is_company', '=', True),
                    ('employer_sdl_no', '=', sdl_no)
                ])

                if not exists:
                    for col_idx, value in enumerate(row, start=1):
                        output_sheet.cell(row=row_to_write, column=col_idx, value=value)
                    row_to_write += 1

        # Save to buffer
        output_buffer = io.BytesIO()
        output_book.save(output_buffer)
        output_data = base64.b64encode(output_buffer.getvalue())
        output_buffer.close()

        attachment = self.env['ir.attachment'].create({
            'name': 'Employer_to_Import.xlsx',
            'type': 'binary',
            'datas': output_data,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'new',
        }


# --- Government Levy ---

class GovernmentLevy(models.Model):
    _name = 'government.levy'
    _description = 'Government Levy Processing'

    # account.period removed; replaced with date or custom fiscal years
    date_creation = fields.Datetime(string='Date', default=fields.Datetime.now)
    scheme_year_id = fields.Many2one('scheme.year', string='Scheme Year')

    def create_invoice_gov_levy(self):
        """
        Migrated to Odoo 18 account.move (Unified Invoices/JVs).
        Calculates government levy and creates draft invoices.
        """
        # 1. Fetch Government Employers
        employers = self.env['res.partner'].search([
            ('is_company', '=', True),
            ('emp_government', '=', True)
        ])

        config = self.env['leavy.income.config'].search([], limit=1)
        if not config or not all([config.percent_one, config.percent_two, config.percent_three]):
            raise UserError(_('Please complete Government Levy configuration in Admin Settings!'))

        # 2. Identify Products
        admin_product = self.env['product.product'].search([('name', '=', 'Admin Grant')], limit=1)
        disc_product = self.env['product.product'].search([('name', '=', 'Discretionary Grant')], limit=1)

        for emp in employers:
            total_payroll = emp.total_annual_payroll
            if not total_payroll:
                continue

            # 3. Calculations
            # Equation: $Final = (((Payroll \times P1\%) \times P2\%) \times P3\%)$
            final_amount = (((total_payroll * (config.percent_one / 100)) * (config.percent_two / 100)) * (
                        config.percent_three / 100))
            admin_amount = final_amount / 3
            disc_amount = (final_amount * 2) / 3

            # 4. Create Invoice (account.move with move_type 'out_invoice')
            self.env['account.move'].create({
                'partner_id': emp.id,
                'move_type': 'out_invoice',
                'ref': 'Government Levy',
                'invoice_date': fields.Date.today(),
                'invoice_line_ids': [
                    Command.create({
                        'product_id': admin_product.id,
                        'name': f"{admin_product.name} / Government Levy",
                        'quantity': 1,
                        'price_unit': admin_amount,
                        'account_id': admin_product.property_account_income.id or admin_product.categ_id.property_account_income_categ.id,
                    }),
                    Command.create({
                        'product_id': disc_product.id,
                        'name': f"{disc_product.name} / Government Levy",
                        'quantity': 1,
                        'price_unit': disc_amount,
                        'account_id': disc_product.property_account_income.id or disc_product.categ_id.property_account_income_categ.id,
                    }),
                ]
            })
        return True