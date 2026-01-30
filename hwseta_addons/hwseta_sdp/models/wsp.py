from odoo import models, fields, api, _
from odoo.exceptions import UserError
from datetime import datetime
import calendar
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import base64
import io
import xlrd
import xlsxwriter
import re
import unicodedata
import os
import logging

_logger = logging.getLogger(__name__)


def dbg(msg):
    _logger.info(msg)


def get_occupation_and_specialization(ofo_code, env):

    if not ofo_code:
        return {
            'occupation_id': False,
            'specialization_ids': [],
        }

    occupation_id = ofo_code.occupation.id if ofo_code.occupation else False

    # Preferred ORM approach
    specialization_ids = ofo_code.specialize_ids.ids or []

    # SQL fallback (ONLY if legacy DB relation exists)
    if not specialization_ids:
        env.cr.execute("""
            SELECT specialise_id
            FROM specialisation_ofo_rel
            WHERE ofo_id = %s
        """, (ofo_code.id,))
        specialization_ids = [row[0] for row in env.cr.fetchall()]

    return {
        'occupation_id': occupation_id,
        'specialization_ids': specialization_ids,
    }


class AetSubject(models.Model):
    _name = 'aet.subject'
    _description = 'AET Subject'

    name = fields.Char("AET Subject", required=True)

    planed_adult_education_training_id = fields.Many2one(
        'planned.adult.education.training.fields',
        string="Planned Adult Education Training"
    )

    actual_adult_education_id = fields.Many2one(
        'actual.adult.education.fields',
        string="Actual Adult Education"
    )

    _sql_constraints = [
        ('aet_subject_uniq', 'unique(name)', 'AET Subject must be unique!')
    ]


class SpecializeSubject(models.Model):
    _name = 'specialize.subject'
    _description = 'Specialization Subject'

    name = fields.Char(string='Specialisation Name', required=True)

    ofo_code_id = fields.Many2one(
        'ofo.code',
        string='OFO Code'
    )

    _sql_constraints = [
        ('specialisation_uniq', 'unique(name)', 'Specialisation must be unique!')
    ]


class OccupationOfo(models.Model):
    _name = 'occupation.ofo'
    _description = 'Occupation OFO'

    name = fields.Char(string='Name', required=True)


class TrainingIntervention(models.Model):
    _name = 'training.intervention'
    _description = 'Training Intervention'

    name = fields.Char(string='Name', required=True)
    pivotal = fields.Boolean(string='Pivotal')
    active = fields.Boolean(default=True)


class OfoCode(models.Model):
    _name = 'ofo.code'
    _description = 'OFO Code'

    name = fields.Char(string='OFO Code', required=True)

    occupation = fields.Many2one(
        'occupation.ofo',
        string='Occupation',
        required=True
    )

    specialize_ids = fields.One2many(
        'specialize.subject',
        'ofo_code_id',
        string='Specialisation'
    )

    specialization_ids = fields.Many2many(
        'specialize.subject',
        'specialisation_ofo_rel',
        'ofo_id',
        'specialise_id',
        string='Specialisations'
    )

    ofo_year = fields.Char(string='OFO Year')

    _sql_constraints = [
        ('ofo_uniq', 'unique(name)', 'OFO Code must be unique!')
    ]

    def name_get(self):
        result = []
        for record in self:
            if record.ofo_year:
                display_name = f"{record.ofo_year} - {record.name}"
            else:
                display_name = record.name

            result.append((record.id, display_name))
        return result

    @api.model
    def _name_search(self, name='', args=None, operator='ilike', limit=100, name_get_uid=None):
        args = args or []

        if name:
            occupation_data = self.env['occupation.ofo'].search([('name', operator, name)])

            if occupation_data:
                ofo_data = self.search([
                    '|',
                    ('occupation', 'in', occupation_data.ids),
                    ('name', operator, name)
                ] + args, limit=limit)
            else:
                ofo_data = self.search([('name', operator, name)] + args, limit=limit)
        else:
            ofo_data = self.search(args, limit=limit)

        return ofo_data.name_get()

    def copy_old_ofo(self):
        old_records = self.env['ofo.code'].search([('ofo_year', '=', '2019')])

        for old in old_records:
            spec_list = old.specialization_ids.ids

            new_name = old.name.replace('2019-', '2021-')

            new_occupation_name = old.occupation.name.replace('2019-', '')

            occupation = self.env['occupation.ofo'].create({
                'name': '2021-' + new_occupation_name
            })

            new_vals = {
                'ofo_year': '2021',
                'name': new_name,
                'occupation': occupation.id,
                'specialization_ids': [(6, 0, spec_list)]
            }

            self.env['ofo.code'].create(new_vals)

    @api.model
    def create(self, vals):
        res = super(OfoCode, self).create(vals)

        occupation_data = res.occupation

        existing = self.search([
            ('occupation', '=', occupation_data.id),
            ('ofo_year', '=', res.ofo_year)
        ])

        if len(existing) > 1:
            raise UserError(
                _('OFO Code for the occupation %s already exists for this year!') % occupation_data.name
            )

        return res


class WspDocUpload(models.Model):
    _name = 'wsp.doc.upload'
    _description = 'WSP Document Upload'

    actual_wsp_plan_id = fields.Many2one(
        'wsp.plan',
        string='Related Actual WSP Plan'
    )

    planned_wsp_plan_id = fields.Many2one(
        'wsp.plan',
        string='Related Planned WSP Plan'
    )

    name = fields.Char(string='Name Of Doc')

    doc_file = fields.Many2one(
        'ir.attachment',
        string='Document'
    )

    @api.onchange('doc_file')
    def onchange_file(self):
        if self.doc_file:
            self.name = self.doc_file.name
        else:
            self.name = False


class IrAttachment(models.Model):
    _inherit = 'ir.attachment'
    _description = 'Attachments Extension'

    # In Odoo 18 it's better to rely on default 'name' instead of datas_fname
    # _rec_name = 'datas_fname'   # Generally not recommended anymore

    @api.model
    def default_get(self, fields_list):
        res = super().default_get(fields_list)
        return res


class WspPlan(models.Model):
    _name = 'wsp.plan'
    _description = 'WSP Plan'

    error_reporting_section = fields.Text()

    request_extension = fields.Boolean(
        string="Request Extension",
        compute="_get_request_extension_state",
        store=False
    )

    @api.depends()
    def _get_request_extension_state(self):
        for record in self:
            admin_config_data = self.env['leavy.income.config'].search([], limit=1)

            wsp_end_date = False
            wsp_ext_date = False
            wsp_ext_request_start_date = False
            wsp_ext_request_end_date = False
            wsp_ext_approval_end_date = False

            if admin_config_data:
                wsp_end_date = admin_config_data.wsp_end_date
                wsp_ext_date = admin_config_data.wsp_extension_date
                wsp_ext_request_start_date = admin_config_data.wsp_extension_request_start_date
                wsp_ext_request_end_date = admin_config_data.wsp_extension_request_end_date
                wsp_ext_approval_end_date = admin_config_data.wsp_extension_approval_end_date

            current_date = datetime.now().date()

            # Default value
            request_ext = False

            if (
                wsp_ext_request_start_date
                and (current_date >= wsp_ext_request_start_date)
                and (current_date <= wsp_ext_request_end_date)
                and not record.allow_extension
                and not record.extension_allowed
            ):
                request_ext = True

            elif (
                wsp_ext_request_end_date
                and (current_date > wsp_ext_request_end_date)
                and (current_date > wsp_ext_date)
                and not record.allow_extension
                and not record.extension_allowed
            ):
                request_ext = False

            elif (
                wsp_ext_request_end_date
                and (current_date > wsp_ext_date)
                and not record.allow_extension
                and not record.extension_allowed
            ):
                request_ext = False

            record.request_extension = request_ext

    @api.depends(
        'allow_extension',
        'extension_allowed',
        'fiscal_year',
        'scheme_year_id'
    )
    def _get_submittable_state(self):
        for record in self:
            submittable = False
            current_date = datetime.now().date()

            admin_config = self.env['leavy.income.config'].search([], limit=1)

            if not admin_config:
                record.submittable = False
                continue

            fiscal = admin_config.wsp_financial_year
            scheme = admin_config.scheme_year_id

            wsp_end_date = admin_config.wsp_end_date
            wsp_ext_date = admin_config.wsp_extension_date

            # Safety check: ensure dates exist
            if not wsp_end_date:
                record.submittable = False
                continue

            # Core Logic
            if current_date > wsp_end_date and not record.allow_extension and not record.extension_allowed:
                submittable = False

            elif current_date <= wsp_end_date:
                submittable = True

            elif (
                    current_date > wsp_end_date
                    and record.extension_allowed
                    and wsp_ext_date
                    and current_date > wsp_ext_date
            ):
                submittable = False

            elif (
                    current_date > wsp_end_date
                    and record.extension_allowed
                    and wsp_ext_date
                    and current_date <= wsp_ext_date
            ):
                submittable = True

            # Fiscal year / scheme validation
            if fiscal != record.fiscal_year or scheme != record.scheme_year_id:
                submittable = False

            record.submittable = submittable

    @api.model
    def _get_default_sdf(self):
        user_id = self.env.uid

        if user_id == 1:
            return False

        sdf_employee = self.env['hr.employee'].search([
            ('user_id', '=', user_id),
            ('is_sdf', '=', True)
        ], limit=1)

        if sdf_employee:
            return sdf_employee.id

        user = self.env['res.users'].browse(user_id)
        return user.sdf_id.id if user.sdf_id else False

    @api.model
    def _get_default_fiscalyr(self):
        admin_config = self.env['leavy.income.config'].search([], limit=1)
        return admin_config.wsp_financial_year.id if admin_config and admin_config.wsp_financial_year else False

    @api.model
    def _get_default_start_date(self):
        admin_config = self.env['leavy.income.config'].search([], limit=1)
        return admin_config.wsp_start_date if admin_config else False

    @api.model
    def _get_default_end_date(self):
        admin_config = self.env['leavy.income.config'].search([], limit=1)
        return admin_config.wsp_end_date if admin_config else False

    @api.model
    def _get_start_period(self):
        admin_config = self.env['leavy.income.config'].search([], limit=1)
        return admin_config.wsp_financial_year.date_start if admin_config and admin_config.wsp_financial_year else False

    @api.model
    def _get_end_period(self):
        admin_config = self.env['leavy.income.config'].search([], limit=1)
        return admin_config.wsp_financial_year.date_stop if admin_config and admin_config.wsp_financial_year else False

    @api.model
    def _get_scheme_year(self):
        admin_config = self.env['leavy.income.config'].search([], limit=1)
        return admin_config.scheme_year_id if admin_config else False

    @api.model
    def default_get(self, fields_list):
        res = super().default_get(fields_list)

        # Default variance lines
        inter_ids = [
            (0, 0, {'type_training': inter_data.id})
            for inter_data in self.env['training.intervention'].search([('pivotal', '=', False)])
        ]

        inter_pivotal_ids = [
            (0, 0, {'type_training': inter_data.id})
            for inter_data in self.env['training.intervention'].search([
                ('pivotal', '=', True),
                ('name', '!=', 'Other')
            ])
        ]

        res.update({
            'variance_ids': inter_ids,
            'variance_pivotal_ids': inter_pivotal_ids
        })

        admin_config = self.env['leavy.income.config'].search([], limit=1)

        if admin_config:
            wsp_end_date = admin_config.wsp_end_date
            wsp_ext_date = admin_config.wsp_extension_date
            current_date = datetime.now().date()

            request_extension = False
            submittable = False

            if wsp_end_date and current_date <= wsp_end_date:
                submittable = True

            elif wsp_end_date and wsp_ext_date and wsp_end_date < current_date <= wsp_ext_date:
                request_extension = True

            res.update({
                'request_extension': request_extension,
                'submittable': submittable
            })

        return res

    @api.depends('total_employment_profile_ids')
    def _get_total_am_tep(self):
        for record in self:
            total_african_male = sum(
                1 for line in record.total_employment_profile_ids
                if line.population_group == 'african' and line.gender == 'male'
            )
            record.total_am_tep = total_african_male

    @api.depends('total_employment_profile_ids')
    def _get_total_af_tep(self):
        for record in self:
            record.total_af_tep = sum(
                1 for line in record.total_employment_profile_ids
                if line.population_group == 'african' and line.gender == 'female'
            )

    @api.depends('total_employment_profile_ids')
    def _get_total_wm_tep(self):
        for record in self:
            record.total_wm_tep = sum(
                1 for line in record.total_employment_profile_ids
                if line.population_group == 'white' and line.gender == 'male'
            )

    @api.depends('total_employment_profile_ids')
    def _get_total_wf_tep(self):
        for record in self:
            record.total_wf_tep = sum(
                1 for line in record.total_employment_profile_ids
                if line.population_group == 'white' and line.gender == 'female'
            )

    @api.depends('total_employment_profile_ids')
    def _get_total_cm_tep(self):
        for record in self:
            record.total_cm_tep = sum(
                1 for line in record.total_employment_profile_ids
                if line.population_group == 'coloured' and line.gender == 'male'
            )

    @api.depends('total_employment_profile_ids')
    def _get_total_cf_tep(self):
        for record in self:
            record.total_cf_tep = sum(
                1 for line in record.total_employment_profile_ids
                if line.population_group == 'coloured' and line.gender == 'female'
            )

    @api.depends('total_employment_profile_ids')
    def _get_total_im_tep(self):
        for record in self:
            record.total_im_tep = sum(
                1 for line in record.total_employment_profile_ids
                if line.population_group == 'indian' and line.gender == 'male'
            )

    @api.depends('total_employment_profile_ids')
    def _get_total_if_tep(self):
        for record in self:
            record.total_if_tep = sum(
                1 for line in record.total_employment_profile_ids
                if line.population_group == 'indian' and line.gender == 'female'
            )

    # Races and Gender total calculation in Actual Training
    @api.depends('training_actual_ids')
    def _compute_actual_training_totals(self):
        for record in self:
            profiles = record.training_actual_ids

            record.total_am_at = sum(1 for l in profiles if l.population_group == 'african' and l.gender == 'male')
            record.total_af_at = sum(1 for l in profiles if l.population_group == 'african' and l.gender == 'female')

            record.total_wm_at = sum(1 for l in profiles if l.population_group == 'white' and l.gender == 'male')
            record.total_wf_at = sum(1 for l in profiles if l.population_group == 'white' and l.gender == 'female')

            record.total_cm_at = sum(1 for l in profiles if l.population_group == 'coloured' and l.gender == 'male')
            record.total_cf_at = sum(1 for l in profiles if l.population_group == 'coloured' and l.gender == 'female')

            record.total_im_at = sum(1 for l in profiles if l.population_group == 'indian' and l.gender == 'male')
            record.total_if_at = sum(1 for l in profiles if l.population_group == 'indian' and l.gender == 'female')

            record.total_pivotal_at = sum(1 for l in profiles if l.training_type == 'pivotal')
            record.total_non_pivotal_at = sum(1 for l in profiles if l.training_type == 'non-pivotal')

            record.total_employed_at = sum(1 for l in profiles if l.socio_economic_status == 'employed')
            record.total_unemployed_at = sum(1 for l in profiles if l.socio_economic_status == 'unemployed')

    # Races and Gender total calculation in Planned Training
    @api.depends('training_planned_ids')
    def _compute_planned_training_totals(self):
        for record in self:
            profiles = record.training_planned_ids

            record.total_am_pt = sum(1 for l in profiles if l.population_group == 'african' and l.gender == 'male')
            record.total_af_pt = sum(1 for l in profiles if l.population_group == 'african' and l.gender == 'female')

            record.total_wm_pt = sum(1 for l in profiles if l.population_group == 'white' and l.gender == 'male')
            record.total_wf_pt = sum(1 for l in profiles if l.population_group == 'white' and l.gender == 'female')

            record.total_cm_pt = sum(1 for l in profiles if l.population_group == 'coloured' and l.gender == 'male')
            record.total_cf_pt = sum(1 for l in profiles if l.population_group == 'coloured' and l.gender == 'female')

            record.total_im_pt = sum(1 for l in profiles if l.population_group == 'indian' and l.gender == 'male')
            record.total_if_pt = sum(1 for l in profiles if l.population_group == 'indian' and l.gender == 'female')

    # For Races and Gender total calculation in Adult Education Training
    @api.depends(
        'planned_adult_education_ids'
    )
    def _get_total_am_apt(self):
        total_african_male = 0
        for planned_trainig_data in self.planned_adult_education_ids:
            if planned_trainig_data.population_group == 'african' and planned_trainig_data.gender == 'male':
                total_african_male += 1
        self.total_am_apt = total_african_male

    @api.depends(
        'planned_adult_education_ids'
    )
    def _get_total_af_apt(self):
        total_african_female = 0
        for planned_trainig_data in self.planned_adult_education_ids:
            if planned_trainig_data.population_group == 'african' and planned_trainig_data.gender == 'female':
                total_african_female += 1
        self.total_af_apt = total_african_female

    @api.depends(
        'planned_adult_education_ids'
    )
    def _get_total_wm_apt(self):
        total_white_male = 0
        for planned_trainig_data in self.planned_adult_education_ids:
            if planned_trainig_data.population_group == 'white' and planned_trainig_data.gender == 'male':
                total_white_male += 1
        self.total_wm_apt = total_white_male

    @api.depends(
        'planned_adult_education_ids'
    )
    def _get_total_wf_apt(self):
        total_white_female = 0
        for planned_trainig_data in self.planned_adult_education_ids:
            if planned_trainig_data.population_group == 'white' and planned_trainig_data.gender == 'female':
                total_white_female += 1
        self.total_wf_apt = total_white_female

    @api.depends(
        'planned_adult_education_ids'
    )
    def _get_total_cm_apt(self):
        total_coloured_male = 0
        for planned_trainig_data in self.planned_adult_education_ids:
            if planned_trainig_data.population_group == 'coloured' and planned_trainig_data.gender == 'male':
                total_coloured_male += 1
        self.total_cm_apt = total_coloured_male

    @api.depends(
        'planned_adult_education_ids'
    )
    def _get_total_cf_apt(self):
        total_coloured_female = 0
        for planned_trainig_data in self.planned_adult_education_ids:
            if planned_trainig_data.population_group == 'coloured' and planned_trainig_data.gender == 'female':
                total_coloured_female += 1
        self.total_cf_apt = total_coloured_female

    @api.depends(
        'planned_adult_education_ids'
    )
    def _get_total_im_apt(self):
        total_indian_male = 0
        for planned_trainig_data in self.planned_adult_education_ids:
            if planned_trainig_data.population_group == 'indian' and planned_trainig_data.gender == 'male':
                total_indian_male += 1
        self.total_im_apt = total_indian_male

    @api.depends(
        'planned_adult_education_ids'
    )
    def _get_total_if_apt(self):
        total_indian_female = 0
        for planned_trainig_data in self.planned_adult_education_ids:
            if planned_trainig_data.population_group == 'indian' and planned_trainig_data.gender == 'female':
                total_indian_female += 1
        self.total_if_apt = total_indian_female

    # For Races and Gender total calculation in Vacancies Hard to Fill.
    @api.depends(
        'scarce_and_critical_skills_ids'
    )
    def _get_total_am_vhf(self):
        total_african_male = 0
        for vacancies_data in self.scarce_and_critical_skills_ids:
            if vacancies_data.population_group == 'african' and vacancies_data.gender == 'male':
                total_african_male += 1
        self.total_am_vhf = total_african_male

    @api.depends(
        'scarce_and_critical_skills_ids'
    )
    def _get_total_af_vhf(self):
        total_african_female = 0
        for vacancies_data in self.scarce_and_critical_skills_ids:
            if vacancies_data.population_group == 'african' and vacancies_data.gender == 'female':
                total_african_female += 1
        self.total_af_vhf = total_african_female

    @api.depends(
        'scarce_and_critical_skills_ids'
    )
    def _get_total_wm_vhf(self):
        total_white_male = 0
        for vacancies_data in self.scarce_and_critical_skills_ids:
            if vacancies_data.population_group == 'white' and vacancies_data.gender == 'male':
                total_white_male += 1
        self.total_wm_vhf = total_white_male

    @api.depends(
        'scarce_and_critical_skills_ids'
    )
    def _get_total_wf_vhf(self):
        total_white_female = 0
        for vacancies_data in self.scarce_and_critical_skills_ids:
            if vacancies_data.population_group == 'white' and vacancies_data.gender == 'female':
                total_white_female += 1
        self.total_wf_vhf = total_white_female

    @api.depends(
        'scarce_and_critical_skills_ids'
    )
    def _get_total_cm_vhf(self):
        total_coloured_male = 0
        for vacancies_data in self.scarce_and_critical_skills_ids:
            if vacancies_data.population_group == 'coloured' and vacancies_data.gender == 'male':
                total_coloured_male += 1
        self.total_cm_vhf = total_coloured_male

    @api.depends(
        'scarce_and_critical_skills_ids'
    )
    def _get_total_cf_vhf(self):
        total_coloured_female = 0
        for vacancies_data in self.scarce_and_critical_skills_ids:
            if vacancies_data.population_group == 'coloured' and vacancies_data.gender == 'female':
                total_coloured_female += 1
        self.total_cf_vhf = total_coloured_female

    @api.depends(
        'scarce_and_critical_skills_ids'
    )
    def _get_total_im_vhf(self):
        total_indian_male = 0
        for vacancies_data in self.scarce_and_critical_skills_ids:
            if vacancies_data.population_group == 'indian' and vacancies_data.gender == 'male':
                total_indian_male += 1
        self.total_im_vhf = total_indian_male

    @api.depends(
        'scarce_and_critical_skills_ids'
    )
    def _get_total_if_vhf(self):
        total_indian_female = 0
        for vacancies_data in self.scarce_and_critical_skills_ids:
            if vacancies_data.population_group == 'indian' and vacancies_data.gender == 'female':
                total_indian_female += 1
        self.total_if_vhf = total_indian_female

    @api.depends(
        'total_am_tep',
        'total_wm_tep',
        'total_cm_tep',
        'total_im_tep'
    )
    def _get_total_male_tep(self):
        self.total_male_tep = self.total_am_tep + \
            self.total_wm_tep + self.total_cm_tep + self.total_im_tep

    @api.depends(
        'total_af_tep',
        'total_wf_tep',
        'total_cf_tep',
        'total_if_tep'
    )
    def _get_total_female_tep(self):
        self.total_female_tep = self.total_af_tep + \
            self.total_wf_tep + self.total_cf_tep + self.total_if_tep

    @api.depends(
        'total_male_tep',
        'total_female_tep',
    )
    def _get_total_mf_tep(self):
        self.total_mf_tep = self.total_male_tep + self.total_female_tep
        self.duplicate_total_mf_tep = self.total_male_tep + self.total_female_tep

    # Total Counts====> Actual Training
    @api.depends(
        'total_am_at',
        'total_wm_at',
        'total_cm_at',
        'total_im_at'
    )
    def _get_total_male_at(self):
        self.total_male_at = self.total_am_at + \
            self.total_wm_at + self.total_cm_at + self.total_im_at

    @api.depends(
        'total_af_at',
        'total_wf_at',
        'total_cf_at',
        'total_if_at'
    )
    def _get_total_female_at(self):
        self.total_female_at = self.total_af_at + \
            self.total_wf_at + self.total_cf_at + self.total_if_at

    @api.depends(
        'total_male_at',
        'total_female_at',
    )
    def _get_total_mf_at(self):
        self.total_mf_at = self.total_male_at + self.total_female_at

    @api.depends(
        'total_am_pt',
        'total_wm_pt',
        'total_cm_pt',
        'total_im_pt'
    )
    def _get_total_male_pt(self):
        self.total_male_pt = self.total_am_pt + \
            self.total_wm_pt + self.total_cm_pt + self.total_im_pt

    @api.depends(
        'total_af_pt',
        'total_wf_pt',
        'total_cf_pt',
        'total_if_pt'
    )
    def _get_total_female_pt(self):
        self.total_female_pt = self.total_af_pt + \
            self.total_wf_pt + self.total_cf_pt + self.total_if_pt

    @api.depends(
        'total_male_pt',
        'total_female_pt',
    )
    def _get_total_male_female_pt(self):
        self.total_male_female_pt = self.total_male_pt + self.total_female_pt

    @api.depends(
        'training_planned_ids',
    )
    def _get_total_count_employed_pt(self):
        emp_count = 0
        if self.training_planned_ids:
            for rec in self.training_planned_ids:
                if rec.socio_economic_status == 'employed':
                    emp_count += 1
        self.total_count_employed_pt = emp_count

    @api.depends(
        'training_planned_ids',
    )
    def _get_total_count_unemployed_pt(self):
        unemp_count = 0
        if self.training_planned_ids:
            for rec in self.training_planned_ids:
                if rec.socio_economic_status == 'unemployed':
                    unemp_count += 1
        self.total_count_unemployed_pt = unemp_count

    def _get_pivotal_total(self):
        pivotal_total_per = 0.0
        if self.variance_pivotal_ids:
            for rec in self.variance_pivotal_ids:
                if rec.variance_percentage:
                    pivotal_total_per += rec.variance_percentage
        self.pivotal_total_per = pivotal_total_per

    def _get_non_pivotal_total(self):
        non_pivotal_total_per = 0.0
        if self.variance_ids:
            for rec in self.variance_ids:
                if rec.variance_percentage:
                    non_pivotal_total_per += rec.variance_percentage
        self.non_pivotal_total_per = non_pivotal_total_per

    @api.depends(
        'total_am_apt',
        'total_wm_apt',
        'total_cm_apt',
        'total_im_apt'
    )
    def _get_total_male_apt(self):
        self.total_male_apt = self.total_am_apt + \
            self.total_wm_apt + self.total_cm_apt + self.total_im_apt

    @api.depends(
        'total_af_apt',
        'total_wf_apt',
        'total_cf_apt',
        'total_if_apt'
    )
    def _get_total_female_apt(self):
        self.total_female_apt = self.total_af_apt + \
            self.total_wf_apt + self.total_cf_apt + self.total_if_apt

    @api.depends(
        'total_am_vhf',
        'total_wm_vhf',
        'total_cm_vhf',
        'total_im_vhf'
    )
    def _get_total_male_vhf(self):
        self.total_male_vhf = self.total_am_vhf + \
            self.total_wm_vhf + self.total_cm_vhf + self.total_im_vhf

    @api.depends(
        'total_af_vhf',
        'total_wf_vhf',
        'total_cf_vhf',
        'total_if_vhf'
    )
    def _get_total_female_vhf(self):
        self.total_female_vhf = self.total_af_vhf + \
            self.total_wf_vhf + self.total_cf_vhf + self.total_if_vhf

    @api.depends(
        'user_id',
        'state'
    )
    def _compute_current_user_forbidden(self):
        """
        Compute a field indicating whether the current user
        shouldn't be able to edit some fields.
        """
        field_status = False
        group_data = self.env['res.groups'].search([('name', '=', 'SDF')])
        group_users = [group_user.id for group_user in group_data.users]
        if self._uid in group_users and self.state == 'submitted':
            field_status = True
        else:
            field_status = False
        self.current_user_forbidden = field_status

    org_size = fields.Selection(string='Organization Size',
                                related='employer_id.organisation_size', store=True)

    query_comments = fields.Text(string="Query Comments")

    name = fields.Char(string='Name')
    sdf_id = fields.Many2one(
        'hr.employee', string='SDF', domain="[('is_sdf','=',True)]", default=_get_default_sdf)
    employer_id = fields.Many2one(
        'res.partner', string='Employer', domain="[('employer','=',True)]")
    emp_province = fields.Many2one(related="employer_id.state_id", store=True)
    wsp_details_type = fields.Selection([('actual', 'Annual Training Report'), ('planned', 'Workplace Skills Plan'),
                                         #                                          ('planned_current','Workplace Skills Plan 2015-2016')
                                         ], string="WSP ATR Details Type", default='actual')
#     wsp_details_type = fields.Selection([('actual','Annual Training Report'),('planned','Workplace Skills Plan')], string="WSP ATR Details Type", default='actual')
    # Fields for Actual and Planned Training.
    # Keeping pivotal and non pivotal in one place using following one2many
    training_actual_ids = fields.One2many(
        'actual.training.fields', 'actual_wsp_id', string='Please fill all required fields from Actual Training')
    training_actual_planned_ids = fields.One2many(
        'actual.training.fields', 'actual_planned_wsp_id', string='Please fill all required fields from Planned Training')
    ##
    actual_adult_education_ids = fields.One2many(
        'actual.adult.education.fields', 'actual_adult_wsp_id', string='Please fill all required fields from Actual Adult Education')
    actual_planned_adult_education_ids = fields.One2many(
        'actual.adult.education.fields', 'actual_planned_adult_wsp_id', string='Please fill all required fields from Planned Adult Education')
    # In future if asked for seperate pivotal and non-pivotal section then we can use below One2many fields.
#     actual_training_ids = fields.One2many('actual.training.d1.fields', 'actual_wsp_id', string='Actual Training D1')
#     actual_pivotal_training_ids = fields.One2many('actual.pivotal.training.fields', 'actual_pivotal_wsp_id', string='Actual Pivotal Training')
    total_employment_profile_ids = fields.One2many(
        'total.employment.profile.fields', 'total_employment_wsp_id', string='Please fill all required fields from Total Employment Profile')
    # Keeping pivotal and non pivotal in one place using following one2many
    training_planned_ids = fields.One2many(
        'planned.training.fields', 'planned_training_non_wsp_id', string='Please fill all required fields from Planned Training Non Pivotal')
    ##
    # In future if asked for seperate pivotal and non-pivotal section then we can use below One2many fields.
#     planned_training_non_pivotal_ids = fields.One2many('planned.training.non.pivotal.fields','planned_training_non_wsp_id', string='Planned Training Non Pivotal')
#     planned_training_pivotal_ids = fields.One2many('planned.training.pivotal.fields','planned_training_wsp_id', string='Planned Training Pivotal')
    planned_adult_education_ids = fields.One2many(
        'planned.adult.education.training.fields', 'planned_adult_education_wsp_id', string='Please fill all required fields from Planned Training Pivotal')
    scarce_and_critical_skills_ids = fields.One2many(
        'scarce.and.critical.skills.fields', 'scarce_and_critical_wsp_id', string='Please fill all required fields from Scarce and Critical Skills')
#     state = fields.Selection([('draft','Draft'),('submitted','Submitted'),('evaluated','Evaluated'),('rejected','Rejected')], string="State", default='draft')
    state = fields.Selection([('draft', 'Draft'), ('submitted', 'Submitted'), ('evaluated', 'Assessment'), ('evaluated2',
                                                                                                            'Evaluated'), ('approved', 'Accepted'), ('query', 'Query'), ('rejected', 'Rejected')], string="State", default='draft')

    fiscal_year = fields.Many2one(
        'account.fiscalyear', string='Financial Year', default=_get_default_fiscalyr)
    wsp_start_period = fields.Date(
        string='WSP Start Period', default=_get_start_period)
    scheme_year_id = fields.Many2one('scheme.year','Scheme Year', default=_get_scheme_year)
    #vacancy_hard_to_fill_bool = fields.Boolean(string='Do you have any vacancies that are hard to fill?', default=False)
    #vacancy_hard_to_fill_bool = fields.Boolean(string='Do you have any vacancies that are hard to fill?', default=False)
    vacancy_hard_to_fill_bool = fields.Selection([('yes','Yes'),('no','No')], string="Do you have vacancies hard to fill?")
    vacancy_hard_to_fill_bool_pre_date = fields.Boolean(default=False)
    wsp_end_period = fields.Date(
        string='WSP End Period', default=_get_end_period)
    start_period = fields.Date(
        string='WSP Submission Start Date', default=_get_default_start_date)
    end_period = fields.Date(
        string='WSP Submission Due Date', default=_get_default_end_date)
    ext_period = fields.Date(string='WSP Extension Date')
    ext_period1 = fields.Date(string='WSP Extension Date')
#     doc_uploaded = fields.Boolean(string="Doc Uploaded")
    submitted = fields.Boolean(string="Submitted")
    # Document Uploads
    training_proof_invoice = fields.Many2one(
        'ir.attachment', string="Proof of Training(Invoices, Receipts)", help='Upload Document')
    training_proof_attendance = fields.Many2one(
        'ir.attachment', string="Proof of Training(Attendance Registers and Certificates)", help='Upload Document')
    consultation_proof_meeting = fields.Many2one(
        'ir.attachment', string="Proof of Consultation(Minutes of meetings)", help='Upload Document')
    consultation_proof_attendance = fields.Many2one(
        'ir.attachment', string="Proof of Consultation(Attendance register)", help='Upload Document')
    authorization_page = fields.Many2one(
        'ir.attachment', string="Authorization Page", help='Upload Document')
    tax_clearance = fields.Many2one(
        'ir.attachment', string="Tax Clearance", help="Upload Tax Clearance Document")

    actual_doc_upload_ids = fields.One2many(
        'wsp.doc.upload', 'actual_wsp_plan_id', string='Other Documents', help='Upload Document')
    planned_doc_upload_ids = fields.One2many(
        'wsp.doc.upload', 'planned_wsp_plan_id', string='Other Documents', help='Upload Document')
    # WSP Impact Assessment
    point_one = fields.Char(
        string="1. On a scale of 1 to 5, please indicate the overall impact of training your staff.")
    point_two = fields.Char(string="2. Impact on the organisation.")
    point_three = fields.Char(
        string="3. To what extent are the following incentives available to staff who undergo training.")
    survey_productivity = fields.Selection([('yes', '1 - Not at All'), ('by_small', '2 - By a small amount'), ('somewhat', '3 - Somewhat'),
                                            ('to_extent', '4 - To an extent'), ('large_extent', '5 - To a very large extent')], string="1.1 Increased Productivity")
    staff_turnover = fields.Selection([('yes', '1 - Not at All'), ('by_small', '2 - By a small amount'), ('somewhat', '3 - Somewhat'),
                                       ('to_extent', '4 - To an extent'), ('large_extent', '5 - To a very large extent')], string="1.2 Reduction in staff turnover")
    increased_efficiency = fields.Selection([('yes', '1 - Not at All'), ('by_small', '2 - By a small amount'), ('somewhat', '3 - Somewhat'), (
        'to_extent', '4 - To an extent'), ('large_extent', '5 - To a very large extent')], string="1.3 Increased efficiency, resulting in financial gain")
    trained_employee = fields.Selection([('yes', '1 - Not at All'), ('by_small', '2 - By a small amount'), ('somewhat', '3 - Somewhat'), ('to_extent',
                                                                                                                                          '4 - To an extent'), ('large_extent', '5 - To a very large extent')], string="2.1 Trained employees makes fewer mistakes than those who have not been trained")
    solve_problems = fields.Selection([('yes', '1 - Not at All'), ('by_small', '2 - By a small amount'), ('somewhat', '3 - Somewhat'), ('to_extent',
                                                                                                                                        '4 - To an extent'), ('large_extent', '5 - To a very large extent')], string="2.2 Trained employees solve problems more quickly than those without training")
    training_helps_skills = fields.Selection([('yes', '1 - Not at All'), ('by_small', '2 - By a small amount'), ('somewhat', '3 - Somewhat'), ('to_extent', '4 - To an extent'),
                                              ('large_extent', '5 - To a very large extent')], string="2.3 Training helps an organisation keep abreast of the latest skills development in the field")
    supervision = fields.Selection([('yes', '1 - Not at All'), ('by_small', '2 - By a small amount'), ('somewhat', '3 - Somewhat'),
                                    ('to_extent', '4 - To an extent'), ('large_extent', '5 - To a very large extent')], string="2.4 There is a decreased need for supervision")
    training_helps_product = fields.Selection([('yes', '1 - Not at All'), ('by_small', '2 - By a small amount'), ('somewhat', '3 - Somewhat'), ('to_extent', '4 - To an extent'),
                                               ('large_extent', '5 - To a very large extent')], string="2.5 Training helps an organisation keep abreast of the latest product development in the field")
    career_path = fields.Selection([('yes', '1 - Not at All'), ('by_small', '2 - By a small amount'), ('somewhat', '3 - Somewhat'), ('to_extent',
                                                                                                                                     '4 - To an extent'), ('large_extent', '5 - To a very large extent')], string="2.6 Training gives employees a clearer sense of career path")
    keep_motivated = fields.Selection([('yes', '1 - Not at All'), ('by_small', '2 - By a small amount'), ('somewhat', '3 - Somewhat'),
                                       ('to_extent', '4 - To an extent'), ('large_extent', '5 - To a very large extent')], string="2.7 Training keeps employees motivated")
    promotion = fields.Selection([('yes', '1 - Not at All'), ('by_small', '2 - By a small amount'), ('somewhat', '3 - Somewhat'),
                                  ('to_extent', '4 - To an extent'), ('large_extent', '5 - To a very large extent')], string="3.1 Improved promotion opportunities")
    remuneration = fields.Selection([('yes', '1 - Not at All'), ('by_small', '2 - By a small amount'), ('somewhat', '3 - Somewhat'),
                                     ('to_extent', '4 - To an extent'), ('large_extent', '5 - To a very large extent')], string="3.2 Improved remuneration prospects")
    training_reported = fields.Text(
        string="4. Indicate how training as reported in the ATR has affected the organisation")
    percentage_payroll = fields.Char(
        string='5. Percentage payroll spent on training')
    wsp_submission_date = fields.Date(string='Date Submitted')
    variance_ids = fields.One2many(
        'variance.calculate', 'wsp_plan_id', string='Variance')
    variance_pivotal_ids = fields.One2many(
        'variance.calculate.pivotal', 'wsp_plan_id', string='Variance')
    wsp_submission_status_ids = fields.One2many(
        'wsp.submission.status', 'wsp_plan_id', string='Status of WSP Submission')
#     wsp_submission_status_wspofficer_ids = fields.One2many('wsp.submission.status.wspofficer','wsp_plan_id', string='Status of WSP Submission WSP Officer')
#     wsp_submission_status_wspmanager_ids = fields.One2many('wsp.submission.status.wspmanager','wsp_plan_id', string='Status of WSP Submission WSP Manager')
#

    user_id = fields.Many2one(
        'res.users', string='Current User', default=lambda self: self.env.user)
    current_user_forbidden = fields.Boolean(
        compute="_compute_current_user_forbidden")

    # Fields for Races and Gender total calculations(TEP).
    total_am_tep = fields.Integer(
        string='Total African Male', compute='_get_total_am_tep')
    total_af_tep = fields.Integer(
        string='Total African Female', compute='_get_total_af_tep')
    total_wm_tep = fields.Integer(
        string='Total White Male', compute='_get_total_wm_tep')
    total_wf_tep = fields.Integer(
        string='Total White Female', compute='_get_total_wf_tep')
    total_cm_tep = fields.Integer(
        string='Total Coloured Male', compute='_get_total_cm_tep')
    total_cf_tep = fields.Integer(
        string='Total Coloured Female', compute='_get_total_cf_tep')
    total_im_tep = fields.Integer(
        string='Total Indian Male', compute='_get_total_im_tep')
    total_if_tep = fields.Integer(
        string='Total Indian Female', compute='_get_total_if_tep')

    total_male_tep = fields.Integer(
        string='Total Male', compute='_get_total_male_tep')
    total_female_tep = fields.Integer(
        string='Total Female', compute='_get_total_female_tep')
    total_mf_tep = fields.Integer(
        string='Total Employees', compute='_get_total_mf_tep')
    duplicate_total_mf_tep = fields.Integer(
        string='Total Employees', compute='_get_total_mf_tep')

    # Fields for Races and Gender total calculations(Actual Training).
    total_am_at = fields.Integer(
        string='Total African Male', compute='_compute_actual_training_totals')
    total_af_at = fields.Integer(
        string='Total African Female', compute='_compute_actual_training_totals')
    total_wm_at = fields.Integer(
        string='Total White Male', compute='_compute_actual_training_totals')
    total_wf_at = fields.Integer(
        string='Total White Female', compute='_compute_actual_training_totals')
    total_cm_at = fields.Integer(
        string='Total Coloured Male', compute='_compute_actual_training_totals')
    total_cf_at = fields.Integer(
        string='Total Coloured Female', compute='_compute_actual_training_totals')
    total_im_at = fields.Integer(
        string='Total Indian Male', compute='_compute_actual_training_totals')
    total_if_at = fields.Integer(
        string='Total Indian Female', compute='_compute_actual_training_totals')
    total_pivotal_at = fields.Integer(
        string='Total Pivotal', compute='_compute_actual_training_totals')
    total_non_pivotal_at = fields.Integer(
        string='Total Non-Pivotal', compute='_compute_actual_training_totals')
    total_employed_at = fields.Integer(
        string='Total Employed', compute='_compute_actual_training_totals')
    total_unemployed_at = fields.Integer(
        string='Total Unemployed', compute='_compute_actual_training_totals')

    total_male_at = fields.Integer(
        string='Total Male', compute='_compute_actual_training_totals')
    total_female_at = fields.Integer(
        string='Total Female', compute='_compute_actual_training_totals')
    total_mf_at = fields.Integer(
        string='Total Employees', compute='_compute_actual_training_totals')
    # Fields for Races and Gender total calculations(Planned Training).
    total_am_pt = fields.Integer(
        string='Total African Male', compute='_compute_planned_training_totals')
    total_af_pt = fields.Integer(
        string='Total African Female', compute='_compute_planned_training_totals')
    total_wm_pt = fields.Integer(
        string='Total White Male', compute='_compute_planned_training_totals')
    total_wf_pt = fields.Integer(
        string='Total White Female', compute='_compute_planned_training_totals')
    total_cm_pt = fields.Integer(
        string='Total Coloured Male', compute='_compute_planned_training_totals')
    total_cf_pt = fields.Integer(
        string='Total Coloured Female', compute='_compute_planned_training_totals')
    total_im_pt = fields.Integer(
        string='Total Indian Male', compute='_compute_planned_training_totals')
    total_if_pt = fields.Integer(
        string='Total Indian Female', compute='_compute_planned_training_totals')

    total_male_pt = fields.Integer(
        string='Total Male', compute='_compute_planned_training_totals')
    total_female_pt = fields.Integer(
        string='Total Female', compute='_compute_planned_training_totals')
    total_male_female_pt = fields.Integer(
        string='NO. OF PLANNED BENEFICIARIES', compute='_compute_planned_training_totals')
    total_count_employed_pt = fields.Integer(
        string='Total Employed', compute='_compute_planned_training_totals')
    total_count_unemployed_pt = fields.Integer(
        string='Total Unemployed', compute='_compute_planned_training_totals')
    # Fields for Races and Gender total calculations(Adult Planned Training).
    total_am_apt = fields.Integer(
        string='Total African Male', compute='_get_total_am_apt')
    total_af_apt = fields.Integer(
        string='Total African Female', compute='_get_total_af_apt')
    total_wm_apt = fields.Integer(
        string='Total White Male', compute='_get_total_wm_apt')
    total_wf_apt = fields.Integer(
        string='Total White Female', compute='_get_total_wf_apt')
    total_cm_apt = fields.Integer(
        string='Total Coloured Male', compute='_get_total_cm_apt')
    total_cf_apt = fields.Integer(
        string='Total Coloured Female', compute='_get_total_cf_apt')
    total_im_apt = fields.Integer(
        string='Total Indian Male', compute='_get_total_im_apt')
    total_if_apt = fields.Integer(
        string='Total Indian Female', compute='_get_total_if_apt')

    total_male_apt = fields.Integer(
        string='Total Male', compute='_get_total_male_apt')
    total_female_apt = fields.Integer(
        string='Total Female', compute='_get_total_female_apt')

    # Fields for Races and Gender total calculations(Vacancies Hard to Fill).
    total_am_vhf = fields.Integer(
        string='Total African Male', compute='_get_total_am_vhf')
    total_af_vhf = fields.Integer(
        string='Total African Female', compute='_get_total_af_vhf')
    total_wm_vhf = fields.Integer(
        string='Total White Male', compute='_get_total_wm_vhf')
    total_wf_vhf = fields.Integer(
        string='Total White Female', compute='_get_total_wf_vhf')
    total_cm_vhf = fields.Integer(
        string='Total Coloured Male', compute='_get_total_cm_vhf')
    total_cf_vhf = fields.Integer(
        string='Total Coloured Female', compute='_get_total_cf_vhf')
    total_im_vhf = fields.Integer(
        string='Total Indian Male', compute='_get_total_im_vhf')
    total_if_vhf = fields.Integer(
        string='Total Indian Female', compute='_get_total_if_vhf')

    total_male_vhf = fields.Integer(
        string='Total Male', compute='_get_total_male_vhf')
    total_female_vhf = fields.Integer(
        string='Total Female', compute='_get_total_female_vhf')

    extension_date = fields.Date(string='WSP Extension Date')
    # Fields for extension
    request_extension = fields.Boolean(string='Request Extension', compute='_get_request_extension_state')
    submittable = fields.Boolean(string='Can be submitted', compute='_get_submittable_state')
    #request_extension = fields.Boolean(string='Request Extension')
    allow_extension = fields.Boolean(string='Allow Extension')
    show_extension_date = fields.Boolean(string='Show Extension')
    extension_allowed = fields.Boolean(string='Extension Allowed')
    request_extension_date = fields.Date(string='Request Extension Date')
    approve_extension_date = fields.Date(string='Approve Extension Date')
#     from_extension = fields.Boolean(string='From Extension')
#     extension_over = fields.Boolean(string='Extension Over')
#     extension_allowed = fields.Boolean(string='Extension Allowed')
    sdl_no = fields.Char(string='SDL No.', size=10)
    approved_by = fields.Many2one('res.users', string='Evaluated By')

    assessment_rec_or_not = fields.Boolean(string="assessment_rec_or_not")
    evaluated_rec_or_not = fields.Boolean(string="evaluated_rec_or_not")
    comments = fields.Text(string='Comments')
    query_status = fields.Selection([('while_assessment', 'While Assessment'), (
        'while_evaluation', 'While Evaluation')], string='Query Status')

#     WSP evaluation form
#     Actual Training Report 2015 2016
    actual_training_selection = fields.Selection(
        [('yes', 'Yes'), ('no', 'No')], string="Actual Training")
    actual_training_comment = fields.Char(string='Comments / Reasons')
    actual_education_and_training_selection = fields.Selection(
        [('yes', 'Yes'), ('no', 'No')], string="Adult Education and Training")
    actual_education_and_training_comment = fields.Char(
        string='Comments / Reasons')
    wsp_impact_assessment_selection = fields.Selection(
        [('yes', 'Yes'), ('no', 'No')], string="WSP  Impact Assessment")
    wsp_impact_assessment_comment = fields.Char(string='Comments / Reasons')
    variance_report_selection = fields.Selection(
        [('yes', 'Yes'), ('no', 'No')], string="Variance Report ( has more than 60% of training been done )")
    variance_report_comment = fields.Char(string='Comments / Reasons')
    proof_of_training_invoice_selection = fields.Selection(
        [('yes', 'Yes'), ('no', 'No')], string="* Proof of Training ( Invoices/ Receipts)")
    proof_of_training_invoice_comment = fields.Char(
        string='Comments / Reasons')
    proof_of_training_attendance_selection = fields.Selection(
        [('yes', 'Yes'), ('no', 'No')], string="* Proof of Training ( Attendance Registers) ")
    proof_of_training_attendance_comment = fields.Char(
        string='Comments / Reasons')

#     Workplace Skills Plan 2016 2017
    tep_selection = fields.Selection(
        [('yes', 'Yes'), ('no', 'No')], string="Total Employment Profile")
    tep_comment = fields.Char(string='Comments / Reasons')
    planned_training_selection = fields.Selection(
        [('yes', 'Yes'), ('no', 'No')], string="Actual Training")
    planned_training_comment = fields.Char(string='Comments / Reasons')
    actual_education_and_training_selection = fields.Selection(
        [('yes', 'Yes'), ('no', 'No')], string="Adult Education and Training")
    actual_education_and_training_comment = fields.Char(
        string='Comments / Reasons')
    vacancies_selection = fields.Selection(
        [('yes', 'Yes'), ('no', 'No')], string="Vacancies Hard to Fill ")
    vacancies_comment = fields.Char(string='Comments / Reasons')
    minute_of_training_selection = fields.Selection(
        [('yes', 'Yes'), ('no', 'No')], string="* Minutes of Training Meeting ")
    minute_of_training_comment = fields.Char(string='Comments / Reasons')
    attendance_register_selection = fields.Selection(
        [('yes', 'Yes'), ('no', 'No')], string="* Attendance Register Of Training Meeting")
    attendance_register_comment = fields.Char(string='Comments / Reasons')
    authorisation_page_selection = fields.Selection(
        [('yes', 'Yes'), ('no', 'No')], string="* Authorisation Page ")
    authorisation_page_comment = fields.Char(string='Comments / Reasons')
    save_some_record = fields.Boolean("Partially Save", default=False)
    rejection_date = fields.Date("WSP Rejection Date")
    approval_date = fields.Date("WSP Approval Date")
    pivotal_total_per = fields.Float('Pivotal', compute='_get_pivotal_total')
    non_pivotal_total_per = fields.Float(
        'Non-Pivotal', compute='_get_non_pivotal_total')

    is_tep_loaded = fields.Boolean("TEP Loaded", default=False)
    is_tep_to_planned_loaded = fields.Boolean(
        "TEP to Planned Loaded", default=False)
    is_prev_wsp_loaded = fields.Boolean("Previous WSP Loaded", default=False)
    extension_doc_statement = fields.Many2one(
        'ir.attachment', string="Extenstion of WSP Reason Statement/Letter", help='Upload Document')
    extension_state = fields.Selection([('requested', 'Requested'), ('resubmit', 'Re-submit'), ('verified', 'Verified'), ('approved', 'Approved'), ('rejected', 'Rejected'),('historic', 'Historic Ext'),], string="State" )
    extension_reason_comment = fields.Text(string='Comments / Reasons')
    extension_verify_reason_comment = fields.Text(string='Comments / Reasons')
    extension_approve_reason_comment = fields.Text(string='Comments / Reasons')
    extension_reject_reason_comment = fields.Text(string='Comments / Reasons')
    extension_approved_by = fields.Many2one('res.users', string='Extension Approved By')
    extension_verified_by = fields.Many2one('res.users', string='Extension Verifired By')
    extension_rejected_by = fields.Many2one('res.users', string='Extension Rejected By')
    allow_resubmit_extension = fields.Boolean("Re-Submit Extension", default=False)
    wsp_extension_status_ids = fields.One2many(
        'wsp.extension.status', 'wsp_plan_id', string='Status of WSP Extension')
    wsp_extension_ids = fields.One2many(
        'wsp.extension', 'wsp_plan_id', string='WSP Extension')


    # Fields for query >>>>>>>>>>>>>START>>>>>>>>>>>>
    query_date = fields.Date(string='WSP Query Date')   
    # request_query = fields.Boolean(string='Request Query', compute='_get_request_query_state')
    allow_query = fields.Boolean(string='Allow Query')
    show_query_date = fields.Boolean(string='Show Query')
    query_allowed = fields.Boolean(string='Query Allowed')
    request_query_date = fields.Date(string='Request Query Date')
    approve_query_date = fields.Date(string='Approve Query Date')
    query_docs_statement = fields.Many2one(
        'ir.attachment', string="Extenstion of WSP Reason Statement/Letter", help='Upload Document') 
    query_doc_statement = fields.Many2one('ir.attachment', string="Extenstion of WSP Reason Statement/Letter", help='Upload Document')
    query_state = fields.Selection([('requested', 'Requested'), ('resubmit', 'Re-submit'), ('verified', 'Verified'), ('approved', 'Approved'), ('rejected', 'Rejected'),('historic', 'Historic Ext'),], string="State" )
    query_reason_comment = fields.Text(string='Comments / Reasons')
    query_verify_reason_comment = fields.Text(string='Comments / Reasons')
    query_approve_reason_comment = fields.Text(string='Comments / Reasons')
    query_reject_reason_comment = fields.Text(string='Comments / Reasons')
    query_approved_by = fields.Many2one('res.users', string='Query Approved By')
    query_verified_by = fields.Many2one('res.users', string='Query Verifired By')
    query_rejected_by = fields.Many2one('res.users', string='Query Rejected By')
    allow_resubmit_query = fields.Boolean("Re-Submit Query", default=False)
    wsp_query_status_ids = fields.One2many('wsp.query.status', 'wsp_plan_id', string='Status of WSP Query')
    wsp_query_id = fields.Integer(string='WSP Query ID')
    before_query_state = fields.Selection([('draft', 'Draft'), ('submitted', 'Submitted'), ('evaluated', 'Assessment'), ('evaluated2',
                                                                                                            'Evaluated'), ('approved', 'Accepted'), ('query', 'Query'), ('rejected', 'Rejected')], string="State")

    wsp_query_ids = fields.One2many('wsp.query', 'wsp_plan_id', string='WSP Query')
    wsp_query_docs_ids = fields.One2many('wsp.query.docs', 'wsp_plan_id', string='WSP Query')

    def action_verify_query(self):
        # wdb.set_trace()    
        return {
            'name': _('WSP Query Verificacation'),
            'res_model': 'wsp.query.verify',
            'view_mode': 'form',
            'context': {
                'active_model': 'wsp.plan',
                'active_ids': self.ids,
            },
            'target': 'new',
            'type': 'ir.actions.act_window',
        }

    def action_approve_query(self):
        # wdb.set_trace()    
        return {
            'name': _('WSP Query Approval'),
            'res_model': 'wsp.query.approval',
            'view_mode': 'form',
            'context': {
                'active_model': 'wsp.plan',
                'active_ids': self.ids,
            },
            'target': 'new',
            'type': 'ir.actions.act_window',
        }
            

    def action_reject_query(self):
        # wdb.set_trace()    
        return {
            'name': _('WSP Query Rejection'),
            'res_model': 'wsp.query.reject',
            'view_mode': 'form',
            'context': {
                'active_model': 'wsp.plan',
                'active_ids': self.ids,
            },
            'target': 'new',
            'type': 'ir.actions.act_window',
        }
    
    def action_request_query(self):
        # wdb.set_trace()           
        return {
            'name': _('Request Query'),
            'res_model': 'wsp.query.request',
            'view_mode': 'form',
            'context': {
                'active_model': 'wsp.plan',
                'active_ids': self.ids,
            },
            'target': 'new',
            'type': 'ir.actions.act_window',
        }


    def action_request_for_query(self):
        ir_model_data_obj = self.env['ir.model.data']
        bronwen_id = ir_model_data_obj.get_object_reference('hwseta_sdp', 'email_template_wsp_request_for_query_bronwen')
        if bronwen_id:
            self.pool['email.template'].send_mail(self.env.cr, self.env.uid, bronwen_id[1], self.id,force_send=True,context=self.env.context)
        luyanda_id = ir_model_data_obj.get_object_reference('hwseta_sdp', 'email_template_wsp_request_for_query_luyanda')
        if luyanda_id:
            self.pool['email.template'].send_mail(self.env.cr, self.env.uid, luyanda_id[1], self.id,force_send=True,context=self.env.context)
        self.write({'allow_query': True, 'request_query':
                    False,'query_state': 'requested', 'request_query_date': datetime.now().date()})
        res_status = {
            'evaluator': self._uid,
            'date_evaluation': datetime.now(),
            'status': 'requested',
            'state': 'requested',
            'wsp_state': self.state,
            'comments': self.query_reason_comment,
            'wsp_plan_id': self.id,
        }
        self.write({'wsp_query_status_ids': [(0, 0, res_status)], 'comments': ''})

        res_ext = {
            'query_state': 'requested',
            'before_query_state': self.before_query_state,
            'query_doc_statement' : self.query_doc_statement,
            'query_reason_comment': self.query_reason_comment,
            'wsp_plan_id': self.id,
        }
        self.write({'wsp_query_ids': [(0, 0, res_ext)], 'comments': ''})
        return True

    def action_resubmit_for_query(self):
        ir_model_data_obj = self.env['ir.model.data']
        bronwen_id = ir_model_data_obj.get_object_reference('hwseta_sdp', 'email_template_wsp_request_for_query_bronwen')
        if bronwen_id:
            self.pool['email.template'].send_mail(self.env.cr, self.env.uid, bronwen_id[1], self.id,force_send=True,context=self.env.context)
        luyanda_id = ir_model_data_obj.get_object_reference('hwseta_sdp', 'email_template_wsp_request_for_query_luyanda')
        if luyanda_id:
            self.pool['email.template'].send_mail(self.env.cr, self.env.uid, luyanda_id[1], self.id,force_send=True,context=self.env.context)
        self.write({'allow_query': True, 'request_query':
                    False,'query_state': 'requested'})
        res_status = {
            'evaluator': self._uid,
            'date_evaluation': datetime.now(),
            'status': 'requested',
            'state': 'requested',
            'wsp_query_id': self.wsp_query_id,
            'wsp_state': self.state,
            'comments': self.query_reason_comment,
            'wsp_plan_id': self.id,
        }
        self.write({'wsp_query_status_ids': [(0, 0, res_status)], 'comments': ''})        
        return True

    def action_verify_for_query(self):
        self.write({'query_state': 'verified'})
        res = {
            'evaluator': self._uid,
            'date_evaluation': datetime.now(),
            'status': 'verified',
            'state': 'verified',
            'wsp_state': self.state,
            'wsp_query_id': self.wsp_query_id,
            'comments': self.query_verify_reason_comment,
            'wsp_plan_id': self.id,
        }
        self.write({'wsp_query_status_ids': [(0, 0, res)], 'comments': ''})      
        return True

    def action_update_for_query(self):
        res_doc = {
                    'query_docs_statement': self.query_docs_statement.id,
                    'comments': self.query_comments,
                    'wsp_plan_id': self.id,
                    'wsp_query_id': self.wsp_query_id,
                }
        self.write({'wsp_query_docs_ids': [(0, 0, res_doc)], 'comments': ''})
        return True

        
    def action_reject_for_query(self):
        self.write({'query_state': 'rejected'})
        res = {
            'evaluator': self._uid,
            'date_evaluation': datetime.now(),
            'status': 'rejected',
            'state': 'rejected',
            'wsp_query_id': self.wsp_query_id,
            'wsp_state': self.state,
            'comments': self.query_reject_reason_comment,
            'wsp_plan_id': self.id,
        }
        self.write({'wsp_query_status_ids': [(0, 0, res)], 'comments': ''})
        return True    
        
    def action_reject_resubmit_for_query(self):
        self.write({'query_state': 'resubmit'})
        res = {
            'evaluator': self._uid,
            'date_evaluation': datetime.now(),
            'status': 'resubmit',
            'state': 'resubmit',
            'wsp_state': self.state,
            'comments': self.query_reject_reason_comment,
            'wsp_query_id': self.wsp_query_id,
            'wsp_plan_id': self.id,
        }
        self.write({'wsp_query_status_ids': [(0, 0, res)], 'comments': ''})
        return True    

    def action_approve_for_query(self):
        # admin_config_data = self.env['leavy.income.config'].search([])
        wsp_query_date = datetime.now().date()
        # wsp_query_date = datetime.strptime( '%Y-%m-%d').date()
        ir_model_data_obj = self.env['ir.model.data']
        bronwen_id = ir_model_data_obj.get_object_reference('hwseta_sdp', 'email_template_wsp_approval_for_query_bronwen')
        if bronwen_id:
            self.pool['email.template'].send_mail(self.env.cr, self.env.uid, bronwen_id[1], self.id,force_send=True,context=self.env.context)
    
        luyanda_id = ir_model_data_obj.get_object_reference('hwseta_sdp', 'email_template_wsp_approval_for_query_luyanda')
        if luyanda_id:
            self.pool['email.template'].send_mail(self.env.cr, self.env.uid, luyanda_id[1], self.id,force_send=True,context=self.env.context)
        self.write({'allow_query': False, 'show_query_date': True, 'query_date':
                    wsp_query_date, 'query_state': 'approved', 'approve_query_date': datetime.now().date(),'query_allowed':True})
        sdl_id = ir_model_data_obj.get_object_reference('hwseta_sdp', 'email_template_wsp_approval_for_query_employer')       
        if sdl_id:
            self.pool['email.template'].send_mail(self.env.cr, self.env.uid, sdl_id[1], self.id,force_send=True,context=self.env.context)

        res = {
            'evaluator': self._uid,
            'date_evaluation': datetime.now(),
            'status': 'approved',
            'state': 'approved',
            'wsp_state': self.state,
            'wsp_query_id': self.wsp_query_id,
            'comments': self.query_approve_reason_comment,
            'wsp_plan_id': self.id,
        }
        self.write({'wsp_query_status_ids': [(0, 0, res)], 'comments': ''})
        return True
    
         # Fields for query >>>>>>>>>>>>>END>>>>>>>>>>>>    
    

    def onchange_employer_id(self, employer_id):
        if not employer_id:
            return False
        else:
            result = {}
            employer_data = self.env['res.partner'].browse(employer_id)
            result = {'value': {
                'sdl_no': employer_data.employer_sdl_no,
            }}
        return result

    @api.onchange('save_some_record')
    def onchange_partial_save(self):
        for rec in self:
            save = rec.save_some_record

            if rec.training_planned_ids:
                rec.training_planned_ids.write({
                    'save_planned_record': save
                })

            if rec.training_actual_ids:
                rec.training_actual_ids.write({
                    'save_actual_planned_record': save
                })

            if rec.total_employment_profile_ids:
                rec.total_employment_profile_ids.write({
                    'save_total_employed_record': save
                })

            if rec.planned_adult_education_ids:
                rec.planned_adult_education_ids.write({
                    'save_planned_adult_record': save
                })

            if rec.actual_adult_education_ids:
                rec.actual_adult_education_ids.write({
                    'save_actual_adult_record': save
                })

            if rec.scarce_and_critical_skills_ids:
                rec.scarce_and_critical_skills_ids.write({
                    'save_variance_record': save
                })


    @api.model
    def request_for_extension(self):
        ''' scheduler method for processing existing draft wsp according to request for extension.
            It will enable Request Extension button if WSP Submission Due date exceeding current date.
        '''
        admin_config_data = self.env['leavy.income.config'].search([])
        wsp_extension_date = datetime.strptime(
            admin_config_data.wsp_extension_date, '%Y-%m-%d').date()
        config_wsp_end_date = datetime.strptime(
            admin_config_data.wsp_end_date, '%Y-%m-%d').date()
        for wsp_data in self.env['wsp.plan'].search([('state', '=', 'draft'), ('fiscal_year', '=', admin_config_data.wsp_financial_year.id)]):

            #             if wsp_data.end_period and not wsp_data.approve_extension_date:
            #                 wsp_end_date = datetime.strptime(wsp_data.end_period, '%Y-%m-%d').date()
            current_date = datetime.now().date()
            if (current_date > config_wsp_end_date):
                wsp_data = wsp_data.with_context(load_previous=True)
                vals = {'request_extension': False,
                        'allow_extension': False,
                        'show_extension_date': True,
                        'extension_date': wsp_extension_date,
                        'request_extension_date': current_date,
                        'approve_extension_date': current_date,
                        }
                wsp_data.write(vals)
        return True

    def send_query_email(self):
        '''
        This function changes the state to query and send mail to SDF of that WSP.
        '''
        if not self.query_comments:
            raise Warning ("Please provide query comments")

        ir_model_data_obj = self.env['ir.model.data']
        mail_template_id = ir_model_data_obj.get_object_reference(
            'hwseta_sdp', 'email_template_query_process_wsp_plan')
        if mail_template_id:
            self.pool['email.template'].send_mail(self.env.cr, self.env.uid, mail_template_id[
                                                  1], self.id, force_send=True, context=self.env.context)
        query_status = 'while_assessment'
        if self.state == 'evaluated':
            query_status = 'while_evaluation'
        self.write({'before_query_state': self.state })
        self.write({'state': 'query', 'query_status': query_status})
        wsp_submission_data = self.env['wsp.submission.track'].search(
            [('employer_id', '=', self.employer_id.id), ('fiscal_year', '=', self.fiscal_year.id), ('name', '=', self.name)])
        if wsp_submission_data:
            user = self.env['res.users'].browse(self._uid)
            wsp_submission_data.write({'status': 'query', })
        return True

    def before_query_stage(self):
        next_state = self.before_query_state
        if self.wsp_submission_date and next_state == 'approved':
            next_state = 'approved'
        elif self.wsp_submission_date and next_state == 'rejected':
            next_state = 'rejected'
        elif self.wsp_submission_date and next_state == 'evaluated':
            next_state = 'evaluated'
        elif self.wsp_submission_date and next_state == 'evaluated2':
            next_state = 'evaluated2'
        elif self.wsp_submission_date and next_state == 'submitted':
            next_state = 'submitted' 
        elif self.wsp_submission_date and next_state == 'query':
            next_state = 'submitted'
        else:
            next_state = 'draft'
        
        self.write({'state': next_state})
        return True


    def go_previous_stage(self):
        '''  This method allows again writes the previous state.i.e., Whether query got while assessment (WSP Administrator) or while evaluation ( WSP Officer ) '''
        next_state = 'submitted'
        if self.query_status == 'while_evaluation':
            next_state = 'evaluated'
        self.write({'state': next_state})
        return True

    def go_previous_stage_for_every(self):
        '''  This method allows again writes the previous state.'''
        wsp_submission_data = self.env['wsp.submission.track'].search(
            [('employer_id', '=', self.employer_id.id), ('fiscal_year', '=', self.fiscal_year.id), ('name', '=', self.name)])
        if self.state == 'evaluated':
            self.write({'state': 'submitted'})
            if wsp_submission_data:
                wsp_submission_data.write({'status': 'submitted'})
        if self.state == 'evaluated2':
            self.write({'state': 'evaluated'})
            if wsp_submission_data:
                wsp_submission_data.write({'status': 'evaluated'})
        if self.state == 'rejected':
            self.write({'state': 'evalauted2'})
            if wsp_submission_data:
                wsp_submission_data.write({'status': 'evaluated2'})
        return True
    # Applying domain on Employer field on change of SDF.

    def onchange_sdf_id(self, sdf_id):
        sdf_data = self.env['hr.employee'].browse(sdf_id)
        employer_ids = []
        approved_employer_ids = []
        for employer_data in sdf_data.employer_ids:
            if employer_data.employer_id and employer_data.status == 'approved':
                employer_ids.append(employer_data.employer_id.id)
        employer_sdf = self.env['sdf.tracking'].search(
            [('status', '=', 'approved'), ('sdf_id', '=', sdf_id)])
        employer_id_list = [
            employer1.partner_id.id for employer1 in employer_sdf]
        for approved_employer in employer_ids:
            if approved_employer in employer_id_list:
                approved_employer_ids.append(approved_employer)
        return {'domain': {'employer_id': [('id', 'in', approved_employer_ids)]}}

    def action_verify_extension(self):
           
        return {
            'name': _('WSP Extension Verificacation'),
            'res_model': 'wsp.extension.verify',
            'view_mode': 'form',
            'context': {
                'active_model': 'wsp.plan',
                'active_ids': self.ids,
            },
            'target': 'new',
            'type': 'ir.actions.act_window',
        }

    def action_approve_extension(self):
           
        return {
            'name': _('WSP Extension Approval'),
            'res_model': 'wsp.extension.approval',
            'view_mode': 'form',
            'context': {
                'active_model': 'wsp.plan',
                'active_ids': self.ids,
            },
            'target': 'new',
            'type': 'ir.actions.act_window',
        }
            

    def action_reject_extension(self):
           
        return {
            'name': _('WSP Extension Rejection'),
            'res_model': 'wsp.extension.reject',
            'view_mode': 'form',
            'context': {
                'active_model': 'wsp.plan',
                'active_ids': self.ids,
            },
            'target': 'new',
            'type': 'ir.actions.act_window',
        }
    
    def action_request_extension(self):
                  
        return {
            'name': _('Request Extension'),
            'res_model': 'wsp.extension.request',
            'view_mode': 'form',
            'context': {
                'active_model': 'wsp.plan',
                'active_ids': self.ids,
            },
            'target': 'new',
            'type': 'ir.actions.act_window',
        }


    def action_request_for_extension(self):       
        ir_model_data_obj = self.env['ir.model.data']
        bronwen_id = ir_model_data_obj.get_object_reference('hwseta_sdp', 'email_template_wsp_request_for_extension_bronwen')
        if bronwen_id:
            self.pool['email.template'].send_mail(self.env.cr, self.env.uid, bronwen_id[1], self.id,force_send=True,context=self.env.context)
        luyanda_id = ir_model_data_obj.get_object_reference('hwseta_sdp', 'email_template_wsp_request_for_extension_luyanda')
        if luyanda_id:
            self.pool['email.template'].send_mail(self.env.cr, self.env.uid, luyanda_id[1], self.id,force_send=True,context=self.env.context)
        self.write({'allow_extension': True, 'request_extension':
                    False,'extension_state': 'requested', 'request_extension_date': datetime.now().date()})
        res_status = {
            'evaluator': self._uid,
            'date_evaluation': datetime.now(),
            'status': 'requested',
            'state': 'requested',
            'wsp_state': self.state,
            'comments': self.extension_reason_comment,
            'wsp_plan_id': self.id,
        }
        self.write({'wsp_extension_status_ids': [(0, 0, res_status)], 'comments': ''})

        res_ext = {
            'extension_state': 'requested',
            'extension_doc_statement' : self.extension_doc_statement,
            'extension_reason_comment': self.extension_reason_comment,
            'wsp_plan_id': self.id,
        }
        self.write({'wsp_extension_ids': [(0, 0, res_ext)], 'comments': ''})
        return True

    def action_resubmit_for_extension(self):       
        ir_model_data_obj = self.env['ir.model.data']
        bronwen_id = ir_model_data_obj.get_object_reference('hwseta_sdp', 'email_template_wsp_request_for_extension_bronwen')
        if bronwen_id:
            self.pool['email.template'].send_mail(self.env.cr, self.env.uid, bronwen_id[1], self.id,force_send=True,context=self.env.context)
        luyanda_id = ir_model_data_obj.get_object_reference('hwseta_sdp', 'email_template_wsp_request_for_extension_luyanda')
        if luyanda_id:
            self.pool['email.template'].send_mail(self.env.cr, self.env.uid, luyanda_id[1], self.id,force_send=True,context=self.env.context)
        self.write({'allow_extension': True, 'request_extension':
                    False,'extension_state': 'requested'})
        res_status = {
            'evaluator': self._uid,
            'date_evaluation': datetime.now(),
            'status': 'requested',
            'state': 'requested',
            'wsp_state': self.state,
            'comments': self.extension_reason_comment,
            'wsp_plan_id': self.id,
        }
        self.write({'wsp_extension_status_ids': [(0, 0, res_status)], 'comments': ''})        
        return True

    def action_verify_for_extension(self):
        self.write({'extension_state': 'verified'})
        res = {
            'evaluator': self._uid,
            'date_evaluation': datetime.now(),
            'status': 'verified',
            'state': 'verified',
            'wsp_state': self.state,
            'comments': self.extension_verify_reason_comment,
            'wsp_plan_id': self.id,
        }
        self.write({'wsp_extension_status_ids': [(0, 0, res)], 'comments': ''})
        return True

        
    def action_reject_for_extension(self):
        self.write({'extension_state': 'rejected'})
        res = {
            'evaluator': self._uid,
            'date_evaluation': datetime.now(),
            'status': 'rejected',
            'state': 'rejected',
            'wsp_state': self.state,
            'comments': self.extension_reject_reason_comment,
            'wsp_plan_id': self.id,
        }
        self.write({'wsp_extension_status_ids': [(0, 0, res)], 'comments': ''})
        return True    
        
    def action_reject_resubmit_for_extension(self):
        self.write({'extension_state': 'resubmit'})
        res = {
            'evaluator': self._uid,
            'date_evaluation': datetime.now(),
            'status': 'resubmit',
            'state': 'resubmit',
            'wsp_state': self.state,
            'comments': self.extension_reject_reason_comment,
            'wsp_plan_id': self.id,
        }
        self.write({'wsp_extension_status_ids': [(0, 0, res)], 'comments': ''})
        return True    

    def action_approve_for_extension(self):        
        admin_config_data = self.env['leavy.income.config'].search([])
        wsp_extension_date = datetime.strptime(
            admin_config_data.wsp_extension_date, '%Y-%m-%d').date()
        ir_model_data_obj = self.env['ir.model.data']
        bronwen_id = ir_model_data_obj.get_object_reference('hwseta_sdp', 'email_template_wsp_approval_for_extension_bronwen')
        if bronwen_id:
            self.pool['email.template'].send_mail(self.env.cr, self.env.uid, bronwen_id[1], self.id,force_send=True,context=self.env.context)
    
        luyanda_id = ir_model_data_obj.get_object_reference('hwseta_sdp', 'email_template_wsp_approval_for_extension_luyanda')
        if luyanda_id:
            self.pool['email.template'].send_mail(self.env.cr, self.env.uid, luyanda_id[1], self.id,force_send=True,context=self.env.context)
        self.write({'allow_extension': False, 'show_extension_date': True, 'extension_date':
                    wsp_extension_date, 'extension_state': 'approved', 'approve_extension_date': datetime.now().date(),'extension_allowed':True})
        sdl_id = ir_model_data_obj.get_object_reference('hwseta_sdp', 'email_template_wsp_approval_for_extension_employer')       
        if sdl_id:
            self.pool['email.template'].send_mail(self.env.cr, self.env.uid, sdl_id[1], self.id,force_send=True,context=self.env.context)

        res = {
            'evaluator': self._uid,
            'date_evaluation': datetime.now(),
            'status': 'approved',
            'state': 'approved',
            'wsp_state': self.state,
            'comments': self.extension_approve_reason_comment,
            'wsp_plan_id': self.id,
        }
        self.write({'wsp_extension_status_ids': [(0, 0, res)], 'comments': ''})
        return True
    
    def action_new_submit_wsp(self):
        dbg("hit action_new_submit_wsp")
        import subprocess
        validator = subprocess.check_output(["python3", "/var/log/odoo/wsp_all.py", str(self.id)])
        if "Fine" not in validator:
            raise Warning(_(validator))
        else:
            # check if there are any False line validators in the iterables
            if check_line_validated_iterables():
                raise Warning(_("Successful - or would be if not for this raise"))

        # Sending Email Notification.
        ir_model_data_obj = self.env['ir.model.data']
        mail_template_id = ir_model_data_obj.get_object_reference(
            'hwseta_sdp', 'email_template_wsp_submission')
        email_template_obj = self.env['email.template']
        temp_obj = email_template_obj.browse(mail_template_id[1])
        if mail_template_id:
            self.pool['email.template'].send_mail(self.env.cr, self.env.uid, mail_template_id[
                                                  1], self.id, force_send=True, context=self.env.context)
        employer_mail_template_id = ir_model_data_obj.get_object_reference(
            'hwseta_sdp', 'email_template_employer_wsp_submission')
        if employer_mail_template_id:
            self.pool['email.template'].send_mail(self.env.cr, self.env.uid, employer_mail_template_id[
                                                  1], self.id, force_send=True, context=self.env.context)

        self.write({'state': 'submitted', 'submitted': True,
                    'wsp_submission_date': datetime.today().date()})
        return True

    def action_submit_wsp(self):
        ''' Previously made only compulsion of ATR if organisation is old , if it is new then only WSP is compulsory.
            But now made both the old and new organisation mandatory.
        '''
        # Attachment generation code
#         att_obj = self.env['ir.attachment']
#         attachment_ids = []
#         report_xml_atr = self.env['report']._get_report_from_name('hwseta_sdp.atr_report')
#         atr, format = self.pool['report'].get_pdf(self._cr, self._uid, [report_xml_atr.id], 'hwseta_sdp.atr_report', data={'wsp_id':self.id}), 'pdf'
#         atr_attachment_data = {
#                 'name': "ATR.pdf",
#                 'datas_fname':'ATR.pdf', # your object File Name
#                 'datas':base64.encodestring(atr),  # your object Data
#                  'type': 'binary',
#                 'res_name': 'wsp_plan',
#                 'res_id': self.id,
#             }
#         attachment_ids.append(att_obj.create(atr_attachment_data)[0].id)
#         report_xml_atr = self.env['report']._get_report_from_name('hwseta_sdp.wsp_report')
#         wsp, format = self.pool['report'].get_pdf(self._cr, self._uid, [report_xml_atr.id], 'hwseta_sdp.wsp_report', data={'wsp_id':self.id}), 'pdf'
#         wsp_attachment_data = {
#                 'name': "WSP.pdf",
#                 'datas_fname':'WSP.pdf', # your object File Name
#                 'datas':base64.encodestring(wsp),  # your object Data
#                  'type': 'binary',
#                 'res_name': 'wsp_plan',
#                 'res_id': self.id,
#             }
#
#         attachment_ids.append(att_obj.create(wsp_attachment_data)[0].id)

        # Validating Actual ATR Detail Type Sections.
#         if self.wsp_details_type == 'actual' :
#             ## Validating Proof of Training Documents
#             if (not self.training_proof_invoice or not self.training_proof_attendance) :
#                 raise Warning(_('Please upload Proof of Training Documents!'))
#         self = self.with_context({'validating' : True})
#         if not self.authorization_page:
#             raise Warning(_('Please provide authorization page in document upload'))
        admin_config_data = self.env['leavy.income.config'].search([])
        if admin_config_data:
            if len(admin_config_data) > 1:
                scheme_year_id_checker = admin_config_data[0].scheme_year_id
                fiscal_year_checker = admin_config_data[0].wsp_financial_year
                wsp_extension_date_checker = admin_config_data[0].wsp_extension_date
                wsp_extension_date_checker = datetime.strptime(wsp_extension_date_checker, '%Y-%m-%d').date()
            else:
                scheme_year_id_checker = admin_config_data.scheme_year_id
                fiscal_year_checker = admin_config_data.wsp_financial_year
                wsp_extension_date_checker = admin_config_data.wsp_extension_date
                wsp_extension_date_checker = datetime.strptime(wsp_extension_date_checker, '%Y-%m-%d').date()

        current_date_checker = datetime.now().date()
        current_fiscal_year_checker = self.fiscal_year
        current_scheme_year_id_checker = self.scheme_year_id

        if current_scheme_year_id_checker != scheme_year_id_checker or current_fiscal_year_checker != fiscal_year_checker:
            raise Warning(_('Your transaction cannot be processed, because the year is not the same as the current period.'))
        if current_date_checker > wsp_extension_date_checker:
            raise Warning(_('Your transaction cannot be processed, because the date is past the extension date.'))
        if self.save_some_record:
            raise Warning(_('Please uncheck partial save!'))
        if self.duplicate_total_mf_tep <= 0:
            raise Warning(_('you have no computed value for total number of employees. please check again or contact administrator'))
        # Avoiding SDF to submit the WSP if he already submit the WSP for the
        # same year and for the same employer.
        fiscal_year_data = self.fiscal_year
        wsp_submission_data = self.env['wsp.submission.track'].search(
            [('employer_id', '=', self.employer_id.id), ('fiscal_year', '=', fiscal_year_data.id), ('status', 'in', ['submitted', 'evaluated'])])
        if wsp_submission_data:
            raise Warning(_('WSP for year %s is already submitted / evaluated for employer %s!') %
                          (fiscal_year_data.name, self.employer_id.name))
        # Checking whether last year wsp for employer is submitted (old employer) or not (new employer) and if it is old employer then validate doc, actual traiing and variance comment.
#         fiscal_year = datetime.strptime(fiscal_year_data.date_start, '%Y-%m-%d').year
#         last_fy = None
#         for fiscal_data in self.env['account.fiscalyear'].search([]) :
#             if datetime.strptime(fiscal_data.date_start, '%Y-%m-%d').year == (fiscal_year-1) :
#                 last_fy = fiscal_data.id
#         wsp_submission_data_last = self.env['wsp.submission.track'].search([('employer_id','=',self.employer_id.id),('fiscal_year','=',last_fy)])
#         if wsp_submission_data_last :
        if not self.training_actual_ids:
            raise Warning(_('Please enter Actual training data!'))
        if not self.training_proof_invoice:
            raise Warning(
                _('Please upload Proof of Training (Invoices/Receipts) Document!'))
        if not self.training_proof_attendance:
            raise Warning(
                _('Please upload Proof of Training (Attendance Registers and Certificates) Document!'))
        for variance_pivotal_data in self.variance_pivotal_ids:
            if (variance_pivotal_data.wsp_planned or variance_pivotal_data.total_cost_planned or variance_pivotal_data.atr_actual or variance_pivotal_data.total_cost_actual) and not variance_pivotal_data.comments:
                raise Warning(_('Please enter comments for %s in Variance!' % (
                    variance_pivotal_data.type_training and variance_pivotal_data.type_training.name)))
        for variance_data in self.variance_ids:
            if (variance_data.wsp_planned or variance_data.total_cost_planned or variance_data.atr_actual or variance_data.total_cost_actual) and not variance_data.comments:
                raise Warning(_('Please enter comments for %s in Variance!' % (
                    variance_data.type_training and variance_data.type_training.name)))
#             if not variance_data.comments :
#                 raise Warning(_('Please enter all comments in variance!'))
        # Validating whether WSP Impact Assessment filled or not.
        if not (self.survey_productivity and self.staff_turnover and self.increased_efficiency
                and self.trained_employee and self.solve_problems and self.training_helps_skills
                and self.supervision and self.training_helps_product and self.career_path and
                self.keep_motivated and self.promotion and self.remuneration and self.training_reported
                and self.percentage_payroll):
            raise Warning(_('Please complete WSP Impact Assessment Entries!'))
#             if not self.total_employment_profile_ids :
#                 raise Warning(_('Please enter Total Employment Profile data!'))

        # Validating whether employee present in TEP.
        for planned_training_data in self.training_planned_ids:
            if planned_training_data.employee_id:
                count = 0
                emp_status = planned_training_data.socio_economic_status == 'employed'
                emp_ids = [
                    tep_data.employee_id for tep_data in self.total_employment_profile_ids]
                if emp_ids:
                    if planned_training_data.socio_economic_status == 'unemployed':
                        pass
                    else:
                        if planned_training_data.employee_id in emp_ids and planned_training_data.socio_economic_status == 'employed':
                            count += 1
                        if count == 0:
                            raise Warning(_('Employee with ID %s not present TEP. Invalid Employee!') % (
                                planned_training_data.employee_id))
                else:
                    if emp_status:
                        raise Warning(
                            _('Please enter Total Employment Profile data!'))

#             if not self.training_planned_ids :
#                 raise Warning(_('Please enter Planned Training data!'))
#         ## For Brand New Organisation.
#         else:
        if not self.total_employment_profile_ids:
            raise Warning(_('Please enter Total Employment Profile data!'))
        if not self.training_planned_ids:
            raise Warning(_('Please enter Planned Training data!'))
        user = self.env['res.users'].browse(self._uid)
        year_env = self.env['scheme.year']
        good_year = year_env.search([('code', '=', fiscal_year_data.scheme_year_id.name)])
        self.employer_id.write({'wsp_submission_ids': [(0, 0, {'name': self.name,'scheme_year_id':good_year.id ,'org_year':good_year.id ,'fiscal_year': fiscal_year_data.id, 'status': 'submitted', 'employer_id':
                                                               self.employer_id.id, 'last_user_evaluated_updated': user.name, 'wsp_date_submitted': datetime.today().date(), 'date_created': self.create_date})]})
        wsp_submission_data = self.env['wsp.submission.track'].search(
            [('employer_id', '=', self.employer_id.id), ('fiscal_year', '=', self.fiscal_year.id),
             ('name', '=', self.name)])
        wsp_sub_dat = wsp_submission_data.read()
        if wsp_sub_dat:
            # raise Warning(wsp_sub_dat)
            for field in wsp_sub_dat[0].keys():
                if type(wsp_sub_dat[0].get(field)) == tuple:
                    wsp_sub_dat[0].update({field: wsp_sub_dat[0].get(field)[0]})
            for child in self.employer_id.child_employer_ids:
                child_partner = child.employer_id
                dbg(child_partner)
                wsp_sub_dat[0].update({'employer_id': child_partner.id})
                wsp_submission_data = self.env['wsp.submission.track'].create(wsp_sub_dat[0])
                dbg(wsp_submission_data)
        ##
        # Validating Planned WSP Detail Type Sections.
        if self.wsp_details_type == 'planned':
            employer_count = len(self.total_employment_profile_ids)
            if employer_count >= 50:
                # Validating Planned Training and Planned AET sections filled
                # or not
                if (not self.consultation_proof_meeting or not self.consultation_proof_attendance):
                    raise Warning(
                        _('Please Upload Proof of Consultation Documents'))
            # Authorisation Page is mandatory for all employer.
            if not self.authorization_page:
                raise Warning(_('Please Upload Document Authorisation Page!'))
            self.employer_id.write({'employees_count': employer_count})
        ##
        # If organisation is old then will require both ATR and Planned. Brand new organisation do not require ATR
#         emp_existing_wsp = self.env['wsp.plan'].search([('employer_id','=',self.employer_id.id),('state','in',['submitted','evaluated'])])
        # For old organisation.
#         if emp_existing_wsp:
#             if not self.training_actual_ids :
#                 raise Warning(_('Please enter Actual Training data!'))
#             if not self.actual_adult_education_ids :
#                 raise Warning(_('Please enter Adult Education and Training data!'))
#             if not self.total_employment_profile_ids :
#                 raise Warning(_('Please enter Total Employment Profile data!'))
#             if not self.training_planned_ids :
#                 raise Warning(_('Please enter Planned Training data!'))
#             if not self.planned_adult_education_ids :
#                 raise Warning(_('Please enter Planned Adult Education data!'))
#             if not self.scarce_and_critical_skills_ids :
#                 raise Warning(_('Please enter Vacancies Hard to fill data!'))
            # Validating whether WSP Impact Assessment filled or not.
#             if not (self.survey_productivity or self.staff_turnover or self.increased_efficiency
#                     or self.trained_employee or self.solve_problems or self.training_helps_skills
#                     or self.supervision or self.training_helps_product or self.career_path or
#                     self.keep_motivated or self.promotion or self.remuneration or self.training_reported
#                     or self.percentage_payroll) :
#                 raise Warning(_('Please complete WSP Impact Assessment Entries!'))
#             if not self.planned_adult_education_ids :
#                 raise Warning(_('Please enter Planned Adult Education data!'))
        # Validation : Non Pivotal Training Applies only to employees
        for actual_data in self.training_actual_ids:
            if actual_data.training_type == 'non-pivotal' and actual_data.socio_economic_status == 'unemployed':
                raise Warning(_('Actual Training : Non Pivotal training should applies to only employed %s !') % (
                    actual_data.code and 'for OFO Code ' + actual_data.code.name or ''))
        for planned_data in self.training_planned_ids:
            if planned_data.training_type == 'non-pivotal' and planned_data.socio_economic_status == 'unemployed':
                raise Warning(_('Planned Training : Non Pivotal training should applies to only employed %s !') % (
                    planned_data.name and planned_data.surname and 'for ' + planned_data.name + ' ' + planned_data.surname or planned_data.code and 'for OFO Code ' + planned_data.code.name or ''))

        ##
        # Sending Email Notification.
        ir_model_data_obj = self.env['ir.model.data']
        mail_template_id = ir_model_data_obj.get_object_reference(
            'hwseta_sdp', 'email_template_wsp_submission')
        email_template_obj = self.env['email.template']
        temp_obj = email_template_obj.browse(mail_template_id[1])
        if mail_template_id:
            #             temp_obj.write({'attachment_ids':[(6, 0, attachment_ids)]})
            self.pool['email.template'].send_mail(self.env.cr, self.env.uid, mail_template_id[
                                                  1], self.id, force_send=True, context=self.env.context)
        employer_mail_template_id = ir_model_data_obj.get_object_reference(
            'hwseta_sdp', 'email_template_employer_wsp_submission')
        if employer_mail_template_id:
            self.pool['email.template'].send_mail(self.env.cr, self.env.uid, employer_mail_template_id[
                                                  1], self.id, force_send=True, context=self.env.context)

        self.write({'state': 'submitted', 'submitted': True,
                    'wsp_submission_date': datetime.today().date()})
#         wsp_track_submission=self.env['wsp.submission.track'].search([('employer_id','=',self.employer_id.id),('fiscal_year','=',fiscal_year_data.id),('status','in',['draft'])])
#         user=self.env['res.users'].browse(self._uid)
#         wsp_track_submission.write({'status':'submitted','last_user_evaluated_updated':user.name,'wsp_date_submitted':datetime.today().date()})
        return True

# This method is added to resend submitted WSP's email notification to SDF and Organisation's
#     
#     def action_resend_mail(self):
#         ir_model_data_obj = self.env['ir.model.data']
#         mail_template_id = ir_model_data_obj.get_object_reference('hwseta_sdp', 'email_template_wsp_submission')
#         email_template_obj = self.env['email.template']
#         temp_obj = email_template_obj.browse(mail_template_id[1])
#         employer_mail_template_id = ir_model_data_obj.get_object_reference('hwseta_sdp', 'email_template_employer_wsp_submission')
#         wsp_records = self.search([('submitted','=',True),('fiscal_year','=', 30)])
#         for record in wsp_records:
#             if mail_template_id:
#                 self.pool['email.template'].send_mail(self.env.cr, self.env.uid, mail_template_id[1], record.id, force_send=True,context=self.env.context)
#             if employer_mail_template_id:
#                 self.pool['email.template'].send_mail(self.env.cr, self.env.uid, employer_mail_template_id[1], record.id, force_send=True,context=self.env.context)

# This method is added to resend submitted WSP's email notification to SDF and Organisation's
#     
#     def action_resend_mail(self):
#         ir_model_data_obj = self.env['ir.model.data']
#         mail_template_id = ir_model_data_obj.get_object_reference('hwseta_sdp', 'email_template_wsp_submission')
#         email_template_obj = self.env['email.template']
#         temp_obj = email_template_obj.browse(mail_template_id[1])
#         employer_mail_template_id = ir_model_data_obj.get_object_reference('hwseta_sdp', 'email_template_employer_wsp_submission')
#         wsp_records = self.search([('submitted','=',True),('fiscal_year','=', 30)])
#         for record in wsp_records:
#             if mail_template_id:
#                 self.pool['email.template'].send_mail(self.env.cr, self.env.uid, mail_template_id[1], record.id, force_send=True,context=self.env.context)
#             if employer_mail_template_id:
#                 self.pool['email.template'].send_mail(self.env.cr, self.env.uid, employer_mail_template_id[1], record.id, force_send=True,context=self.env.context)

    def validate_wsp(self):
        context = self._context
        actual_csv_data = ''
        planned_csv_data = ''
        planned_current_csv_data = ''

        if self.save_some_record:
            raise Warning(_('Please uncheck partial save!'))
        if self.duplicate_total_mf_tep <= 0:
            raise Warning(
                _('you have no computed value for total number of employees. please check again or contact administrator'))
        # Avoiding SDF to submit the WSP if he already submit the WSP for the
        # same year and for the same employer.
        fiscal_year_data = self.fiscal_year
        wsp_submission_data = self.env['wsp.submission.track'].search(
            [('employer_id', '=', self.employer_id.id), ('fiscal_year', '=', fiscal_year_data.id),
             ('status', 'in', ['submitted', 'evaluated'])])
        if wsp_submission_data:
            raise Warning(_('WSP for year %s is already submitted / evaluated for employer %s!') %
                          (fiscal_year_data.name, self.employer_id.name))

        if not self.training_actual_ids:
            raise Warning(_('Please enter Actual training data!'))
        if not self.training_proof_invoice:
            raise Warning(
                _('Please upload Proof of Training (Invoices/Receipts) Document!'))
        if not self.training_proof_attendance:
            raise Warning(
                _('Please upload Proof of Training (Attendance Registers and Certificates) Document!'))
        for variance_pivotal_data in self.variance_pivotal_ids:
            if (
                    variance_pivotal_data.wsp_planned or variance_pivotal_data.total_cost_planned or variance_pivotal_data.atr_actual or variance_pivotal_data.total_cost_actual) and not variance_pivotal_data.comments:
                raise Warning(_('Please enter comments for %s in Variance!' % (
                        variance_pivotal_data.type_training and variance_pivotal_data.type_training.name)))
        for variance_data in self.variance_ids:
            if (
                    variance_data.wsp_planned or variance_data.total_cost_planned or variance_data.atr_actual or variance_data.total_cost_actual) and not variance_data.comments:
                raise Warning(_('Please enter comments for %s in Variance!' % (
                        variance_data.type_training and variance_data.type_training.name)))

        if not (self.survey_productivity and self.staff_turnover and self.increased_efficiency
                and self.trained_employee and self.solve_problems and self.training_helps_skills
                and self.supervision and self.training_helps_product and self.career_path and
                self.keep_motivated and self.promotion and self.remuneration and self.training_reported
                and self.percentage_payroll):
            raise Warning(_('Please complete WSP Impact Assessment Entries!'))
        #             if not self.total_employment_profile_ids :
        #                 raise Warning(_('Please enter Total Employment Profile data!'))

        # Validating whether employee present in TEP.
        for planned_training_data in self.training_planned_ids:
            if planned_training_data.employee_id:
                count = 0
                emp_status = planned_training_data.socio_economic_status == 'employed'
                emp_ids = [
                    tep_data.employee_id for tep_data in self.total_employment_profile_ids]
                if emp_ids:
                    if planned_training_data.socio_economic_status == 'unemployed':
                        pass
                    else:
                        if planned_training_data.employee_id in emp_ids and planned_training_data.socio_economic_status == 'employed':
                            count += 1
                        if count == 0:
                            raise Warning(_('Employee with ID %s not present TEP. Invalid Employee!') % (
                                planned_training_data.employee_id))
                else:
                    if emp_status:
                        raise Warning(
                            _('Please enter Total Employment Profile data!'))


        if not self.total_employment_profile_ids:
            raise Warning(_('Please enter Total Employment Profile data!'))
        if not self.training_planned_ids:
            raise Warning(_('Please enter Planned Training data!'))
        user = self.env['res.users'].browse(self._uid)
        year_env = self.env['scheme.year']
        good_year = year_env.search([('code', '=', fiscal_year_data.scheme_year_id.name)])

        wsp_submission_data = self.env['wsp.submission.track'].search(
            [('employer_id', '=', self.employer_id.id), ('fiscal_year', '=', self.fiscal_year.id),
             ('name', '=', self.name)])
        wsp_sub_dat = wsp_submission_data.read()
        # if wsp_sub_dat:
        #     # raise Warning(wsp_sub_dat)
        #     for field in wsp_sub_dat[0].keys():
        #         if type(wsp_sub_dat[0].get(field)) == tuple:
        #             wsp_sub_dat[0].update({field: wsp_sub_dat[0].get(field)[0]})
        #     for child in self.employer_id.child_employer_ids:
        #         child_partner = child.employer_id
        #         dbg(child_partner)
        #         wsp_sub_dat[0].update({'employer_id': child_partner.id})
        #         wsp_submission_data = self.env['wsp.submission.track'].create(wsp_sub_dat[0])
        #         dbg(wsp_submission_data)

        ##
        # Validating Planned WSP Detail Type Sections.
        if self.wsp_details_type == 'planned':
            employer_count = len(self.total_employment_profile_ids)
            if employer_count >= 50:
                # Validating Planned Training and Planned AET sections filled
                # or not
                if (not self.consultation_proof_meeting or not self.consultation_proof_attendance):
                    raise Warning(
                        _('Please Upload Proof of Consultation Documents'))
            # Authorisation Page is mandatory for all employer.
            if not self.authorization_page:
                raise Warning(_('Please Upload Document Authorisation Page!'))
            # self.employer_id.write({'employees_count': employer_count})


        ##
        # If organisation is old then will require both ATR and Planned. Brand new organisation do not require ATR
        #         emp_existing_wsp = self.env['wsp.plan'].search([('employer_id','=',self.employer_id.id),('state','in',['submitted','evaluated'])])
        # For old organisation.
        #         if emp_existing_wsp:
        #             if not self.training_actual_ids :
        #                 raise Warning(_('Please enter Actual Training data!'))
        #             if not self.actual_adult_education_ids :
        #                 raise Warning(_('Please enter Adult Education and Training data!'))
        #             if not self.total_employment_profile_ids :
        #                 raise Warning(_('Please enter Total Employment Profile data!'))
        #             if not self.training_planned_ids :
        #                 raise Warning(_('Please enter Planned Training data!'))
        #             if not self.planned_adult_education_ids :
        #                 raise Warning(_('Please enter Planned Adult Education data!'))
        #             if not self.scarce_and_critical_skills_ids :
        #                 raise Warning(_('Please enter Vacancies Hard to fill data!'))
        # Validating whether WSP Impact Assessment filled or not.
        #             if not (self.survey_productivity or self.staff_turnover or self.increased_efficiency
        #                     or self.trained_employee or self.solve_problems or self.training_helps_skills
        #                     or self.supervision or self.training_helps_product or self.career_path or
        #                     self.keep_motivated or self.promotion or self.remuneration or self.training_reported
        #                     or self.percentage_payroll) :
        #                 raise Warning(_('Please complete WSP Impact Assessment Entries!'))
        #             if not self.planned_adult_education_ids :
        #                 raise Warning(_('Please enter Planned Adult Education data!'))
        # Validation : Non Pivotal Training Applies only to employees
        
        for actual_data in self.training_actual_ids:
            if actual_data.training_type == 'non-pivotal' and actual_data.socio_economic_status == 'unemployed':
                raise Warning(_('Actual Training : Non Pivotal training should applies to only employed %s !') % (
                        actual_data.code and 'for OFO Code ' + actual_data.code.name or ''))
        for planned_data in self.training_planned_ids:
            if planned_data.training_type == 'non-pivotal' and planned_data.socio_economic_status == 'unemployed':
                raise Warning(_('Planned Training : Non Pivotal training should applies to only employed %s !') % (
                        planned_data.name and planned_data.surname and 'for ' + planned_data.name + ' ' + planned_data.surname or planned_data.code and 'for OFO Code ' + planned_data.code.name or ''))


    
    def action_evaluate_wsp(self):
        #         self.write({'state':'evaluated'})
        self.write({'before_query_state': 'evaluated','state': 'evaluated', 'assessment_rec_or_not': True})

        # Sending Email Notification.
#         ir_model_data_obj = self.env['ir.model.data']
#         mail_template_id = ir_model_data_obj.get_object_reference('hwseta_sdp', 'email_template_wsp_acceptance')
#         if mail_template_id:
#             self.pool['email.template'].send_mail(self.env.cr, self.env.uid, mail_template_id[1], self.id, force_send=True,context=self.env.context)
        # Putting entry under Employer master.
        wsp_submission_data = self.env['wsp.submission.track'].search(
            [('employer_id', '=', self.employer_id.id), ('fiscal_year', '=', self.fiscal_year.id), ('name', '=', self.name)])
        wsp_sub_dat = False
        if wsp_submission_data:
            wsp_submission_data.write({'status': 'evaluated'})
            wsp_sub_dat = wsp_submission_data.read()
            # todo: write the chile status loop
        if wsp_sub_dat:
            # raise Warning(wsp_sub_dat)
            for field in wsp_sub_dat[0].keys():
                if type(wsp_sub_dat[0].get(field)) == tuple:
                    wsp_sub_dat[0].update({field: wsp_sub_dat[0].get(field)[0]})
            for child in self.employer_id.child_employer_ids:
                child_partner = child.employer_id
                child_wsp_submission_data = self.env['wsp.submission.track'].search(
                    [('employer_id', '=', child_partner.id), ('fiscal_year', '=', self.fiscal_year.id),
                     ('name', '=', self.name)])

                dbg(child_partner)
                wsp_sub_dat[0].update({'employer_id':child_partner.id})
                child_wsp_submission_data.write(wsp_sub_dat[0])
                dbg(wsp_submission_data)
        self = self.with_context({'from_wsp_evaluated': True})
        self.employer_id.write({'wsp_submitted': True})
        self.write({'approved_by': self._uid})

        res = {
            'evaluator': self._uid,
            'date_evaluation': datetime.now(),
            'status': 'recommended',
            'comments': self.comments,
            'state': 'evaluated',
            'wsp_plan_id': self.id,
        }
        self.write(
            {'wsp_submission_status_ids': [(0, 0, res)], 'comments': ''})

        return True

    
    def action_evaluate_not_recommended_wsp(self):
        self.write({'before_Query_state': 'evaluated','state': 'evaluated', 'assessment_rec_or_not': False})
        # Sending Email Notification.
#         ir_model_data_obj = self.env['ir.model.data']
#         mail_template_id = ir_model_data_obj.get_object_reference('hwseta_sdp', 'email_template_wsp_acceptance')
#         if mail_template_id:
#             self.pool['email.template'].send_mail(self.env.cr, self.env.uid, mail_template_id[1], self.id, force_send=True,context=self.env.context)
        # Putting entry under Employer master.
        wsp_submission_data = self.env['wsp.submission.track'].search(
            [('employer_id', '=', self.employer_id.id), ('fiscal_year', '=', self.fiscal_year.id), ('name', '=', self.name)])
        wsp_sub_dat = False
        if wsp_submission_data:
            wsp_submission_data.write({'status': 'evaluated'})
            wsp_sub_dat = wsp_submission_data.read()
            # todo: add the child write loop
        if wsp_sub_dat:
            # raise Warning(wsp_sub_dat)
            for field in wsp_sub_dat[0].keys():
                if type(wsp_sub_dat[0].get(field)) == tuple:
                    wsp_sub_dat[0].update({field: wsp_sub_dat[0].get(field)[0]})
            for child in self.employer_id.child_employer_ids:
                child_partner = child.employer_id
                child_wsp_submission_data = self.env['wsp.submission.track'].search(
                    [('employer_id', '=', child_partner.id), ('fiscal_year', '=', self.fiscal_year.id),
                     ('name', '=', self.name)])

                dbg(child_partner)
                wsp_sub_dat[0].update({'employer_id':child_partner.id})
                child_wsp_submission_data.write(wsp_sub_dat[0])
                dbg(wsp_submission_data)
        self = self.with_context({'from_wsp_evaluated': True})
        self.employer_id.write({'wsp_submitted': True})
        self.write({'approved_by': self._uid})

        res = {
            'evaluator': self._uid,
            'date_evaluation': datetime.now(),
            'status': 'not_recommended',
            'state': 'evaluated',
            'comments': self.comments,
            'wsp_plan_id': self.id,
        }
        self.write(
            {'wsp_submission_status_ids': [(0, 0, res)], 'comments': ''})
        return True

    
    def action_evaluate2_wsp(self):
        #         self.write({'state':'evaluated'})
        self.write({'before_query_state': 'evaluated2','state': 'evaluated2', 'evaluated_rec_or_not': True})
        res = {
            'evaluator': self._uid,
            'date_evaluation': datetime.now(),
            'status': 'recommended',
            'state': 'evaluated2',
            'comments': self.comments,
            'wsp_plan_id': self.id,
        }
        self.write(
            {'wsp_submission_status_ids': [(0, 0, res)], 'comments': ''})
        return True

    
    def action_evaluate2_not_recommended_wsp(self):
        self.write({'before_query_state': 'evaluated2','state': 'evaluated2', 'evaluated_rec_or_not': False})
        res = {
            'evaluator': self._uid,
            'date_evaluation': datetime.now(),
            'status': 'not_recommended',
            'state': 'evaluated2',
            'comments': self.comments,
            'wsp_plan_id': self.id,
        }
        self.write(
            {'wsp_submission_status_ids': [(0, 0, res)], 'comments': ''})
        return True

    
    def action_wsp_approve(self):
        self.write({'before_query_state': 'approved','state': 'approved'})
        user = self.env['res.users'].browse(self._uid)
        res = {
            'evaluator': self._uid,
            'date_evaluation': datetime.now(),
            'status': 'recommended',
            'state': 'approved',
            'comments': self.comments,
            'wsp_plan_id': self.id,
        }
        self.write(
            {'wsp_submission_status_ids': [(0, 0, res)], 'comments': ''})

        wsp_submission_data = self.env['wsp.submission.track'].search(
            [('employer_id', '=', self.employer_id.id), ('fiscal_year', '=', self.fiscal_year.id), ('name', '=', self.name)])
        wsp_sub_dat = False
        if wsp_submission_data:
            wsp_submission_data.write(
                {'status': 'accepted', 'approved_by': user.name, 'approved_date': datetime.now(), })
            wsp_sub_dat = wsp_submission_data.read()
        if wsp_sub_dat:
            # raise Warning(wsp_sub_dat)
            for field in wsp_sub_dat[0].keys():
                if type(wsp_sub_dat[0].get(field)) == tuple:
                    wsp_sub_dat[0].update({field: wsp_sub_dat[0].get(field)[0]})
            for child in self.employer_id.child_employer_ids:
                child_partner = child.employer_id
                child_wsp_submission_data = self.env['wsp.submission.track'].search(
                    [('employer_id', '=', child_partner.id), ('fiscal_year', '=', self.fiscal_year.id),
                     ('name', '=', self.name)])

                dbg(child_partner)
                wsp_sub_dat[0].update({'employer_id':child_partner.id})
                child_wsp_submission_data.write(wsp_sub_dat[0])
                dbg(wsp_submission_data)
        # Sending Email Notification.
        self.write({'approval_date': datetime.today().date()})
        ir_model_data_obj = self.env['ir.model.data']
        mail_template_id = ir_model_data_obj.get_object_reference(
            'hwseta_sdp', 'email_template_wsp_approval')
        if mail_template_id:
            self.pool['email.template'].send_mail(self.env.cr, self.env.uid, mail_template_id[
                                                  1], self.id, force_send=True, context=self.env.context)
        return True

    
    def action_reject_wsp(self):
        self.write({'state': 'rejected'})
        res = {
            'evaluator': self._uid,
            'date_evaluation': datetime.now(),
            'status': 'not_recommended',
            'state': 'rejected',
            'comments': self.comments,
            'wsp_plan_id': self.id,
        }
        self.write(
            {'wsp_submission_status_ids': [(0, 0, res)], 'comments': ''})
        wsp_submission_data = self.env['wsp.submission.track'].search(
            [('employer_id', '=', self.employer_id.id), ('fiscal_year', '=', self.fiscal_year.id), ('name', '=', self.name)])
        wsp_sub_dat = False
        if wsp_submission_data:
            user = self.env['res.users'].browse(self._uid)
            wsp_submission_data.write(
                {'status': 'rejected', 'rejected_by': user.name, 'rejected_date': datetime.now(), })
            wsp_sub_dat = wsp_submission_data.read()
        if wsp_sub_dat:
            # raise Warning(wsp_sub_dat)
            for field in wsp_sub_dat[0].keys():
                if type(wsp_sub_dat[0].get(field)) == tuple:
                    wsp_sub_dat[0].update({field: wsp_sub_dat[0].get(field)[0]})
            for child in self.employer_id.child_employer_ids:
                child_partner = child.employer_id
                child_wsp_submission_data = self.env['wsp.submission.track'].search(
                    [('employer_id', '=', child_partner.id), ('fiscal_year', '=', self.fiscal_year.id),
                     ('name', '=', self.name)])

                dbg(child_partner)
                wsp_sub_dat[0].update({'employer_id': child_partner.id})
                child_wsp_submission_data.write(wsp_sub_dat[0])
                dbg(child_wsp_submission_data)
        # Sending Email Notification.
        self.write({'rejection_date': datetime.today().date()})
        ir_model_data_obj = self.env['ir.model.data']
        mail_template_id = ir_model_data_obj.get_object_reference(
            'hwseta_sdp', 'email_template_wsp_rejection')
        if mail_template_id:
            self.pool['email.template'].send_mail(self.env.cr, self.env.uid, mail_template_id[
                                                  1], self.id, force_send=True, context=self.env.context)
        return True

    
    def action_set_draft(self):
        # Unlinking the existing values and set back to draft state.
        self.write({
            'state': 'draft',
            'training_actual_ids': [(2, training_field_data.id) for training_field_data in self.env['actual.training.fields'].search([('actual_wsp_id', '=', self.id)])],
            'actual_adult_education_ids': [(2, actual_adult_data.id) for actual_adult_data in self.env['actual.adult.education.fields'].search([('actual_adult_wsp_id', '=', self.id)])],
            'total_employment_profile_ids': [(2, tot_empl.id) for tot_empl in self.env['total.employment.profile.fields'].search([('total_employment_wsp_id', '=', self.id)])],
            'training_planned_ids': [(2, plan_train_non.id) for plan_train_non in self.env['planned.training.fields'].search([('planned_training_non_wsp_id', '=', self.id)])],
            'planned_adult_education_ids': [(2, plan_adult.id) for plan_adult in self.env['planned.adult.education.training.fields'].search([('planned_adult_education_wsp_id', '=', self.id)])],
            'scarce_and_critical_skills_ids': [(2, scarce.id) for scarce in self.env['scarce.and.critical.skills.fields'].search([('scarce_and_critical_wsp_id', '=', self.id)])],
        })
        return True

    
    def import_previous(self):
        ''' This Method will import last year planned wsp for employer into current year. '''
        prev_fisc_id = None
        self = self.with_context(load_previous=True)
        wsp_fy_start_date = datetime.strptime(
            self.fiscal_year.date_start, '%Y-%m-%d').date()
        dbg('wsp_fy_start_date')
        dbg(wsp_fy_start_date)
        for fiscal_year in self.env['account.fiscalyear'].search([]):
            dbg(fiscal_year)
            start_date = datetime.strptime(
                fiscal_year.date_start, '%Y-%m-%d').date()
            if start_date.year == (wsp_fy_start_date.year - 1):
                dbg("got start date")
                prev_fisc_id = fiscal_year.id
        if prev_fisc_id:
            dbg(prev_fisc_id)
            # Getting previous year wsp record for the same employer.
            prev_wsp = self.search(
                [('employer_id', '=', self.employer_id.id), ('fiscal_year', '=', prev_fisc_id)])
            dbg(prev_wsp)
            if len(prev_wsp) > 1:
                raise Warning(_('System should not have multiple wsp for employer %s') % (
                    self.employer_id.name))
            # Getting planned training,adult education records and total employment prfile.
# 7/12/12 commented code
#             actual_training_list = []
#             actual_adult_training_list = []
#             total_employment_prfile_list=[]
            # Flushing record if one2many already consisting of records.
            if self.training_actual_ids:
                self.write({'training_actual_ids': [
                           (2, actual_data.id) for actual_data in self.training_actual_ids]})
            if self.actual_adult_education_ids:
                self.write({'actual_adult_education_ids': [
                           (2, adult_data.id) for adult_data in self.actual_adult_education_ids]})

#             for planned_train_data in prev_wsp.training_actual_planned_ids:
            for planned_train_data in prev_wsp.training_planned_ids:
                # 7/12/12   commented code
                #                 actual_training_list.append((0,0,{
                #                                                     'name' : planned_train_data.name,
                #                                                     'surname' : planned_train_data.surname,
                #                                                     'employee_id':planned_train_data.employee_id,
                #                                                     'code' : planned_train_data.code and planned_train_data.code.id,
                #                                                     'training_type' : planned_train_data.training_type,
                #                                                     'occupation' : planned_train_data.occupation and planned_train_data.occupation.id,
                #                                                     'specialization' : planned_train_data.specialization and planned_train_data.specialization.id,
                # #                                                     'municipality_id' : planned_train_data.municipality_id and planned_train_data.municipality_id.id,
                #                                                     'city_id' : planned_train_data.city_id and planned_train_data.city_id.id,
                #                                                     'urban' : planned_train_data.urban,
                #                                                     'learner_province' : planned_train_data.learner_province and planned_train_data.learner_province.id,
                #                                                     'socio_economic_status' : planned_train_data.socio_economic_status,
                #                                                     'type_training' : planned_train_data.type_training and planned_train_data.type_training.id,
                #                                                     'other_type_of_intervention':planned_train_data.other_type_of_intervention,
                #                                                     'name_training' : planned_train_data.name_training,
                #                                                     'pivotal_programme_qualification':planned_train_data.pivotal_programme_qualification,
                #                                                     'pivotal_programme_institution' : planned_train_data.pivotal_programme_institution,
                #                                                     'nqf_aligned' : planned_train_data.nqf_aligned,
                #                                                     'nqf_level' : planned_train_data.nqf_level or '',
                #                                                     'training_cost' : planned_train_data.training_cost,
                #                                                     'start_date' : planned_train_data.start_date,
                #                                                     'end_date' : planned_train_data.end_date,
                #                                                     ## For this year we will use this
                #                                                     'african_male':planned_train_data.african_male,
                #                                                     'african_female':planned_train_data.african_female,
                #                                                     'african_dissabled':planned_train_data.african_dissabled,
                #                                                     'coloured_male':planned_train_data.coloured_male,
                #                                                     'coloured_female':planned_train_data.coloured_female,
                #                                                     'coloured_dissabled':planned_train_data.coloured_dissabled,
                #                                                     'indian_male':planned_train_data.indian_male,
                #                                                     'indian_female':planned_train_data.indian_female,
                #                                                     'indian_dissabled':planned_train_data.indian_dissabled,
                #                                                     'white_male':planned_train_data.white_male,
                #                                                     'white_female':planned_train_data.white_female,
                #                                                     'white_dissabled':planned_train_data.white_dissabled,
                #                                                     'total_male' : planned_train_data.total_male,
                #                                                     'total_female' : planned_train_data.total_female,
                #                                                     'total_dissabled' : planned_train_data.total_dissabled,
                #                                                     'age_group_less' : planned_train_data.age_group_less,
                #                                                     'age_group_upto' : planned_train_data.age_group_upto,
                #                                                     'age_group_greater' : planned_train_data.age_group_greater,
                #                                                     'total_cost' : planned_train_data.total_cost,
                #                                                     ##
                #                                                     ## For next year needs to use like follows
                #                                                      'population_group' : planned_train_data.population_group,
                # #                                                     'age_group' : planned_train_data.age_group,
                #                                                      'gender' : planned_train_data.gender,
                #                                                     ##
                # #                                                     'major' : planned_train_data.major,
                # #                                                     'sub_major_group' : planned_train_data.sub_major_group,
                #                                                      'pivotal_programme_qualification' : planned_train_data.pivotal_programme_qualification,
                # #                                                     'pivotal_programme_type' : planned_train_data.pivotal_programme_type,
                #                                                      'dissability' : planned_train_data.dissability
                #                                                    }))

                # ofo_code = planned_train_data.code and planned_train_data.code.id or None
                # start getting the latest set of ofo codes that match the previous wsp if the set has been updated
                stripped_name = planned_train_data.code.name[5:]
                new_ofo_rec = self.env['ofo.code'].search([('name','like',stripped_name)])
                latest = 0
                for x in new_ofo_rec:
                    if x.id > latest:
                        latest = x.id
                latest = self.env['ofo.code'].browse(latest)
                if latest == planned_train_data.code:
                    dbg('using old codes ')
                    ofo_code = planned_train_data.code and planned_train_data.code.id or None
                else:
                    ofo_code = latest and latest.id or None
                occupation = latest.occupation and latest.occupation.id or None
                specialization = None
                for spec in latest.specialization_ids:
                    if planned_train_data.specialization.name.lower() == spec.name.lower():
                        specialization = spec.id
                province = planned_train_data.learner_province and planned_train_data.learner_province.id or None
                city = planned_train_data.city_id and planned_train_data.city_id.id or None
                training = planned_train_data.type_training and planned_train_data.type_training.id or None
                # New code to insert data using query 7/12/12
                self._cr.execute("insert into actual_training_fields (training_type,name,surname,employee_id,code,occupation, specialization, learner_province, city_id, urban, socio_economic_status, type_training,other_type_of_intervention,name_training,pivotal_programme_qualification,pivotal_programme_institution, training_cost, start_date, end_date,nqf_aligned, nqf_level, population_group,gender,dissability,actual_wsp_id) values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                                 (planned_train_data.training_type, planned_train_data.name, planned_train_data.surname, planned_train_data.employee_id, ofo_code, occupation, specialization, province, city, planned_train_data.urban, planned_train_data.socio_economic_status, training, planned_train_data.other_type_of_intervention or '', planned_train_data.name_training or '', planned_train_data.pivotal_programme_qualification or '', planned_train_data.pivotal_programme_institution or '', planned_train_data.training_cost, planned_train_data.start_date, planned_train_data.end_date, planned_train_data.nqf_aligned, planned_train_data.nqf_level or '', planned_train_data.population_group, planned_train_data.gender, planned_train_data.dissability, self.id))

            for adult_train_data in prev_wsp.actual_planned_adult_education_ids:
                # 7/12/12 commented code
                #                 actual_adult_training_list.append((0,0,{
                # #                                                         'ofo_code' : adult_train_data.ofo_code and adult_train_data.ofo_code.id,
                #                                                         'name' : adult_train_data.name,
                #                                                         'surname' : adult_train_data.surname,
                #                                                         'id_number' : adult_train_data.id_number,
                # #                                                         'african_male' : adult_train_data.african_male,
                # #                                                         'african_female' : adult_train_data.african_female,
                # #                                                         'african_dissabled' : adult_train_data.african_dissabled,
                # #                                                         'coloured_male' : adult_train_data.coloured_male,
                # #                                                         'coloured_female' : adult_train_data.coloured_female,
                # #                                                         'coloured_dissabled' : adult_train_data.coloured_dissabled,
                # #                                                         'indian_male' : adult_train_data.indian_male,
                # #                                                         'indian_female' : adult_train_data.indian_female,
                # #                                                         'indian_dissabled' : adult_train_data.indian_dissabled,
                # #                                                         'white_male' : adult_train_data.white_male,
                # #                                                         'white_female' : adult_train_data.white_female,
                # #                                                         'white_dissabled' : adult_train_data.white_dissabled,
                # #                                                         'total_male' : adult_train_data.total_male,
                # #                                                         'total_female' : adult_train_data.total_female,
                # #                                                         'total_dissabled' : adult_train_data.total_dissabled,
                #                                                         'population_group' : adult_train_data.population_group,
                #                                                         'gender' : adult_train_data.gender,
                #                                                         'dissability_status_and_type' : adult_train_data.dissability_status_and_type,
                #                                                         'municipality_id' : adult_train_data.municipality_id and adult_train_data.municipality_id.id,
                #                                                         'city_id' : adult_train_data.city_id and adult_train_data.city_id.id,
                #                                                         'province' : adult_train_data.province and adult_train_data.province.id,
                #                                                         'urban' : adult_train_data.urban,
                #                                                         'start_date' : adult_train_data.start_date,
                #                                                         'end_date' : adult_train_data.end_date,
                #                                                         'provider' : adult_train_data.provider,
                #                                                         'aet_level' : adult_train_data.aet_level,
                #                                                         'aet_subject' : adult_train_data.aet_subject,
                #                                                         }))

                province = adult_train_data.province and adult_train_data.province.id or None
                city = adult_train_data.city_id and adult_train_data.city_id.id or None

                # New code to insert data using query  7/12/12
                self._cr.execute("insert into actual_adult_education_fields (name, surname, id_number, population_group, gender, dissability_status_and_type, province, city_id, urban, start_date, end_date, provider, aet_level, actual_adult_wsp_id) values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)", (adult_train_data.name,
                                                                                                                                                                                                                                                                                                             adult_train_data.surname, adult_train_data.id_number, adult_train_data.population_group, adult_train_data.gender, adult_train_data.dissability_status_and_type, province, city, adult_train_data.urban, adult_train_data.start_date, adult_train_data.end_date, adult_train_data.provider, adult_train_data.aet_level, self.id))
                self._cr.execute(
                    "select max(id) from actual_adult_education_fields")
                adult_education_id = self._cr.fetchone()
                for aet_subject_id in adult_train_data.aet_subject:
                    self._cr.execute(
                        "insert into aet_subject_rel(actual_adult_education_ids,aet_subject_id) values(%s,%s)", (adult_education_id[0], aet_subject_id))


#             for total_employment_data in prev_wsp.total_employment_profile_ids:
#                 total_employment_prfile_list.append((0,0,{
#                                                         'sdl_number' : total_employment_data.sdl_number and total_employment_data.sdl_number.id ,
#                                                         'name':total_employment_data.name,
#                                                         'surname' : total_employment_data.surname,
#                                                         'citizen_resident_status_code':total_employment_data.citizen_resident_status_code,
#                                                         'id_type':total_employment_data.id_type,
#                                                         'employee_id':total_employment_data.employee_id,
#                                                         'dob':total_employment_data.dob,
#                                                         'ofo_code':total_employment_data.ofo_code and total_employment_data.ofo_code.id,
#                                                         'occupation':total_employment_data.occupation and total_employment_data.occupation.id,
#                                                         'specialization':total_employment_data.specialization and total_employment_data.specialization.id,
#                                                         'province':total_employment_data.province and total_employment_data.province.id,
#                                                         'city_id':total_employment_data.city_id and total_employment_data.city_id.id,
#                                                         'urban':total_employment_data.urban,
#                                                         'highest_education_level':total_employment_data.highest_education_level,
#                                                         'population_group':total_employment_data.population_group,
#                                                         'gender':total_employment_data.gender,
#                                                         'dissabilty':total_employment_data.dissability,
#                                                         'selection':True,
#                                                         }))
            # 7/12/12 commented code
            # self.with_context({'load_previous':True}).write({'training_actual_ids':actual_training_list,'actual_adult_education_ids':actual_adult_training_list,'total_employment_profile_ids':total_employment_prfile_list})
        self.write({'is_prev_wsp_loaded': True})
        return True

    
    def import_total_emp_profile(self):
        prev_fisc_id = None
        self = self.with_context(load_previous=True)
        wsp_fy_start_date = datetime.strptime(
            self.fiscal_year.date_start, '%Y-%m-%d').date()
        for fiscal_year in self.env['account.fiscalyear'].search([]):
            start_date = datetime.strptime(
                fiscal_year.date_start, '%Y-%m-%d').date()
            if start_date.year == (wsp_fy_start_date.year - 1):
                prev_fisc_id = fiscal_year.id
        if prev_fisc_id:
            # Getting previous year wsp record for the same employer.
            prev_wsp = self.search(
                [('employer_id', '=', self.employer_id.id), ('fiscal_year', '=', prev_fisc_id)])
            if len(prev_wsp) > 1:
                raise Warning(_('System should not have multiple wsp for employer %s') % (
                    self.employer_id.name))
            # Getting planned training,adult education records and total employment profile.
# 7/12/12 commented code
#            total_employment_profile_list=[]
        for total_employment_data in prev_wsp.total_employment_profile_ids:
            # 7/12/12 commented code
            #                 total_employment_profile_list.append((0,0,{
            #                                                         'sdl_number' : total_employment_data.sdl_number and total_employment_data.sdl_number.id ,
            #                                                         'name':total_employment_data.name,
            #                                                         'surname' : total_employment_data.surname,
            #                                                         'citizen_resident_status_code':total_employment_data.citizen_resident_status_code,
            #                                                         'id_type':total_employment_data.id_type,
            #                                                         'employee_id':total_employment_data.employee_id,
            #                                                         'dob':total_employment_data.dob,
            #                                                         'ofo_code':total_employment_data.ofo_code and total_employment_data.ofo_code.id,
            #                                                         'occupation':total_employment_data.occupation and total_employment_data.occupation.id,
            #                                                         'specialization':total_employment_data.specialization and total_employment_data.specialization.id,
            #                                                         'province':total_employment_data.province and total_employment_data.province.id,
            #                                                         'city_id':total_employment_data.city_id and total_employment_data.city_id.id,
            #                                                         'urban':total_employment_data.urban,
            #                                                         'highest_education_level':total_employment_data.highest_education_level,
            #                                                         'population_group':total_employment_data.population_group,
            #                                                         'gender':total_employment_data.gender,
            #                                                         'dissabilty':total_employment_data.dissability,
            #                                                         'selection':True,
            #                                                         }))
            sdl_number = total_employment_data.sdl_number and total_employment_data.sdl_number.id or None
            ofo_code = total_employment_data.ofo_code and total_employment_data.ofo_code.id or None
            occupation = total_employment_data.occupation and total_employment_data.occupation.id or None
            specialization = total_employment_data.specialization and total_employment_data.specialization.id or None
            province = total_employment_data.province and total_employment_data.province.id or None
            city = total_employment_data.city_id and total_employment_data.city_id.id or None
            # Newly added code to insert data using query
            self._cr.execute("insert into total_employment_profile_fields (selection,sdl_number, name, surname, citizen_resident_status_code, employee_id, id_type, dob, ofo_code, occupation, specialization, province, city_id, urban, highest_education_level, population_group, gender, dissability, total_employment_wsp_id) values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)", (True, sdl_number, total_employment_data.name,
                                                                                                                                                                                                                                                                                                                                                                                                      total_employment_data.surname, total_employment_data.citizen_resident_status_code, total_employment_data.employee_id, total_employment_data.id_type, total_employment_data.dob, ofo_code, occupation, specialization, province, city, total_employment_data.urban, total_employment_data.highest_education_level, total_employment_data.population_group, total_employment_data.gender, total_employment_data.dissability, self.id))
        # 7/12/12 commented code
        # self.with_context({'load_previous':True}).write({'total_employment_profile_ids':total_employment_profile_list})
        self.write({'is_tep_loaded': True})
        return True

    
    def import_selected_tep_to_plan(self):
        # Getting planned training from selected total employment profile.
        # 7/12/12 commented code
        #        planed_training_list = []
        if not self.total_employment_profile_ids:
            raise Warning(_('Please Load Total Empoyement Profile First..'))
        if self.total_employment_profile_ids:
            for total_employment_data in self.total_employment_profile_ids:
                if total_employment_data.selection:
                    # 7/12/12 commented code
                    #                         planed_training_list.append((0,0,{
                    #                                                                 'name':total_employment_data.name,
                    #                                                                 'surname' : total_employment_data.surname,
                    #                                                                 'employee_id':total_employment_data.employee_id,
                    #                                                                 'code':total_employment_data.ofo_code and total_employment_data.ofo_code.id,
                    #                                                                 'occupation':total_employment_data.occupation and total_employment_data.occupation.id,
                    #                                                                 'specialization':total_employment_data.specialization and total_employment_data.specialization.id,
                    #                                                                 'learner_province':total_employment_data.province and total_employment_data.province.id,
                    #                                                                 'city_id':total_employment_data.city_id and total_employment_data.city_id.id,
                    #                                                                 'urban':total_employment_data.urban,
                    #                                                                 'population_group':total_employment_data.population_group,
                    #                                                                 'gender':total_employment_data.gender,
                    #                                                                 'dissabilty':total_employment_data.dissability,
                    #                                                                 }))
                    stripped_name = total_employment_data.ofo_code.name[5:]
                    new_ofo_rec = self.env['ofo.code'].search([('name', 'like', stripped_name)])
                    latest = 0
                    for x in new_ofo_rec:
                        if x.id > latest:
                            latest = x.id
                    latest = self.env['ofo.code'].browse(latest)
                    if latest == total_employment_data.ofo_code:
                        dbg('using old codes ')
                        ofo_code = total_employment_data.ofo_code and total_employment_data.ofo_code.id or None
                    else:
                        ofo_code = latest and latest.id or None
                    occupation = latest.occupation and latest.occupation.id or None
                    specialization = None
                    for spec in latest.specialization_ids:
                        if total_employment_data.specialization.name.lower() == spec.name.lower():
                            specialization = spec.id
                    # ofo_code = total_employment_data.ofo_code and total_employment_data.ofo_code.id or None
                    # occupation = total_employment_data.occupation and total_employment_data.occupation.id or None
                    # specialization = total_employment_data.specialization and total_employment_data.specialization.id or None
                    province = total_employment_data.province and total_employment_data.province.id or None
                    city = total_employment_data.city_id and total_employment_data.city_id.id or None
                    # New Code to insert data using query 7/12/12
                    self._cr.execute("insert into planned_training_fields (training_type, name, surname, employee_id, code, occupation, specialization, learner_province, city_id, urban, population_group, gender, dissability, planned_training_non_wsp_id) values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)", (
                        '', total_employment_data.name, total_employment_data.surname, total_employment_data.employee_id, ofo_code, occupation, specialization, province, city, total_employment_data.urban, total_employment_data.population_group, total_employment_data.gender, total_employment_data.dissability, self.id))

# 7/12/12 commented code
#            self.write({'training_planned_ids':planed_training_list})
        self.write({'is_tep_to_planned_loaded': True})
        return True

    
    def get_population_group(self, race_value):
        race = ''
        if race_value == 'african':
            race = 'African'
        if race_value == 'coloured':
            race = 'Coloured'
        if race_value == 'indian':
            race = 'Indian'
        if race_value == 'white':
            race = 'White'
        return race

    
    def get_genders(self, gender_value):
        gender = ''
        if gender_value == 'male':
            gender = 'M - Male'
        if gender_value == 'female':
            gender = 'F - Female'
        return gender

    
    def get_disability_status(self, disability_value):
        disability_status = ''
        if disability_value == 'site':
            disability_status = '01 - Sight ( even with glasses )'
        if disability_value == 'hearing':
            disability_status = '02 - Hearing ( even with h.aid )'
        if disability_value == 'communication':
            disability_status = '03 - Communication ( talk/listen)'
        if disability_value == 'physical':
            disability_status = '04 - Physical ( move/stand, etc)'
        if disability_value == 'intellectual':
            disability_status = '05 - Intellectual ( learn,etc)'
        if disability_value == 'emotional':
            disability_status = '06 - Emotional ( behav/psych)'
        if disability_value == 'multiple':
            disability_status = '07 - Multiple'
        if disability_value == 'disabled':
            disability_status = '08 - Disabled'
        if disability_value == 'N-None':
            disability_status = 'none'
        return disability_status

    
    def get_urban_rural(self, urban_value):
        urban_rural = ''
        if urban_value == 'urban':
            urban_rural = 'Urban'
        if urban_value == 'rural':
            urban_rural = 'Rural'
        if urban_value == 'unknown':
            urban_rural = 'Unknown'
        return urban_rural

    
    def get_aet_level(self, aet_level_value):
        aet_level = ''
        if aet_level_value == 'aet_level_1':
            aet_level = 'AET Level 1'
        if aet_level_value == 'aet_level_2':
            aet_level = 'AET Level 2'
        if aet_level_value == 'aet_level_3':
            aet_level = 'AET Level 3'
        if aet_level_value == 'aet_level_4':
            aet_level = 'AET Level 4'
        return aet_level

    
    def get_aet_subject(self, aet_subject_value):
        aet_subject = ''
        if aet_subject_value == 'life_skills':
            aet_subject = 'Life Skills'
        if aet_subject_value == 'numeracy':
            aet_subject = 'Numeracy'
        if aet_subject_value == 'literacy':
            aet_subject = 'Literacy'
        return aet_subject

    
    def get_training_type(self, training_type_value):
        training_type = ''
        if training_type_value == 'pivotal':
            training_type = 'Pivotal'
        if training_type_value == 'non-pivotal':
            training_type = 'Non Pivotal'
        return training_type

    
    def get_socio_eco_status(self, socio_value):
        socio = ''
        if socio_value == 'employed':
            socio = 'Employed'
        if socio_value == 'unemployed':
            socio = 'Unemployed'
        return socio

    
    def get_nqf_aligned(self, nqf_aligned_value):
        nqf_aligned = ''
        if nqf_aligned_value == 'yes':
            nqf_aligned = 'Yes'
        if nqf_aligned_value == 'no':
            nqf_aligned = 'No'
        return nqf_aligned

    
    def get_nqf_level(self, nqf_level_value):
        nqf_level = ''
        if nqf_level_value == 'abet':
            nqf_level = 'ABET'
        if nqf_level_value == '':
            nqf_level = 'Level 1'
        if nqf_level_value == 'level2':
            nqf_level = 'Level 2'
        if nqf_level_value == 'level3':
            nqf_level = 'Level 3'
        if nqf_level_value == 'level4':
            nqf_level = 'Level 4'
        if nqf_level_value == 'level5':
            nqf_level = 'Level 5'
        if nqf_level_value == 'level6':
            nqf_level = 'Level 6'
        if nqf_level_value == 'level7':
            nqf_level = 'Level 7'
        if nqf_level_value == 'level8':
            nqf_level = 'Level 8'
        if nqf_level_value == 'level9':
            nqf_level = 'Level 9'
        if nqf_level_value == 'level10':
            nqf_level = 'Level 10'
        return nqf_level

    
    def get_citizen_status(self, citizen_value):
        citizen_status = ''
        if citizen_value == 'dual':
            citizen_status = 'D - Dual (SA plus other)'
        if citizen_value == 'other':
            citizen_status = 'O - Other'
        if citizen_value == 'sa':
            citizen_status = 'SA - South Africa'
        if citizen_value == 'unknown':
            citizen_status = 'U - Unknown'
        return citizen_status

    
    def get_id_type(self, id_value):
        id_status = ''
        if id_value == 'id_document':
            id_status = 'ID Document'
        if id_value == 'passport':
            id_status = 'Passport'
        return id_status

    
    def get_highest_edu(self, high_edu_value):
        highest_edu = ''
        if high_edu_value == 'abet_level_1':
            highest_edu = 'Abet Level 1'
        if high_edu_value == 'abet_level_2':
            highest_edu = 'Abet Level 2'
        if high_edu_value == 'abet_level_3':
            highest_edu = 'Abet Level 3'
        if high_edu_value == 'abet_level_4':
            highest_edu = 'Abet Level 4'
        if high_edu_value == 'nqf123':
            highest_edu = 'NQF 1,2,3'
        if high_edu_value == 'nqf45':
            highest_edu = 'NQF 4,5'
        if high_edu_value == 'nqf67':
            highest_edu = 'NQF 6,7'
        if high_edu_value == 'nqf8910':
            highest_edu = 'NQF 8,9,10'
        return highest_edu

    
    def export_previous(self):
        ''' This Method will export wsp template. '''

#         workbook = xlwt.Workbook(encoding = 'utf-8')
#         buffer = cStringIO.StringIO()
#
#         font = xlwt.Font()
#         font.bold = True
#         font.height = 350
#         style = xlwt.XFStyle()
#         style.font = font
#
#         font1 = xlwt.Font()
#         font1.bold = True
# #         font1.height = 200
#         style1 = xlwt.XFStyle()
#         style1.font = font1
        workbook = Workbook()
        buffer = cStringIO.StringIO()

#         if self.wsp_details_type == "actual":
        if self.training_actual_ids:
            row = col = 1
#                 worksheet1 = workbook.create_sheet(title="Actual Training")
            worksheet1 = workbook.active
            worksheet1.title = "Actual Training"

            h1 = worksheet1.cell(column=col, row=row, value="Type of Training")
            worksheet1.merge_cells(
                start_row=row, start_column=col, end_row=row + 1, end_column=col)
            h1.alignment = Alignment(horizontal='center', vertical='center')
            h1.font = Font(size=12, bold=True)
            col += 1

            h2 = worksheet1.cell(column=col, row=row, value="Occupation")
            worksheet1.merge_cells(
                start_row=row, start_column=col, end_row=row + 1, end_column=col)
            h2.alignment = Alignment(horizontal='center', vertical='center')
            h2.font = Font(size=12, bold=True)
            col += 1

            h3 = worksheet1.cell(column=col, row=row, value="Specialisation")
            worksheet1.merge_cells(
                start_row=row, start_column=col, end_row=row + 1, end_column=col)
            h3.alignment = Alignment(horizontal='center', vertical='center')
            h3.font = Font(size=12, bold=True)
            col += 1

            h4 = worksheet1.cell(column=col, row=row, value="Municipality")
            worksheet1.merge_cells(
                start_row=row, start_column=col, end_row=row + 1, end_column=col)
            h4.alignment = Alignment(horizontal='center', vertical='center')
            h4.font = Font(size=12, bold=True)
            col += 1

            h5 = worksheet1.cell(column=col, row=row, value="Urban/Rural")
            worksheet1.merge_cells(
                start_row=row, start_column=col, end_row=row + 1, end_column=col)
            h5.alignment = Alignment(horizontal='center', vertical='center')
            h5.font = Font(size=12, bold=True)
            col += 1

            h6 = worksheet1.cell(column=col, row=row, value="Province")
            worksheet1.merge_cells(
                start_row=row, start_column=col, end_row=row + 1, end_column=col)
            h6.alignment = Alignment(horizontal='center', vertical='center')
            h6.font = Font(size=12, bold=True)
            col += 1

            h7 = worksheet1.cell(
                column=col, row=row, value="Socio Economic\nstatus")
            worksheet1.merge_cells(
                start_row=row, start_column=col, end_row=row + 1, end_column=col)
            h7.alignment = Alignment(horizontal='center', vertical='center')
            h7.font = Font(size=12, bold=True)
            col += 1

            h8 = worksheet1.cell(
                column=col, row=row, value="Type of Training\nIntervention/Pivotal\nProgramme Type")
            worksheet1.merge_cells(
                start_row=row, start_column=col, end_row=row + 1, end_column=col)
            h8.alignment = Alignment(horizontal='center', vertical='center')
            h8.font = Font(size=12, bold=True)
            col += 1

            h9 = worksheet1.cell(
                column=col, row=row, value="Name of Training\nIntervention/Pivotal\nProgramme Qualification")
            worksheet1.merge_cells(
                start_row=row, start_column=col, end_row=row + 1, end_column=col)
            h9.alignment = Alignment(horizontal='center', vertical='center')
            h9.font = Font(size=12, bold=True)
            col += 1

            h10 = worksheet1.cell(
                column=col, row=row, value="Piotal Programme\nInstitution")
            worksheet1.merge_cells(
                start_row=row, start_column=col, end_row=row + 1, end_column=col)
            h10.alignment = Alignment(horizontal='center', vertical='center')
            h10.font = Font(size=12, bold=True)
            col += 1

            h11 = worksheet1.cell(column=col, row=row, value="NQF\nAligned")
            worksheet1.merge_cells(
                start_row=row, start_column=col, end_row=row + 1, end_column=col)
            h11.alignment = Alignment(horizontal='center', vertical='center')
            h11.font = Font(size=12, bold=True)
            col += 1

            h12 = worksheet1.cell(column=col, row=row, value="NQF\nLevel")
            worksheet1.merge_cells(
                start_row=row, start_column=col, end_row=row + 1, end_column=col)
            h12.alignment = Alignment(horizontal='center', vertical='center')
            h12.font = Font(size=12, bold=True)
            col += 1

            h13 = worksheet1.cell(column=col, row=row, value="Cost")
            worksheet1.merge_cells(
                start_row=row, start_column=col, end_row=row + 1, end_column=col)
            h13.alignment = Alignment(horizontal='center', vertical='center')
            h13.font = Font(size=12, bold=True)
            col += 1

            h14 = worksheet1.cell(column=col, row=row, value="Start Date")
            worksheet1.merge_cells(
                start_row=row, start_column=col, end_row=row + 1, end_column=col)
            h14.alignment = Alignment(horizontal='center', vertical='center')
            h14.font = Font(size=12, bold=True)
            col += 1

            h15 = worksheet1.cell(column=col, row=row, value="End Date")
            worksheet1.merge_cells(
                start_row=row, start_column=col, end_row=row + 1, end_column=col)
            h15.alignment = Alignment(horizontal='center', vertical='center')
            h15.font = Font(size=12, bold=True)
            col += 1

            h16 = worksheet1.cell(column=col, row=row, value="AFRICAN")
            h16.alignment = Alignment(horizontal='center', vertical='center')
            h16.font = Font(size=12, bold=True)
            worksheet1.merge_cells(
                start_row=1, start_column=col, end_row=1, end_column=col + 2)

            h16i = worksheet1.cell(column=col, row=row + 1, value="M")
            h16i.font = Font(size=12, bold=True)

            col += 1
            h16j = worksheet1.cell(column=col, row=row + 1, value="F")
            h16j.font = Font(size=12, bold=True)

            col += 1
            h16k = worksheet1.cell(column=col, row=row + 1, value="D")
            h16k.font = Font(size=12, bold=True)

            col += 1
            h17 = worksheet1.cell(column=col, row=row, value="COLOURED")
            h17.alignment = Alignment(horizontal='center', vertical='center')
            h17.font = Font(size=12, bold=True)
            worksheet1.merge_cells(
                start_row=1, start_column=col, end_row=1, end_column=col + 2)

            h17i = worksheet1.cell(column=col, row=row + 1, value="M")
            h17i.font = Font(size=12, bold=True)

            col += 1
            h17j = worksheet1.cell(column=col, row=row + 1, value="F")
            h17j.font = Font(size=12, bold=True)

            col += 1
            h17k = worksheet1.cell(column=col, row=row + 1, value="D")
            h17k.font = Font(size=12, bold=True)

            col += 1
            h18 = worksheet1.cell(column=col, row=row, value="INDIAN")
            h18.alignment = Alignment(horizontal='center', vertical='center')
            h18.font = Font(size=12, bold=True)
            worksheet1.merge_cells(
                start_row=1, start_column=col, end_row=1, end_column=col + 2)

            h18i = worksheet1.cell(column=col, row=row + 1, value="M")
            h18i.font = Font(size=12, bold=True)

            col += 1
            h18j = worksheet1.cell(column=col, row=row + 1, value="F")
            h18j.font = Font(size=12, bold=True)

            col += 1
            h18k = worksheet1.cell(column=col, row=row + 1, value="D")
            h18k.font = Font(size=12, bold=True)

            col += 1
            h19 = worksheet1.cell(column=col, row=row, value="WHITE")
            h19.alignment = Alignment(horizontal='center', vertical='center')
            h19.font = Font(size=12, bold=True)
            worksheet1.merge_cells(
                start_row=1, start_column=col, end_row=1, end_column=col + 2)

            h19i = worksheet1.cell(column=col, row=row + 1, value="M")
            h19i.font = Font(size=12, bold=True)

            col += 1
            h19j = worksheet1.cell(column=col, row=row + 1, value="F")
            h19j.font = Font(size=12, bold=True)

            col += 1
            h19k = worksheet1.cell(column=col, row=row + 1, value="D")
            h19k.font = Font(size=12, bold=True)

            col += 1
            h20 = worksheet1.cell(column=col, row=row, value="Total")
            h20.alignment = Alignment(horizontal='center', vertical='center')
            h20.font = Font(size=12, bold=True)
            worksheet1.merge_cells(
                start_row=1, start_column=col, end_row=1, end_column=col + 2)

            h20i = worksheet1.cell(column=col, row=row + 1, value="M")
            h20i.font = Font(size=12, bold=True)

            col += 1
            h20j = worksheet1.cell(column=col, row=row + 1, value="F")
            h20j.font = Font(size=12, bold=True)

            col += 1
            h20k = worksheet1.cell(column=col, row=row + 1, value="D")
            h20k.font = Font(size=12, bold=True)

            col += 1
            h21 = worksheet1.cell(column=col, row=row, value="AGE GROUP")
            h21.alignment = Alignment(horizontal='center', vertical='center')
            h21.font = Font(size=12, bold=True)
            worksheet1.merge_cells(
                start_row=1, start_column=col, end_row=1, end_column=col + 2)

            h21i = worksheet1.cell(column=col, row=row + 1, value="<35")
            h21i.font = Font(size=12, bold=True)

            col += 1
            h21j = worksheet1.cell(column=col, row=row + 1, value="35-55")
            h21j.font = Font(size=12, bold=True)

            col += 1
            h21k = worksheet1.cell(column=col, row=row + 1, value=">55")
            h21k.font = Font(size=12, bold=True)

            col += 1
            h22 = worksheet1.cell(column=col, row=row, value="Total Cost")
            worksheet1.merge_cells(
                start_row=row, start_column=col, end_row=row + 1, end_column=col)
            h22.alignment = Alignment(horizontal='center', vertical='center')
            h22.font = Font(size=12, bold=True)
            col += 1

            row = 3
            col = 1
            for actual_training_line in self.training_actual_ids:
                row += 1
                col = 1
                for actual_training_data in actual_training_line:

                    if actual_training_data.training_type != 0:
                        worksheet1.cell(column=col, row=row, value=self.get_training_type(
                            actual_training_data.training_type))
                        col += 1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col += 1

                    if actual_training_data.occupation:
                        worksheet1.cell(
                            column=col, row=row, value=actual_training_data.occupation and actual_training_data.occupation.name)
                        col += 1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col += 1
                    if actual_training_data.specialization:
                        worksheet1.cell(
                            column=col, row=row, value=actual_training_data.specialization and actual_training_data.specialization.name)
                        col += 1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col += 1
                    if actual_training_data.municipality_id:
                        worksheet1.cell(
                            column=col, row=row, value=actual_training_data.municipality_id and actual_training_data.municipality_id.name)
                        col += 1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col += 1

                    if actual_training_data.urban != 0:
                        worksheet1.cell(
                            column=col, row=row, value=self.get_urban_rural(actual_training_data.urban))
                        col += 1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col += 1

                    if actual_training_data.learner_province:
                        worksheet1.cell(
                            column=col, row=row, value=actual_training_data.learner_province and actual_training_data.learner_province.name)
                        col += 1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col += 1

                    if actual_training_data.socio_economic_status != 0:
                        worksheet1.cell(column=col, row=row, value=self.get_socio_eco_status(
                            actual_training_data.socio_economic_status))
                        col += 1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col += 1

                    if actual_training_data.type_training:
                        worksheet1.cell(
                            column=col, row=row, value=actual_training_data.type_training and actual_training_data.type_training.name)
                        col += 1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col += 1

                    if actual_training_data.name_training != 0:
                        worksheet1.cell(
                            column=col, row=row, value=actual_training_data.name_training)
                        col += 1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col += 1

                    if actual_training_data.pivotal_programme_institution != 0:
                        worksheet1.cell(
                            column=col, row=row, value=actual_training_data.pivotal_programme_institution)
                        col += 1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col += 1
                    if actual_training_data.nqf_aligned != 0:
                        worksheet1.cell(column=col, row=row, value=self.get_nqf_aligned(
                            actual_training_data.nqf_aligned))
                        col += 1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col += 1
                    if actual_training_data.nqf_level != 0:

                        worksheet1.cell(
                            column=col, row=row, value=self.get_nqf_level(actual_training_data.nqf_level))
                        col += 1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col += 1

                    if actual_training_data.training_cost != 0:
                        worksheet1.cell(
                            column=col, row=row, value=actual_training_data.training_cost)
                        col += 1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col += 1
                    if actual_training_data.start_date != 0:
                        worksheet1.cell(
                            column=col, row=row, value=actual_training_data.start_date)
                        col += 1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col += 1

                    if actual_training_data.end_date != 0:
                        worksheet1.cell(
                            column=col, row=row, value=actual_training_data.end_date)
                        col += 1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col += 1
                    worksheet1.cell(
                        column=col, row=row, value=actual_training_data.african_male)
                    col += 1
                    worksheet1.cell(
                        column=col, row=row, value=actual_training_data.african_female)
                    col += 1
                    worksheet1.cell(
                        column=col, row=row, value=actual_training_data.african_dissabled)
                    col += 1
                    worksheet1.cell(
                        column=col, row=row, value=actual_training_data.coloured_male)
                    col += 1
                    worksheet1.cell(
                        column=col, row=row, value=actual_training_data.coloured_female)
                    col += 1
                    worksheet1.cell(
                        column=col, row=row, value=actual_training_data.coloured_dissabled)
                    col += 1
                    worksheet1.cell(
                        column=col, row=row, value=actual_training_data.indian_male)
                    col += 1
                    worksheet1.cell(
                        column=col, row=row, value=actual_training_data.indian_female)
                    col += 1
                    worksheet1.cell(
                        column=col, row=row, value=actual_training_data.indian_dissabled)
                    col += 1
                    worksheet1.cell(
                        column=col, row=row, value=actual_training_data.white_male)
                    col += 1
                    worksheet1.cell(
                        column=col, row=row, value=actual_training_data.white_female)
                    col += 1
                    worksheet1.cell(
                        column=col, row=row, value=actual_training_data.white_dissabled)
                    col += 1
                    worksheet1.cell(
                        column=col, row=row, value=actual_training_data.total_male)
                    col += 1
                    worksheet1.cell(
                        column=col, row=row, value=actual_training_data.total_female)
                    col += 1
                    worksheet1.cell(
                        column=col, row=row, value=actual_training_data.total_dissabled)
                    col += 1
                    worksheet1.cell(
                        column=col, row=row, value=actual_training_data.age_group_less)
                    col += 1
                    worksheet1.cell(
                        column=col, row=row, value=actual_training_data.age_group_upto)
                    col += 1
                    worksheet1.cell(
                        column=col, row=row, value=actual_training_data.age_group_greater)
                    col += 1
                    worksheet1.cell(
                        column=col, row=row, value=actual_training_data.total_cost)

        if self.actual_adult_education_ids:
            row = col = 1
            worksheet2 = workbook.create_sheet(
                title="Adult Education and Training")

            h1 = worksheet2.cell(column=col, row=row, value="First Name")
            h1.alignment = Alignment(horizontal='center', vertical='center')
            h1.font = Font(size=12, bold=True)
            col += 1

            h2 = worksheet2.cell(column=col, row=row, value="SurName")
            h2.alignment = Alignment(horizontal='center', vertical='center')
            h2.font = Font(size=12, bold=True)
            col += 1

            h3 = worksheet2.cell(column=col, row=row, value="Id Number")
            h3.alignment = Alignment(horizontal='center', vertical='center')
            h3.font = Font(size=12, bold=True)
            col += 1

            h33 = worksheet2.cell(
                column=col, row=row, value="Population Group")
            h33.alignment = Alignment(horizontal='center', vertical='center')
            h33.font = Font(size=12, bold=True)
            col += 1

            h4 = worksheet2.cell(column=col, row=row, value="Gender")
            h4.alignment = Alignment(horizontal='center', vertical='center')
            h4.font = Font(size=12, bold=True)
            col += 1

            h5 = worksheet2.cell(
                column=col, row=row, value="Disability Status And Type")
            h5.alignment = Alignment(horizontal='center', vertical='center')
            h5.font = Font(size=12, bold=True)
            col += 1

            h6 = worksheet2.cell(column=col, row=row, value="Municipality")
            h6.alignment = Alignment(horizontal='center', vertical='center')
            h6.font = Font(size=12, bold=True)
            col += 1

            h7 = worksheet2.cell(column=col, row=row, value="Learner Province")
            h7.alignment = Alignment(horizontal='center', vertical='center')
            h7.font = Font(size=12, bold=True)
            col += 1

            h8 = worksheet2.cell(column=col, row=row, value="Urban/Rural")
            h8.alignment = Alignment(horizontal='center', vertical='center')
            h8.font = Font(size=12, bold=True)
            col += 1

            h9 = worksheet2.cell(column=col, row=row, value="AET Start Date")
            h9.alignment = Alignment(horizontal='center', vertical='center')
            h9.font = Font(size=12, bold=True)
            col += 1

            h10 = worksheet2.cell(column=col, row=row, value="AET End Date")
            h10.alignment = Alignment(horizontal='center', vertical='center')
            h10.font = Font(size=12, bold=True)
            col += 1

            h10 = worksheet2.cell(column=col, row=row, value="Provider")
            h10.alignment = Alignment(horizontal='center', vertical='center')
            h10.font = Font(size=12, bold=True)
            col += 1

            h11 = worksheet2.cell(column=col, row=row, value="AET Level")
            h11.alignment = Alignment(horizontal='center', vertical='center')
            h11.font = Font(size=12, bold=True)
            col += 1

            h12 = worksheet2.cell(column=col, row=row, value="AET Subject")
            h12.alignment = Alignment(horizontal='center', vertical='center')
            h12.font = Font(size=12, bold=True)

            for actual_adult_line in self.actual_adult_education_ids:
                row += 1
                col = 1
                for actual_adult_data in actual_adult_line:

                    if actual_adult_data.name != 0:
                        worksheet2.cell(
                            column=col, row=row, value=actual_adult_data.name)
                        col += 1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col += 1

                    if actual_adult_data.surname != 0:
                        worksheet2.cell(
                            column=col, row=row, value=actual_adult_data.surname)
                        col += 1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col += 1

                    if actual_adult_data.id_number != 0:
                        worksheet2.cell(
                            column=col, row=row, value=actual_adult_data.id_number)
                        col += 1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col += 1

                    if actual_adult_data.population_group != 0:
                        worksheet2.cell(column=col, row=row, value=self.get_population_group(
                            actual_adult_data.population_group))
                        col += 1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col += 1

                    if actual_adult_data.gender != 0:
                        worksheet2.cell(
                            column=col, row=row, value=self.get_genders(actual_adult_data.gender))
                        col += 1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col += 1

                    if actual_adult_data.dissability_status_and_type != 0:
                        worksheet2.cell(column=col, row=row, value=self.get_disability_status(
                            actual_adult_data.dissability_status_and_type))
                        col += 1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col += 1

                    if actual_adult_data.municipality_id:
                        worksheet2.cell(
                            column=col, row=row, value=actual_adult_data.municipality_id and actual_adult_data.municipality_id.name)
                        col += 1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col += 1

                    if actual_adult_data.province:
                        worksheet2.cell(
                            column=col, row=row, value=actual_adult_data.province and actual_adult_data.province.name)
                        col += 1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col += 1

                    if actual_adult_data.urban != 0:
                        worksheet2.cell(
                            column=col, row=row, value=self.get_urban_rural(actual_adult_data.urban))
                        col += 1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col += 1

                    if actual_adult_data.start_date != 0:
                        worksheet2.cell(
                            column=col, row=row, value=actual_adult_data.start_date)
                        col += 1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col += 1

                    if actual_adult_data.end_date != 0:
                        worksheet2.cell(
                            column=col, row=row, value=actual_adult_data.end_date)
                        col += 1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col += 1

                    if actual_adult_data.provider != 0:
                        worksheet2.cell(
                            column=col, row=row, value=actual_adult_data.provider)
                        col += 1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col += 1

                    if actual_adult_data.aet_level != 0:
                        worksheet2.cell(
                            column=col, row=row, value=self.get_aet_level(actual_adult_data.aet_level))
                        col += 1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col += 1

                    if actual_adult_data.aet_subject != 0:
                        worksheet2.cell(
                            column=col, row=row, value=self.get_aet_subject(actual_adult_data.aet_subject))
                        col += 1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col += 1

        workbook.save(buffer)
        out_data = base64.encodestring(buffer.getvalue())
        attachment_obj = self.env['ir.attachment']
        new_attach = attachment_obj.create({
            'name': 'wsp.xlsx',
            'res_name': 'wsp_data',
            'type': 'binary',
            'res_model': 'wsp.plan',
            'datas': out_data,
        })

        return {
            'type': 'ir.actions.act_url',
            'url': '/web/binary/saveas?model=ir.attachment&field=datas&filename_field=name&id=%s' % (new_attach.id, ),
            'target': 'self',
        }

    
    def export_previous_wsp(self):
        ''' This Method will export wsp template. '''

        workbook = Workbook()
        buffer = cStringIO.StringIO()

        if self.total_employment_profile_ids:
            row = col = 1
#                 worksheet1 = workbook.create_sheet(title="Actual Training")
            worksheet1 = workbook.active
            worksheet1.title = "Total Employment Profile"

            h1 = worksheet1.cell(column=col, row=row, value="SDL Number")
            worksheet1.merge_cells(
                start_row=row, start_column=col, end_row=row + 1, end_column=col)
            h1.alignment = Alignment(horizontal='center', vertical='center')
            h1.font = Font(size=12, bold=True)
            col += 1

            h2 = worksheet1.cell(column=col, row=row, value="First Name")
            worksheet1.merge_cells(
                start_row=row, start_column=col, end_row=row + 1, end_column=col)
            h2.alignment = Alignment(horizontal='center', vertical='center')
            h2.font = Font(size=12, bold=True)
            col += 1

            h3 = worksheet1.cell(column=col, row=row, value="Last Name")
            worksheet1.merge_cells(
                start_row=row, start_column=col, end_row=row + 1, end_column=col)
            h3.alignment = Alignment(horizontal='center', vertical='center')
            h3.font = Font(size=12, bold=True)
            col += 1

            h4 = worksheet1.cell(column=col, row=row, value="Citizen Status")
            worksheet1.merge_cells(
                start_row=row, start_column=col, end_row=row + 1, end_column=col)
            h4.alignment = Alignment(horizontal='center', vertical='center')
            h4.font = Font(size=12, bold=True)
            col += 1

            h5 = worksheet1.cell(column=col, row=row, value="Employee Id")
            worksheet1.merge_cells(
                start_row=row, start_column=col, end_row=row + 1, end_column=col)
            h5.alignment = Alignment(horizontal='center', vertical='center')
            h5.font = Font(size=12, bold=True)
            col += 1

            h6 = worksheet1.cell(column=col, row=row, value="ID Type")
            worksheet1.merge_cells(
                start_row=row, start_column=col, end_row=row + 1, end_column=col)
            h6.alignment = Alignment(horizontal='center', vertical='center')
            h6.font = Font(size=12, bold=True)
            col += 1

            h7 = worksheet1.cell(column=col, row=row, value="Date Of Birth")
            worksheet1.merge_cells(
                start_row=row, start_column=col, end_row=row + 1, end_column=col)
            h7.alignment = Alignment(horizontal='center', vertical='center')
            h7.font = Font(size=12, bold=True)
            col += 1

            h8 = worksheet1.cell(column=col, row=row, value="OFO Code")
            worksheet1.merge_cells(
                start_row=row, start_column=col, end_row=row + 1, end_column=col)
            h8.alignment = Alignment(horizontal='center', vertical='center')
            h8.font = Font(size=12, bold=True)
            col += 1

            h9 = worksheet1.cell(column=col, row=row, value="Occupation")
            worksheet1.merge_cells(
                start_row=row, start_column=col, end_row=row + 1, end_column=col)
            h9.alignment = Alignment(horizontal='center', vertical='center')
            h9.font = Font(size=12, bold=True)
            col += 1

            h10 = worksheet1.cell(column=col, row=row, value="Specialisation")
            worksheet1.merge_cells(
                start_row=row, start_column=col, end_row=row + 1, end_column=col)
            h10.alignment = Alignment(horizontal='center', vertical='center')
            h10.font = Font(size=12, bold=True)
            col += 1

            h11 = worksheet1.cell(column=col, row=row, value="City")
            worksheet1.merge_cells(
                start_row=row, start_column=col, end_row=row + 1, end_column=col)
            h11.alignment = Alignment(horizontal='center', vertical='center')
            h11.font = Font(size=12, bold=True)
            col += 1

            h12 = worksheet1.cell(column=col, row=row, value="Province")
            worksheet1.merge_cells(
                start_row=row, start_column=col, end_row=row + 1, end_column=col)
            h12.alignment = Alignment(horizontal='center', vertical='center')
            h12.font = Font(size=12, bold=True)
            col += 1

            h13 = worksheet1.cell(column=col, row=row, value="Urban/Rural")
            worksheet1.merge_cells(
                start_row=row, start_column=col, end_row=row + 1, end_column=col)
            h13.alignment = Alignment(horizontal='center', vertical='center')
            h13.font = Font(size=12, bold=True)
            col += 1

            h14 = worksheet1.cell(
                column=col, row=row, value="Highest Education Level")
            worksheet1.merge_cells(
                start_row=row, start_column=col, end_row=row + 1, end_column=col)
            h14.alignment = Alignment(horizontal='center', vertical='center')
            h14.font = Font(size=12, bold=True)
            col += 1

            h15 = worksheet1.cell(column=col, row=row, value="Race")
            worksheet1.merge_cells(
                start_row=row, start_column=col, end_row=row + 1, end_column=col)
            h15.alignment = Alignment(horizontal='center', vertical='center')
            h15.font = Font(size=12, bold=True)
            col += 1

            h16 = worksheet1.cell(column=col, row=row, value="Gender")
            worksheet1.merge_cells(
                start_row=row, start_column=col, end_row=row + 1, end_column=col)
            h16.alignment = Alignment(horizontal='center', vertical='center')
            h16.font = Font(size=12, bold=True)
            col += 1

            h17 = worksheet1.cell(column=col, row=row, value="Disability")
            worksheet1.merge_cells(
                start_row=row, start_column=col, end_row=row + 1, end_column=col)
            h17.alignment = Alignment(horizontal='center', vertical='center')
            h17.font = Font(size=12, bold=True)

            row = 3
            col = 1
            for total_employment_profile_line in self.total_employment_profile_ids:
                row += 1
                col = 1
                for total_employment_profile_data in total_employment_profile_line:

                    if total_employment_profile_data.sdl_number:
                        worksheet1.cell(
                            column=col, row=row, value=total_employment_profile_data.sdl_number and total_employment_profile_data.sdl_number.name)
                        col += 1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col += 1

                    if total_employment_profile_data.name != 0:
                        worksheet1.cell(
                            column=col, row=row, value=total_employment_profile_data.name)
                        col += 1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col += 1

                    if total_employment_profile_data.surname != 0:
                        worksheet1.cell(
                            column=col, row=row, value=total_employment_profile_data.surname)
                        col += 1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col += 1

                    if total_employment_profile_data.citizen_resident_status_code != 0:
                        worksheet1.cell(column=col, row=row, value=self.get_citizen_status(
                            total_employment_profile_data.citizen_resident_status_code))
                        col += 1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col += 1

                    if total_employment_profile_data.employee_id != 0:
                        worksheet1.cell(
                            column=col, row=row, value=total_employment_profile_data.employee_id)
                        col += 1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col += 1

                    if total_employment_profile_data.id_type != 0:
                        worksheet1.cell(column=col, row=row, value=self.get_id_type(
                            total_employment_profile_data.id_type))
                        col += 1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col += 1

                    if total_employment_profile_data.dob != 0:
                        worksheet1.cell(
                            column=col, row=row, value=total_employment_profile_data.dob)
                        col += 1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col += 1

                    if total_employment_profile_data.ofo_code:
                        worksheet1.cell(
                            column=col, row=row, value=total_employment_profile_data.ofo_code and total_employment_profile_data.ofo_code.name)
                        col += 1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col += 1

                    if total_employment_profile_data.occupation:
                        worksheet1.cell(
                            column=col, row=row, value=total_employment_profile_data.occupation and total_employment_profile_data.occupation.name)
                        col += 1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col += 1

                    if total_employment_profile_data.specialization:
                        worksheet1.cell(
                            column=col, row=row, value=total_employment_profile_data.ofo_code and total_employment_profile_data.specialization.name)
                        col += 1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col += 1

                    if total_employment_profile_data.city_id:
                        worksheet1.cell(
                            column=col, row=row, value=total_employment_profile_data.city_id and total_employment_profile_data.city_id.name)
                        col += 1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col += 1

                    if total_employment_profile_data.province:
                        worksheet1.cell(
                            column=col, row=row, value=total_employment_profile_data.province and total_employment_profile_data.province.name)
                        col += 1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col += 1

                    if total_employment_profile_data.urban != 0:
                        worksheet1.cell(column=col, row=row, value=self.get_urban_rural(
                            total_employment_profile_data.urban))
                        col += 1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col += 1

                    if total_employment_profile_data.highest_education_level != 0:
                        worksheet1.cell(column=col, row=row, value=self.get_highest_edu(
                            total_employment_profile_data.highest_education_level))
                        col += 1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col += 1

                    if total_employment_profile_data.population_group != 0:
                        worksheet1.cell(column=col, row=row, value=self.get_population_group(
                            total_employment_profile_data.population_group))
                        col += 1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col += 1

                    if total_employment_profile_data.gender != 0:
                        worksheet1.cell(column=col, row=row, value=self.get_genders(
                            total_employment_profile_data.gender))
                        col += 1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col += 1

                    if total_employment_profile_data.dissability != 0:
                        worksheet1.cell(column=col, row=row, value=self.get_nqf_aligned(
                            total_employment_profile_data.dissability))
                        col += 1
                    else:
                        worksheet1.cell(column=col, row=row, value='')
                        col += 1

        if self.training_planned_ids:
            row = col = 1
            worksheet2 = workbook.create_sheet(title="Planned Training")

            h1 = worksheet2.cell(column=col, row=row, value="Type Of Training")
            h1.alignment = Alignment(horizontal='center', vertical='center')
            h1.font = Font(size=12, bold=True)
            col += 1

            h2 = worksheet2.cell(column=col, row=row, value="Name")
            h2.alignment = Alignment(horizontal='center', vertical='center')
            h2.font = Font(size=12, bold=True)
            col += 1

            h3 = worksheet2.cell(column=col, row=row, value="SurName")
            h3.alignment = Alignment(horizontal='center', vertical='center')
            h3.font = Font(size=12, bold=True)
            col += 1

            h4 = worksheet2.cell(column=col, row=row, value="Employee Id")
            h4.alignment = Alignment(horizontal='center', vertical='center')
            h4.font = Font(size=12, bold=True)
            col += 1

            h5 = worksheet2.cell(column=col, row=row, value="OFO Code")
            h5.alignment = Alignment(horizontal='center', vertical='center')
            h5.font = Font(size=12, bold=True)
            col += 1

            h6 = worksheet2.cell(column=col, row=row, value="Occupation")
            h6.alignment = Alignment(horizontal='center', vertical='center')
            h6.font = Font(size=12, bold=True)
            col += 1

            h7 = worksheet2.cell(column=col, row=row, value="Specialisation")
            h7.alignment = Alignment(horizontal='center', vertical='center')
            h7.font = Font(size=12, bold=True)
            col += 1

            h8 = worksheet2.cell(column=col, row=row, value="City")
            h8.alignment = Alignment(horizontal='center', vertical='center')
            h8.font = Font(size=12, bold=True)
            col += 1

            h9 = worksheet2.cell(column=col, row=row, value="Urban/Rural")
            h9.alignment = Alignment(horizontal='center', vertical='center')
            h9.font = Font(size=12, bold=True)
            col += 1

            h10 = worksheet2.cell(column=col, row=row, value="Province")
            h10.alignment = Alignment(horizontal='center', vertical='center')
            h10.font = Font(size=12, bold=True)
            col += 1

            h11 = worksheet2.cell(
                column=col, row=row, value="Employed/UnEmployed")
            h11.alignment = Alignment(horizontal='center', vertical='center')
            h11.font = Font(size=12, bold=True)
            col += 1

            h12 = worksheet2.cell(
                column=col, row=row, value="Type of Training Intervention")
            h12.alignment = Alignment(horizontal='center', vertical='center')
            h12.font = Font(size=12, bold=True)
            col += 1

            h13 = worksheet2.cell(
                column=col, row=row, value="Other Type Of Training Intervention/\nPivotal programme Type")
            h13.alignment = Alignment(horizontal='center', vertical='center')
            h13.font = Font(size=12, bold=True)
            col += 1

            h14 = worksheet2.cell(
                column=col, row=row, value="Name of training\nIntervention")
            h14.alignment = Alignment(horizontal='center', vertical='center')
            h14.font = Font(size=12, bold=True)
            col += 1

            h16 = worksheet2.cell(
                column=col, row=row, value="Pivotal Programme\nQualification")
            h16.alignment = Alignment(horizontal='center', vertical='center')
            h16.font = Font(size=12, bold=True)
            col += 1

            h17 = worksheet2.cell(
                column=col, row=row, value="Pivotal Programme\ninstitution")
            h17.alignment = Alignment(horizontal='center', vertical='center')
            h17.font = Font(size=12, bold=True)
            col += 1

            h18 = worksheet2.cell(
                column=col, row=row, value="Cost Per Learner")
            h18.alignment = Alignment(horizontal='center', vertical='center')
            h18.font = Font(size=12, bold=True)
            col += 1

            h19 = worksheet2.cell(column=col, row=row, value="Start Date")
            h19.alignment = Alignment(horizontal='center', vertical='center')
            h19.font = Font(size=12, bold=True)
            col += 1

            h20 = worksheet2.cell(column=col, row=row, value="End Date")
            h20.alignment = Alignment(horizontal='center', vertical='center')
            h20.font = Font(size=12, bold=True)
            col += 1

            h21 = worksheet2.cell(column=col, row=row, value="NQF\nAligned")
            h21.alignment = Alignment(horizontal='center', vertical='center')
            h21.font = Font(size=12, bold=True)
            col += 1

            h22 = worksheet2.cell(column=col, row=row, value="NQF Level")
            h22.alignment = Alignment(horizontal='center', vertical='center')
            h22.font = Font(size=12, bold=True)
            col += 1

            h23 = worksheet2.cell(column=col, row=row, value="Race")
            h23.alignment = Alignment(horizontal='center', vertical='center')
            h23.font = Font(size=12, bold=True)
            col += 1

            h24 = worksheet2.cell(column=col, row=row, value="Gender")
            h24.alignment = Alignment(horizontal='center', vertical='center')
            h24.font = Font(size=12, bold=True)
            col += 1

            h25 = worksheet2.cell(column=col, row=row, value="Disability")
            h25.alignment = Alignment(horizontal='center', vertical='center')
            h25.font = Font(size=12, bold=True)
            col += 1

            for training_planned_line in self.training_planned_ids:
                row += 1
                col = 1
                for training_planned_data in training_planned_line:

                    if training_planned_data.training_type != 0:
                        worksheet2.cell(column=col, row=row, value=self.get_training_type(
                            training_planned_data.training_type))
                        col += 1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col += 1

                    if training_planned_data.name != 0:
                        worksheet2.cell(
                            column=col, row=row, value=training_planned_data.name)
                        col += 1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col += 1

                    if training_planned_data.surname != 0:
                        worksheet2.cell(
                            column=col, row=row, value=training_planned_data.surname)
                        col += 1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col += 1

                    if training_planned_data.employee_id != 0:
                        worksheet2.cell(
                            column=col, row=row, value=training_planned_data.employee_id)
                        col += 1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col += 1

                    if training_planned_data.code:
                        worksheet2.cell(
                            column=col, row=row, value=training_planned_data.code and training_planned_data.code.name)
                        col += 1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col += 1

                    if training_planned_data.occupation:
                        worksheet2.cell(
                            column=col, row=row, value=training_planned_data.occupation and training_planned_data.occupation.name)
                        col += 1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col += 1

                    if training_planned_data.specialization:
                        worksheet2.cell(
                            column=col, row=row, value=training_planned_data.specialization and training_planned_data.specialization.name)
                        col += 1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col += 1

                    if training_planned_data.city_id:
                        worksheet2.cell(
                            column=col, row=row, value=training_planned_data.city_id and training_planned_data.city_id.name)
                        col += 1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col += 1

                    if training_planned_data.urban != 0:
                        worksheet2.cell(
                            column=col, row=row, value=self.get_urban_rural(training_planned_data.urban))
                        col += 1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col += 1

                    if training_planned_data.learner_province:
                        worksheet2.cell(
                            column=col, row=row, value=training_planned_data.learner_province and training_planned_data.learner_province.name)
                        col += 1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col += 1

                    if training_planned_data.socio_economic_status != 0:
                        worksheet2.cell(column=col, row=row, value=self.get_socio_eco_status(
                            training_planned_data.socio_economic_status))
                        col += 1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col += 1

                    if training_planned_data.type_training:
                        worksheet2.cell(
                            column=col, row=row, value=training_planned_data.type_training and training_planned_data.type_training.name)
                        col += 1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col += 1

                    if training_planned_data.other_type_of_intervention != 0:
                        worksheet2.cell(
                            column=col, row=row, value=training_planned_data.other_type_of_intervention)
                        col += 1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col += 1

                    if training_planned_data.name_training != 0:
                        worksheet2.cell(
                            column=col, row=row, value=training_planned_data.name_training)
                        col += 1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col += 1

                    if training_planned_data.pivotal_programme_qualification != 0:
                        worksheet2.cell(
                            column=col, row=row, value=training_planned_data.pivotal_programme_qualification)
                        col += 1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col += 1

                    if training_planned_data.pivotal_programme_institution != 0:
                        worksheet2.cell(
                            column=col, row=row, value=training_planned_data.pivotal_programme_institution)
                        col += 1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col += 1

                    if training_planned_data.training_cost != 0:
                        worksheet2.cell(
                            column=col, row=row, value=training_planned_data.training_cost)
                        col += 1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col += 1

                    if training_planned_data.start_date != 0:
                        worksheet2.cell(
                            column=col, row=row, value=training_planned_data.start_date)
                        col += 1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col += 1

                    if training_planned_data.end_date != 0:
                        worksheet2.cell(
                            column=col, row=row, value=training_planned_data.end_date)
                        col += 1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col += 1

                    if training_planned_data.nqf_aligned != 0:
                        worksheet2.cell(column=col, row=row, value=self.get_nqf_aligned(
                            training_planned_data.nqf_aligned))
                        col += 1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col += 1

                    if training_planned_data.nqf_level != 0:
                        worksheet2.cell(
                            column=col, row=row, value=self.get_nqf_level(training_planned_data.nqf_level))
                        col += 1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col += 1

                    if training_planned_data.population_group != 0:
                        worksheet2.cell(column=col, row=row, value=self.get_population_group(
                            training_planned_data.population_group))
                        col += 1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col += 1

                    if training_planned_data.gender != 0:
                        worksheet2.cell(
                            column=col, row=row, value=self.get_genders(training_planned_data.gender))
                        col += 1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col += 1

                    if training_planned_data.dissability != 0:
                        worksheet2.cell(column=col, row=row, value=self.get_nqf_aligned(
                            training_planned_data.dissability))
                        col += 1
                    else:
                        worksheet2.cell(column=col, row=row, value='')
                        col += 1

        if self.planned_adult_education_ids:
            row = col = 1
            worksheet3 = workbook.create_sheet(
                title="Adult Education and Training")

            h1 = worksheet3.cell(column=col, row=row, value="First Name")
            h1.alignment = Alignment(horizontal='center', vertical='center')
            h1.font = Font(size=12, bold=True)
            col += 1

            h2 = worksheet3.cell(column=col, row=row, value="Sur Name")
            h2.alignment = Alignment(horizontal='center', vertical='center')
            h2.font = Font(size=12, bold=True)
            col += 1

            h4 = worksheet3.cell(column=col, row=row, value="ID Number")
            h4.alignment = Alignment(horizontal='center', vertical='center')
            h4.font = Font(size=12, bold=True)
            col += 1

            h5 = worksheet3.cell(column=col, row=row, value="Population Group")
            h5.alignment = Alignment(horizontal='center', vertical='center')
            h5.font = Font(size=12, bold=True)
            col += 1

            h6 = worksheet3.cell(column=col, row=row, value="Gender")
            h6.alignment = Alignment(horizontal='center', vertical='center')
            h6.font = Font(size=12, bold=True)
            col += 1

            h7 = worksheet3.cell(
                column=col, row=row, value="Disability Status And Type")
            h7.alignment = Alignment(horizontal='center', vertical='center')
            h7.font = Font(size=12, bold=True)
            col += 1

            h8 = worksheet3.cell(column=col, row=row, value="City")
            h8.alignment = Alignment(horizontal='center', vertical='center')
            h8.font = Font(size=12, bold=True)
            col += 1

            h9 = worksheet3.cell(column=col, row=row, value="Learner Province")
            h9.alignment = Alignment(horizontal='center', vertical='center')
            h9.font = Font(size=12, bold=True)
            col += 1

            h10 = worksheet3.cell(column=col, row=row, value="Urban/Rural")
            h10.alignment = Alignment(horizontal='center', vertical='center')
            h10.font = Font(size=12, bold=True)
            col += 1

            h11 = worksheet3.cell(column=col, row=row, value="AET Start Date")
            h11.alignment = Alignment(horizontal='center', vertical='center')
            h11.font = Font(size=12, bold=True)
            col += 1

            h12 = worksheet3.cell(column=col, row=row, value="AET End Date")
            h12.alignment = Alignment(horizontal='center', vertical='center')
            h12.font = Font(size=12, bold=True)
            col += 1

            h13 = worksheet3.cell(column=col, row=row, value="Provider")
            h13.alignment = Alignment(horizontal='center', vertical='center')
            h13.font = Font(size=12, bold=True)
            col += 1

            h14 = worksheet3.cell(column=col, row=row, value="AET Level")
            h14.alignment = Alignment(horizontal='center', vertical='center')
            h14.font = Font(size=12, bold=True)
            col += 1

            h15 = worksheet3.cell(column=col, row=row, value="AET Subject")
            h15.alignment = Alignment(horizontal='center', vertical='center')
            h15.font = Font(size=12, bold=True)
            col += 1

            for planned_adult_education_line in self.planned_adult_education_ids:
                row += 1
                col = 1
                for planned_adult_education_data in planned_adult_education_line:

                    if planned_adult_education_data.name != 0:
                        worksheet3.cell(
                            column=col, row=row, value=planned_adult_education_data.name)
                        col += 1
                    else:
                        worksheet3.cell(column=col, row=row, value='')
                        col += 1

                    if planned_adult_education_data.surname != 0:
                        worksheet3.cell(
                            column=col, row=row, value=planned_adult_education_data.surname)
                        col += 1
                    else:
                        worksheet3.cell(column=col, row=row, value='')
                        col += 1

                    if planned_adult_education_data.id_number != 0:
                        worksheet3.cell(
                            column=col, row=row, value=planned_adult_education_data.id_number)
                        col += 1
                    else:
                        worksheet3.cell(column=col, row=row, value='')
                        col += 1

                    if planned_adult_education_data.population_group != 0:
                        worksheet3.cell(column=col, row=row, value=self.get_population_group(
                            planned_adult_education_data.population_group))
                        col += 1
                    else:
                        worksheet3.cell(column=col, row=row, value='')
                        col += 1

                    if planned_adult_education_data.gender != 0:
                        worksheet3.cell(column=col, row=row, value=self.get_genders(
                            planned_adult_education_data.gender))
                        col += 1
                    else:
                        worksheet3.cell(column=col, row=row, value='')
                        col += 1

                    if planned_adult_education_data.dissability_status_and_type != 0:
                        worksheet3.cell(column=col, row=row, value=self.get_disability_status(
                            planned_adult_education_data.dissability_status_and_type))
                        col += 1
                    else:
                        worksheet3.cell(column=col, row=row, value='')
                        col += 1

                    if planned_adult_education_data.city_id:
                        worksheet3.cell(
                            column=col, row=row, value=planned_adult_education_data.city_id and planned_adult_education_data.city_id.name)
                        col += 1
                    else:
                        worksheet3.cell(column=col, row=row, value='')
                        col += 1

                    if planned_adult_education_data.province:
                        worksheet3.cell(
                            column=col, row=row, value=planned_adult_education_data.province and planned_adult_education_data.province.name)
                        col += 1
                    else:
                        worksheet3.cell(column=col, row=row, value='')
                        col += 1

                    if planned_adult_education_data.urban != 0:
                        worksheet3.cell(column=col, row=row, value=self.get_urban_rural(
                            planned_adult_education_data.urban))
                        col += 1
                    else:
                        worksheet3.cell(column=col, row=row, value='')
                        col += 1

                    if planned_adult_education_data.start_date != 0:
                        worksheet3.cell(
                            column=col, row=row, value=planned_adult_education_data.start_date)
                        col += 1
                    else:
                        worksheet3.cell(column=col, row=row, value='')
                        col += 1

                    if planned_adult_education_data.end_date != 0:
                        worksheet3.cell(
                            column=col, row=row, value=planned_adult_education_data.end_date)
                        col += 1
                    else:
                        worksheet3.cell(column=col, row=row, value='')
                        col += 1

                    if planned_adult_education_data.provider != 0:
                        worksheet3.cell(
                            column=col, row=row, value=planned_adult_education_data.provider)
                        col += 1
                    else:
                        worksheet3.cell(column=col, row=row, value='')
                        col += 1

                    if planned_adult_education_data.aet_level != 0:
                        worksheet3.cell(column=col, row=row, value=self.get_aet_level(
                            planned_adult_education_data.aet_level))
                        col += 1
                    else:
                        worksheet3.cell(column=col, row=row, value='')
                        col += 1

                    if planned_adult_education_data.aet_subject != 0:
                        worksheet3.cell(column=col, row=row, value=self.get_aet_subject(
                            planned_adult_education_data.aet_subject))
                        col += 1
                    else:
                        worksheet3.cell(column=col, row=row, value='')
                        col += 1

        if self.scarce_and_critical_skills_ids:
            row = col = 1
            worksheet4 = workbook.create_sheet(title="Vacancies Hard to Fill")

            h1 = worksheet4.cell(column=col, row=row, value="OFO Code")
            h1.alignment = Alignment(horizontal='center', vertical='center')
            h1.font = Font(size=12, bold=True)
            col += 1

            h2 = worksheet4.cell(column=col, row=row, value="Occupation")
            h2.alignment = Alignment(horizontal='center', vertical='center')
            h2.font = Font(size=12, bold=True)
            col += 1

            h4 = worksheet4.cell(column=col, row=row, value="Specialisation")
            h4.alignment = Alignment(horizontal='center', vertical='center')
            h4.font = Font(size=12, bold=True)
            col += 1

            h5 = worksheet4.cell(column=col, row=row, value="Province")
            h5.alignment = Alignment(horizontal='center', vertical='center')
            h5.font = Font(size=12, bold=True)
            col += 1

            h6 = worksheet4.cell(
                column=col, row=row, value="Number of Vacancies")
            h6.alignment = Alignment(horizontal='center', vertical='center')
            h6.font = Font(size=12, bold=True)
            col += 1

            h7 = worksheet4.cell(column=col, row=row, value="Gender")
            h7.alignment = Alignment(horizontal='center', vertical='center')
            h7.font = Font(size=12, bold=True)
            col += 1

            h8 = worksheet4.cell(column=col, row=row, value="Race")
            h8.alignment = Alignment(horizontal='center', vertical='center')
            h8.font = Font(size=12, bold=True)
            col += 1

            h9 = worksheet4.cell(
                column=col, row=row, value="Number of months\nposition has")
            h9.alignment = Alignment(horizontal='center', vertical='center')
            h9.font = Font(size=12, bold=True)
            col += 1

            h10 = worksheet4.cell(column=col, row=row, value="Comments")
            h10.alignment = Alignment(horizontal='center', vertical='center')
            h10.font = Font(size=12, bold=True)

            for scarce_and_critical_skills_line in self.scarce_and_critical_skills_ids:
                row += 1
                col = 1
                for scarce_and_critical_skills_data in scarce_and_critical_skills_line:

                    if scarce_and_critical_skills_data.ofo_code:
                        worksheet4.cell(
                            column=col, row=row, value=scarce_and_critical_skills_data.ofo_code and scarce_and_critical_skills_data.ofo_code.name)
                        col += 1
                    else:
                        worksheet4.cell(column=col, row=row, value='')
                        col += 1

                    if scarce_and_critical_skills_data.occupation:
                        worksheet4.cell(
                            column=col, row=row, value=scarce_and_critical_skills_data.occupation and scarce_and_critical_skills_data.occupation.name)
                        col += 1
                    else:
                        worksheet4.cell(column=col, row=row, value='')
                        col += 1

                    if scarce_and_critical_skills_data.specialization:
                        worksheet4.cell(
                            column=col, row=row, value=scarce_and_critical_skills_data.specialization and scarce_and_critical_skills_data.specialization.name)
                        col += 1
                    else:
                        worksheet4.cell(column=col, row=row, value='')
                        col += 1

                    if scarce_and_critical_skills_data.province:
                        worksheet4.cell(
                            column=col, row=row, value=scarce_and_critical_skills_data.province and scarce_and_critical_skills_data.province.name)
                        col += 1
                    else:
                        worksheet4.cell(column=col, row=row, value='')
                        col += 1

                    if scarce_and_critical_skills_data.number_of_vacancies != 0:
                        worksheet4.cell(
                            column=col, row=row, value=scarce_and_critical_skills_data.number_of_vacancies)
                        col += 1
                    else:
                        worksheet4.cell(column=col, row=row, value='')
                        col += 1

                    if scarce_and_critical_skills_data.gender != 0:
                        worksheet4.cell(column=col, row=row, value=self.get_genders(
                            scarce_and_critical_skills_data.gender))
                        col += 1
                    else:
                        worksheet4.cell(column=col, row=row, value='')
                        col += 1

                    if scarce_and_critical_skills_data.population_group != 0:
                        worksheet4.cell(column=col, row=row, value=self.get_population_group(
                            scarce_and_critical_skills_data.population_group))
                        col += 1
                    else:
                        worksheet4.cell(column=col, row=row, value='')
                        col += 1

                    if scarce_and_critical_skills_data.no_of_months != 0:
                        worksheet4.cell(
                            column=col, row=row, value=scarce_and_critical_skills_data.no_of_months)
                        col += 1
                    else:
                        worksheet4.cell(column=col, row=row, value='')
                        col += 1

                    if scarce_and_critical_skills_data.comments != 0:
                        worksheet4.cell(
                            column=col, row=row, value=scarce_and_critical_skills_data.comments)
                        col += 1
                    else:
                        worksheet4.cell(column=col, row=row, value='')
                        col += 1

        workbook.save(buffer)
        out_data = base64.encodestring(buffer.getvalue())

        attachment_obj = self.env['ir.attachment']
        new_attach = attachment_obj.create({
            'name': 'wsp.xlsx',
            'res_name': 'wsp_data',
            'type': 'binary',
            'res_model': 'wsp.plan',
            'datas': out_data,
        })

        return {
            'type': 'ir.actions.act_url',
            'url': '/web/binary/saveas?model=ir.attachment&field=datas&filename_field=name&id=%s' % (new_attach.id, ),
            'target': 'self',
        }
        return True

    
    def check_variance(self):
        return True

#     
#     def action_request_extension(self):
#         email_template_obj = self.env['email.template']
#         ir_model_data_obj = self.env['ir.model.data']
#         bronwen_id = ir_model_data_obj.get_object_reference('hwseta_sdp', 'email_template_wsp_request_for_extension_bronwen')
#         email_template_obj.send_mail(self.env.cr, self.env.uid, bronwen_id[1], self.id,force_send=True,context=self.env.context)
#         luyanda_id = ir_model_data_obj.get_object_reference('hwseta_sdp', 'email_template_wsp_request_for_extension_luyanda')
#         email_template_obj.send_mail(self.env.cr, self.env.uid, luyanda_id[1], self.id,force_send=True,context=self.env.context)
#         self.write({'request_extension':False,'allow_extension':True})
#         return True

#     
#     def action_allow_extension(self):
#         #         email_template_obj = self.env['email.template']
#         #         ir_model_data_obj = self.env['ir.model.data']
#         #         bronwen_id = ir_model_data_obj.get_object_reference('hwseta_sdp', 'email_template_wsp_approval_for_extension_bronwen')
#         #         email_template_obj.send_mail(self.env.cr, self.env.uid, bronwen_id[1], self.id,force_send=True,context=self.env.context)
#         #         luyanda_id = ir_model_data_obj.get_object_reference('hwseta_sdp', 'email_template_wsp_request_for_extension_luyanda')
#         #         email_template_obj.send_mail(self.env.cr, self.env.uid, luyanda_id[1], self.id,force_send=True,context=self.env.context)
#         self.write(
#             {'allow_extension': False, 'from_extension': False, 'extension_allowed': True})
#         return True

    # 
    # def action_reject_extension(self):
    #     self.write({'state': 'rejected'})
    #     return True

    
    def copy(self):
        ''' Inherited to avoid duplicating records '''
        raise Warning(
            _('Forbidden to Duplicate! Is not possible to duplicate the record, please create the new one.!'))
        return super(wsp_plan, self).copy()

    
    def unlink(self):
        if self.state != 'draft':
            raise Warning(
                _('Forbidden to Delete! Is not possible to delete the record in %s state') % (self.state))
        return super(wsp_plan, self).unlink()

    @api.model
    def create(self, vals):
        #        res = super(wsp_plan, self).create(vals)
        #        for planned_training_data in res.training_planned_ids :
        #            if planned_training_data.start_date and planned_training_data.end_date :
        #                msg = 'Planned Training '+(planned_training_data.name and planned_training_data.surname and 'for '+planned_training_data.name+' '+planned_training_data.surname or '')
        #                self.validate_date(planned_training_data.start_date , planned_training_data.end_date ,msg)
        
        #check for hard to fills yes no
        dbg("BOOLMAN:"+str(vals['vacancy_hard_to_fill_bool']))
        if vals['vacancy_hard_to_fill_bool'] == 'yes':
            if vals['scarce_and_critical_skills_ids']==[]:
                raise Warning(_("You indicated that you have vacancies that are hard to fill, but did not add any lines to the scarce and critical skills list!"))
            dbg("VHTFYES:"+str(vals))
        if vals['vacancy_hard_to_fill_bool'] not in ('yes', 'no'):
            raise Warning(_("You must select yes or no to whether you have vacancies that are hard to fill in order to submit the WSP."))

        admin_config_data = self.env['leavy.income.config'].search([])
        config_wsp_end_date = datetime.strptime(
            admin_config_data.wsp_end_date, '%Y-%m-%d').date()
        current_date = datetime.now().date()
        if config_wsp_end_date and (current_date > config_wsp_end_date) and self._uid not in [1, 3830, 3815]:
            raise Warning(_('WSP end date( %s ) for currect year is expried, you are not allowed to create new WSP') % (
                config_wsp_end_date))
        fiscal_year = ''
        current_date = datetime.now().date()
        fiscal_year_data = self.env['account.fiscalyear'].search(
            [('date_start', '<=', current_date), ('date_stop', '>=', current_date)], limit=1)
        if fiscal_year_data:
            fiscal_year = fiscal_year_data.name
        if config_wsp_end_date and (current_date > config_wsp_end_date) and self._uid != 1:
            raise Warning(_('The WSP %s Deadline for Submission has passed %s. Your organisation will not be able to create a WSPATR, please contact the HWSETA for further assistance.') % (
                fiscal_year, config_wsp_end_date))

        # current fiscal year
#         fiscal_year = None
#         current_date = datetime.now().date()
#         fiscal_year_data = self.env['account.fiscalyear'].search([('date_start','<=',current_date),('date_stop','>=',current_date)],limit=1)
#         if fiscal_year_data :
#             fiscal_year = fiscal_year_data.id
        ##
        # Getting WSP Start date, WSP End Date and WSP Extension Date from Admin Configuration.
#         admin_config_data = self.env['leavy.income.config'].search([('wsp_financial_year','=',fiscal_year)])
#         if admin_config_data :
#             vals.update({
#                          'start_period' : admin_config_data.wsp_start_date,
#                          'end_period' : admin_config_data.wsp_end_date,
# #                          'ext_period' : admin_config_data.wsp_extension_date,
# #                          'ext_period1' : admin_config_data.wsp_extension_date,
#                          })
        # Validations for avoiding WSP creation if WSP for the same
        # sdf,employer and financial year already created in the system
        admin_config_data = self.env['leavy.income.config'].search([], limit=1)
        if admin_config_data:
            fy_data = admin_config_data.wsp_financial_year and admin_config_data.wsp_financial_year
        #wsp_exist_check_data = self.env['wsp.plan'].search([('sdf_id','=',vals.get('sdf_id',False)),('employer_id','=',vals.get('employer_id',False)),('fiscal_year','=',fy_data.id)])
        # There should be only one wsp for one employer in one fiscal year
        wsp_exist_check_data = self.env['wsp.plan'].search(
            [('employer_id', '=', vals.get('employer_id', False)), ('fiscal_year', '=', fy_data.id)])
        sdf_name = self.env['hr.employee'].browse(
            vals.get('sdf_id', False)).name
        emp_name = self.env['res.partner'].browse(
            vals.get('employer_id', False)).name
        if wsp_exist_check_data:
            #raise Warning(_('WSP for SDF %s and Employer %s for the year %s is already exist!')%(sdf_name,emp_name,fy_data.name))
            raise Warning(_('WSP for Employer %s for the year %s is already exist!') % (
                emp_name, fy_data.name))
        ##
        vals['name'] = self.env['ir.sequence'].get('wsp.plan.reference')
        res = super(wsp_plan, self).create(vals)
        # Validating start and end date while creating WSP.
        ofo_object = self.env['ofo.code']
        if vals.get('training_actual_ids', False):
            for train_actual_list in vals['training_actual_ids']:
                #                 if train_actual_list[2].get('start_date',False) and train_actual_list[2].get('end_date',False) :
                #                     msg = 'Actual Training '+ (train_actual_list[2].get('code', False) and 'for OFO '+ofo_object.browse(train_actual_list[2]['code']).name or '')
                #                     self.validate_date(train_actual_list[2]['start_date'],train_actual_list[2]['end_date'],msg)
                # Validating Age total with total male, total female total
                #                 total_races = train_actual_list[2].get('african_male',0) + train_actual_list[2].get('african_female',0) + train_actual_list[2].get('african_dissabled',0) + train_actual_list[2].get('coloured_male',0) + train_actual_list[2].get('coloured_female',0) + train_actual_list[2].get('coloured_dissabled',0) + train_actual_list[2].get('white_male',0) + train_actual_list[2].get('white_female',0) + train_actual_list[2].get('white_dissabled',0) + train_actual_list[2].get('indian_male',0) + train_actual_list[2].get('indian_female',0) + train_actual_list[2].get('indian_dissabled',0)
                total_races = train_actual_list[2].get('african_male', 0) + train_actual_list[2].get('african_female', 0) + train_actual_list[2].get('coloured_male', 0) + train_actual_list[2].get(
                    'coloured_female', 0) + train_actual_list[2].get('white_male', 0) + train_actual_list[2].get('white_female', 0) + train_actual_list[2].get('indian_male', 0) + train_actual_list[2].get('indian_female', 0)
                total_ages = train_actual_list[2].get('age_group_less', 0) + train_actual_list[
                    2].get('age_group_greater', 0) + train_actual_list[2].get('age_group_upto', 0)
                if total_ages > total_races:
                    raise Warning(_('Age total should not be greater than total of races %s in Actual Training') % (
                        train_actual_list[2].get('code') and 'for OFO ' + ofo_object.browse(train_actual_list[2]['code']).name or ''))

                # Validations on occupation and code
                if train_actual_list[2].get('occupation', ''):
                    ofo_code_data = self.env['ofo.code'].search(
                        [('occupation', '=', train_actual_list[2].get('occupation', ''))])
                    if len(ofo_code_data) > 1:
                        if train_actual_list[2].get('code', '') != '':
                            if train_actual_list[2].get('code', 0) != ofo_code_data[0].id:
                                raise Warning(_('Wrong occupation %s in Actual Training') % (train_actual_list[2].get(
                                    'code') and 'for OFO ' + ofo_object.browse(train_actual_list[2]['code']).name or ''))
                    else:
                        if train_actual_list[2].get('code', '') != '':
                            if train_actual_list[2].get('code', 0) != ofo_code_data.id:
                                raise Warning(_('Wrong occupation %s in Actual Training') % (train_actual_list[2].get(
                                    'code') and 'for OFO ' + ofo_object.browse(train_actual_list[2]['code']).name or ''))

                # Validations on social economic status
                if train_actual_list[2].get('training_type', '') == 'non-pivotal' and train_actual_list[2].get('socio_economic_status', '') == 'unemployed':
                    raise Warning(_('Socio Economic Status should not be Unemployed %s in Actual Training, If Non pivotal is selected!!') % (
                        train_actual_list[2].get('code') and 'for OFO ' + ofo_object.browse(train_actual_list[2]['code']).name or ''))

                # Validating total male, total female should not grater than
                # total dissabled
                if train_actual_list[2].get('african_dissabled', 0) > train_actual_list[2].get('african_male', 0) + train_actual_list[2].get('african_female', 0):
                    raise Warning(_('In actual training AD should not be greater than total of AM + AF %s') % (
                        train_actual_list[2].get('code') and 'for OFO ' + ofo_object.browse(train_actual_list[2]['code']).name or ''))
                if train_actual_list[2].get('coloured_dissabled', 0) > train_actual_list[2].get('coloured_male', 0) + train_actual_list[2].get('coloured_female', 0):
                    raise Warning(_('In actual training CD should not be greater than total of CM + CF %s') % (
                        train_actual_list[2].get('code') and 'for OFO ' + ofo_object.browse(train_actual_list[2]['code']).name or ''))
                if train_actual_list[2].get('white_dissabled', 0) > train_actual_list[2].get('white_male', 0) + train_actual_list[2].get('white_female', 0):
                    raise Warning(_('In actual training WD should not be greater than total of WM + wF %s') % (
                        train_actual_list[2].get('code') and 'for OFO ' + ofo_object.browse(train_actual_list[2]['code']).name or ''))
                if train_actual_list[2].get('indian_dissabled', 0) > train_actual_list[2].get('indian_male', 0) + train_actual_list[2].get('indian_female', 0):
                    raise Warning(_('In actual training ID should not be greater than total of IM + IF %s') % (
                        train_actual_list[2].get('code') and 'for OFO ' + ofo_object.browse(train_actual_list[2]['code']).name or ''))
#         else:
#             raise Warning(_('Please enter Actual Training data!'))

#         if vals.get('actual_adult_education_ids',False):
#             pass
#         else:
#             raise Warning(_('Please enter Actual Adult Education data!'))
        if vals.get('actual_adult_education_ids', False):
            for actual_adult_list in vals['actual_adult_education_ids']:
                #                 if actual_adult_list[2].get('start_date',False) and actual_adult_list[2].get('end_date',False) :
                #                     msg = 'Adult Education and Training '+ (actual_adult_list[2].get('name', False) and actual_adult_list[2].get('surname', False) and 'for '+actual_adult_list[2]['name']+' '+actual_adult_list[2]['surname'] or '')
                #                     self.validate_date(actual_adult_list[2]['start_date'],actual_adult_list[2]['end_date'],msg)
                # Validations on Name and Surname
                if actual_adult_list[2].get('name', False):
                    if actual_adult_list[2]['name'].isdigit():
                        raise Warning(_('Name should not be numeric: %s in Adult Education Training!') % (
                            actual_adult_list[2]['name']))
                if actual_adult_list[2].get('surname', False):
                    if actual_adult_list[2]['surname'].isdigit():
                        raise Warning(_('Surname should not be numeric: %s! in Adult Education Training') % (
                            actual_adult_list[2]['surname']))
        # Validations for Employee ID
        if vals.get('total_employment_profile_ids', False):
            for total_employment_data in vals['total_employment_profile_ids']:
                if total_employment_data[2].get('citizen_resident_status_code', False) and (total_employment_data[2]['citizen_resident_status_code'] in ['sa', 'dual']):
                    id_number = total_employment_data[
                        2].get('employee_id', False)
                    # If Citizen Status is Dual and ID Type is Passport Then
                    # allow entering alphanumeric values.
                    if id_number:
                        if total_employment_data[2]['citizen_resident_status_code'] in ['sa', 'dual'] and total_employment_data[2].get('id_type', False) == 'passport':
                            pass
                        else:
                            if not total_employment_data[2]['employee_id'].isdigit():
                                raise Warning(
                                    _('Employee ID %s should be numeric in Total Employment Profile!') % (str(id_number)))
                            else:
                                year = id_number[:2]
                                id_number = id_number[2:]
                                month = id_number[:2]
                                id_number = id_number[2:]
                                day = id_number[:2]
                                if int(month) > 12 or int(month) < 1 or int(day) > 31 or int(day) < 1:
                                    raise Warning(
                                        _('Incorrect Identification Number %s in Total Employment Profile!') % (str(id_number)))
                                else:
                                    # # Calculating last day of month.
                                    x_year = int(year)
                                    if x_year == 00:
                                        x_year = 2000
                                    last_day = calendar.monthrange(
                                        int(x_year), int(month))[1]
                                    if int(day) > last_day:
                                        raise Warning(
                                            _('Incorrect Identification Number %s in Total Employment Profile!') % (str(id_number)))
                # Validations on Name and Surname
                if total_employment_data[2].get('name', False):
                    if total_employment_data[2]['name'].isdigit():
                        raise Warning(_('Name should not be numeric: %s in Total Employment Profile!') % (
                            total_employment_data[2]['name']))
                if total_employment_data[2].get('surname', False):
                    if total_employment_data[2]['surname'].isdigit():
                        raise Warning(_('Surname should not be numeric: %s! in Total Employment Profile!') % (
                            total_employment_data[2]['surname']))

                # Validations on occupation and code
                if total_employment_data[2].get('occupation', ''):
                    ofo_code_data = self.env['ofo.code'].search(
                        [('occupation', '=', total_employment_data[2].get('occupation', ''))])
                    if total_employment_data[2].get('ofo_code', '') != '':
                        if total_employment_data[2].get('ofo_code', 0) != ofo_code_data.id:
                            raise Warning(_('Wrong occupation for OFO code %s in Total Employment Profile!') % (
                                ofo_object.browse(total_employment_data[2]['ofo_code']).name or ''))

        if vals.get('training_planned_ids', False):
            if not vals.get('total_employment_profile_ids', False):
                raise Warning(
                    _('Please enter Total Employment Profile Information!'))
            for planned_train_list in vals['training_planned_ids']:
                #                 if planned_train_list[2].get('start_date',False) and planned_train_list[2].get('end_date',False) :
                #                     msg = 'Planned Training '+ (planned_train_list[2].get('name', False) and planned_train_list[2].get('surname', False) and 'for '+planned_train_list[2]['name']+' '+planned_train_list[2]['surname'] or '')
                #                     self.validate_date(planned_train_list[2]['start_date'], planned_train_list[2]['end_date'] , msg)
                # Validating whether employee present in TEP.
                if planned_train_list[2].get('employee_id', False):
                    #                     count = 0
                    emp_ids = []
#                     emp_status=planned_train_list[2].get('socio_economic_status',False) == 'employed'
                    emp_status = planned_train_list[2].get(
                        'socio_economic_status', False)
                    for tep_data in vals.get('total_employment_profile_ids', False):
                        if tep_data[2].get('employee_id', False):
                            emp_ids.append(tep_data[2]['employee_id'])
                    if not planned_train_list[2]['employee_id'] in emp_ids and emp_status == 'employed':
                        raise Warning(_('Employee with ID %s not in TEP. Invalid Employee!') % (
                            planned_train_list[2]['employee_id']))
#                     if emp_ids:
#                         if planned_train_list[2]['employee_id'] in emp_ids and emp_status == 'employed' :
#                             count += 1
#                         if count == 0 :
#                             raise Warning(_('Employee with ID %s not in TEP. Invalid Employee!')%(planned_train_list[2]['employee_id']))
#                     else :
#                         if emp_status:
#                             raise Warning(_('Please enter Total Employment Profile data!'))
                # Validations on Name and Surname
                if planned_train_list[2].get('name', False):
                    if planned_train_list[2]['name'].isdigit():
                        raise Warning(_('Name should not be numeric: %s in Planned Training!') % (
                            planned_train_list[2]['name']))
                if planned_train_list[2].get('surname', False):
                    if planned_train_list[2]['surname'].isdigit():
                        raise Warning(_('Surname should not be numeric: %s in Planned Training!') % (
                            planned_train_list[2]['surname']))

                # Validations on occupation and code
                if planned_train_list[2].get('occupation', ''):
                    ofo_code_data = self.env['ofo.code'].search(
                        [('occupation', '=', planned_train_list[2].get('occupation', ''))])
                    if planned_train_list[2].get('code', '') != '':
                        if planned_train_list[2].get('code', 0) != ofo_code_data.id:
                            raise Warning(_('Wrong occupation for OFO code %s in Planned Training!') % (
                                ofo_object.browse(planned_train_list[2]['code']).name or ''))

                # Validations on social economic status
                if planned_train_list[2].get('training_type', '') == 'non-pivotal' and planned_train_list[2].get('socio_economic_status', '') == 'unemployed':
                    raise Warning(_('Socio Economic Status should not be Unemployed %s in Planned Training if Non pivotal is selected!') % (
                        planned_train_list[2].get('code') and 'for OFO ' + ofo_object.browse(planned_train_list[2]['code']).name or ''))

#         if vals.get('planned_adult_education_ids',False):
#             for planned_adult_list in vals['planned_adult_education_ids'] :
# #                 if planned_adult_list[2].get('start_date',False) and planned_adult_list[2].get('end_date',False) :
# #                     msg = 'Planned Adult Education and Training '+ (planned_adult_list[2].get('name', False) and planned_adult_list[2].get('surname', False) and 'for '+planned_adult_list[2]['name']+' '+planned_adult_list[2]['surname'] or '')
# #                     self.validate_date(planned_adult_list[2]['start_date'],planned_adult_list[2]['end_date'],msg)
#                 ### Validations on Name and Surname
#                 if planned_adult_list[2].get('name',False) :
#                     if planned_adult_list[2]['name'].isdigit() :
#                         raise Warning(_('Name should not be numeric: %s in Planned Adult Education Training!')%(planned_adult_list[2]['name']))
#                 if planned_adult_list[2].get('surname',False) :
#                     if planned_adult_list[2]['surname'].isdigit() :
#                         raise Warning(_('Surname should not be numeric: %s! in Planned Adult Education Training!')%(planned_adult_list[2]['surname']))
        if vals.get('scarce_and_critical_skills_ids', False):
            for scarce_list in vals['scarce_and_critical_skills_ids']:
                if scarce_list[2].get('no_of_months', 0):
                    pass
                else:
                    raise Warning(
                        '"No of months position has been vacant" should not be 0 in Scarce and critical skills!')
#         if vals.get('training_actual_planned_ids',False):
#             for planned_current_list in vals['training_actual_planned_ids'] :
#                 if planned_current_list[2].get('start_date',False) and planned_current_list[2].get('end_date',False) :
#                     msg = 'Planned Training for 2015-2016 '+ (planned_current_list[2].get('code', False) and 'for OFO '+ofo_object.browse(planned_current_list[2]['code']).name or '')
#                     self.validate_date(planned_current_list[2]['start_date'],planned_current_list[2]['end_date'],msg)
#         if vals.get('actual_planned_adult_education_ids',False):
#             for planned_adult_current_list in vals['actual_planned_adult_education_ids'] :
#                 if planned_adult_current_list[2].get('start_date',False) and planned_adult_current_list[2].get('end_date',False) :
#                     msg = 'Planned Adult Education and Training for 2015-2016 '+ (planned_adult_current_list[2].get('name', False) and planned_adult_current_list[2].get('surname', False) and 'for '+planned_adult_current_list[2]['name']+' '+planned_adult_current_list[2]['surname'] or '')
#                     self.validate_date(planned_adult_current_list[2]['start_date'],planned_adult_current_list[2]['end_date'],'Planned Adult Education and Training for 2015-2016')

#         user=self.env['res.users'].browse(self._uid)
#         if vals.get('employer_id',False):
#             wsp_track={'name':vals.get('name'),'fiscal_year':self._get_default_fiscalyr(),'status':'draft','last_user_evaluated_updated':user.name,'employer_id':vals.get('employer_id')}
#             self.env['wsp.submission.track'].create(wsp_track)

        return res

    

    
    def write(self, vals):
        #check for hard to fills yes no
        if 'vacancy_hard_to_fill_bool' in vals.keys():
            if vals['vacancy_hard_to_fill_bool'] not in ('yes','no'):
                raise Warning(_("You must select yes or no to whether you have vacancies that are hard to fill in order to submit the WSP."))
        res = super(wsp_plan, self).write(vals)
        # Code for Variance Calculation.
        # Getting Previous year planned WSP for the same employer.
        # Flushing existing records
        if self.save_some_record == False:
            if self.variance_ids:
                for variance_data in self.variance_ids:
                    if variance_data.line_validated:
                        continue
                    variance_data.write(
                        {'atr_actual': 0, 'total_cost_actual': 0})
            if self.variance_pivotal_ids:
                for variance_data in self.variance_pivotal_ids:
                    if variance_data.line_validated:
                        continue
                    variance_data.write(
                        {'atr_actual': 0, 'total_cost_actual': 0})
            ##
            prev_fisc_id = None
            wsp_fy_start_date = datetime.strptime(
                self.fiscal_year.date_start, '%Y-%m-%d').date()
            if wsp_fy_start_date:
                for fiscal_year in self.env['account.fiscalyear'].search([]):
                    start_date = datetime.strptime(
                        fiscal_year.date_start, '%Y-%m-%d').date()
                    if start_date.year == (wsp_fy_start_date.year - 1):
                        prev_fisc_id = fiscal_year.id
            prev_wsp = ''
            if prev_fisc_id:
                prev_wsp = self.search(
                    [('employer_id', '=', self.employer_id.id), ('fiscal_year', '=', prev_fisc_id)])
                if len(prev_wsp) > 1:
                    raise Warning(_('System should not have multiple wsp for employer %s') % (
                        self.employer_id.name))
            # Getting number of records per intervention type from last year
            # planned WSP.
            training_last_actual_dict = {}
            training_last_actual_dict_pivotal = {}

            if prev_wsp:
                #             for training_actual_planned in prev_wsp.training_actual_planned_ids :
                #             added for 2017-18 WSP variance calculation
                for training_actual_planned in prev_wsp.training_planned_ids:
                    if training_actual_planned.line_validated:
                        continue
                    type_training = training_actual_planned.type_training
                    if type_training:
                        if type_training.pivotal:
                            no_of_learner = training_actual_planned.total_male + \
                                training_actual_planned.total_female
                            total_cost = training_actual_planned.total_cost
                            if training_last_actual_dict_pivotal.get(type_training.id, {}):
                                training_last_actual_dict_pivotal[type_training.id][
                                    'no_of_learner'] = training_last_actual_dict_pivotal[type_training.id]['no_of_learner'] + no_of_learner
                                training_last_actual_dict_pivotal[type_training.id][
                                    'total_cost'] = training_last_actual_dict_pivotal[type_training.id]['total_cost'] + total_cost
                            else:
                                training_last_actual_dict_pivotal.update(
                                    {type_training.id: {'no_of_learner': no_of_learner, 'total_cost': total_cost}})
                        else:
                            #                     no_of_learner = training_actual_planned.total_male + training_actual_planned.total_female + training_actual_planned.total_dissabled
                            no_of_learner = training_actual_planned.total_male + \
                                training_actual_planned.total_female
                            total_cost = training_actual_planned.total_cost
                            if training_last_actual_dict.get(type_training.id, {}):
                                training_last_actual_dict[type_training.id]['no_of_learner'] = training_last_actual_dict[
                                    type_training.id]['no_of_learner'] + no_of_learner
                                training_last_actual_dict[type_training.id]['total_cost'] = training_last_actual_dict[
                                    type_training.id]['total_cost'] + total_cost
                            else:
                                training_last_actual_dict.update(
                                    {type_training.id: {'no_of_learner': no_of_learner, 'total_cost': total_cost}})
                    training_actual_planned.write({'line_validated': True})

            # Getting number of records per intervention type from current WSP.
            training_actual_dict = {}
            training_actual_dict_pivotal = {}
            for training_actual_data in self.training_actual_ids:
                if training_actual_data.line_validated:
                    continue
                type_training = training_actual_data.type_training
                if type_training:
                    if type_training.pivotal:
                        no_of_learner = training_actual_data.total_male + \
                            training_actual_data.total_female
                        total_cost = training_actual_data.total_cost
                        if training_actual_dict_pivotal.get(type_training.id, {}):
                            training_actual_dict_pivotal[type_training.id]['no_of_learner'] = training_actual_dict_pivotal[
                                type_training.id]['no_of_learner'] + no_of_learner
                            training_actual_dict_pivotal[type_training.id][
                                'total_cost'] = training_actual_dict_pivotal[type_training.id]['total_cost'] + total_cost
                        else:
                            training_actual_dict_pivotal.update(
                                {type_training.id: {'no_of_learner': no_of_learner, 'total_cost': total_cost}})
                    else:
                        #                 no_of_learner = training_actual_data.total_male + training_actual_data.total_female + training_actual_data.total_dissabled
                        no_of_learner = training_actual_data.total_male + \
                            training_actual_data.total_female
                        total_cost = training_actual_data.total_cost
                        if training_actual_dict.get(type_training.id, {}):
                            training_actual_dict[type_training.id]['no_of_learner'] = training_actual_dict[
                                type_training.id]['no_of_learner'] + no_of_learner
                            training_actual_dict[type_training.id]['total_cost'] = training_actual_dict[
                                type_training.id]['total_cost'] + total_cost
                        else:
                            training_actual_dict.update(
                                {type_training.id: {'no_of_learner': no_of_learner, 'total_cost': total_cost}})
                    training_actual_data.write({'line_validated': True})
            # Updating Variance Tab Non Pivotal.
            for variance_data in self.variance_ids:
                dbg(variance_data.type_training.name)
                # Updating last year planned values.
                if variance_data.type_training.id in training_last_actual_dict.keys():
                    variance_data.write({
                        'wsp_planned': training_last_actual_dict[variance_data.type_training.id].get('no_of_learner', 0),
                        'total_cost_planned': training_last_actual_dict[variance_data.type_training.id].get('total_cost', 0)
                    })
                # Updating current year actual values.
                if variance_data.type_training.id in training_actual_dict.keys():
                    variance_data.write({
                        'atr_actual': training_actual_dict[variance_data.type_training.id].get('no_of_learner', 0),
                        'total_cost_actual': training_actual_dict[variance_data.type_training.id].get('total_cost', 0),
                    })
                if variance_data.wsp_planned != 0:
                    dbg('variance percentage')
                    dbg('actual:' + str(variance_data.atr_actual) + '-planned:' + str(variance_data.wsp_planned))
                    dbg((float(variance_data.atr_actual) / float(variance_data.wsp_planned)) * 100)
                    variance_data.write({'variance_percentage': (
                        float(variance_data.atr_actual) / float(variance_data.wsp_planned)) * 100})
                else:
                    dbg('no variance percentage')
                #passed validations
                variance_data.write({'line_validated': True})
            # Updating Variance Tab Pivotal.
            for variance_data in self.variance_pivotal_ids:
                # Updating last year planned values.
                if variance_data.type_training.id in training_last_actual_dict_pivotal.keys():
                    variance_data.write({
                        'wsp_planned': training_last_actual_dict_pivotal[variance_data.type_training.id].get('no_of_learner', 0),
                        'total_cost_planned': training_last_actual_dict_pivotal[variance_data.type_training.id].get('total_cost', 0)
                    })
                # Updating current year actual values.
                if variance_data.type_training.id in training_actual_dict_pivotal.keys():
                    variance_data.write({
                        'atr_actual': training_actual_dict_pivotal[variance_data.type_training.id].get('no_of_learner', 0),
                        'total_cost_actual': training_actual_dict_pivotal[variance_data.type_training.id].get('total_cost', 0),
                    })
                if variance_data.wsp_planned != 0:
                    variance_data.write({'variance_percentage': (
                        float(variance_data.atr_actual) / float(variance_data.wsp_planned)) * 100})
                variance_data.write({'line_validated': True})
            if not self._context.get('load_previous', False):
                # Validating start and end date while editing WSP.
                for actual_data in self.training_actual_ids:
                    if actual_data.line_validated:
                        continue
                    #                     if actual_data.start_date and actual_data.end_date :
                    #                         msg = 'Actual Training '+(actual_data.name and actual_data.surname and 'for '+actual_data.name+' '+actual_data.surname or '')
                    #                         self.validate_date( actual_data.start_date , actual_data.end_date ,msg)

                    # Validations on occupation and code
                    if actual_data.occupation:
                        ofo_code_data = self.env['ofo.code'].search(
                            [('occupation', '=', actual_data.occupation.name)])
                        if actual_data.code:
                            if len(ofo_code_data) > 1:
                                if actual_data.code.id != ofo_code_data[0].id:
                                    raise Warning(_('Wrong occupation %s in Actual Training!') % (
                                        actual_data.code and 'for OFO ' + str(actual_data.code.name) or ''))
                            else:
                                if actual_data.code.id != ofo_code_data.id:
                                    raise Warning(_('Wrong occupation %s in Actual Training!') % (
                                        actual_data.code and 'for OFO ' + str(actual_data.code.name) or ''))
                    # Validations on socio economic status
                    if actual_data.training_type == 'non-pivotal' and actual_data.socio_economic_status == 'unemployed':
                        raise Warning(_('Socio Economic Status should not be Unemployed %s in Actual Training, If Non pivotal is selected!') % (
                            actual_data.code and 'for OFO ' + str(actual_data.code.name) or ''))

                    # Validating total male, total female should not grater
                    # than total dissabled
                    if actual_data.african_dissabled > actual_data.african_male + actual_data.african_female:
                        raise Warning(_('In actual training AD should not be greater than total of AM + AF %s') % (
                            actual_data.code and 'for OFO ' + str(actual_data.code.name) or ''))
                    if actual_data.coloured_dissabled > actual_data.coloured_male + actual_data.coloured_female:
                        raise Warning(_('In actual training CD should not be greater than total of CM + CF %s') % (
                            actual_data.code and 'for OFO ' + str(actual_data.code.name) or ''))
                    if actual_data.white_dissabled > actual_data.white_male + actual_data.white_female:
                        raise Warning(_('In actual training WD should not be greater than total of WM + wF %s') % (
                            actual_data.code and 'for OFO ' + str(actual_data.code.name) or ''))
                    if actual_data.indian_dissabled > actual_data.indian_male + actual_data.indian_female:
                        raise Warning(_('In actual training ID should not be greater than total of IM + IF %s') % (
                            actual_data.code and 'for OFO ' + str(actual_data.code.name) or ''))

#                 for actual_adult_data in self.actual_adult_education_ids :
#                     if actual_adult_data.start_date and actual_adult_data.end_date :
#                         msg = 'Adult Education and Training ' + (actual_adult_data.name and actual_adult_data.surname and 'for '+actual_adult_data.name+' '+actual_adult_data.surname or '')
#                         self.validate_date(actual_adult_data.start_date , actual_adult_data.end_date ,msg)
                    # Validations on Name and Surname
    #                 if actual_adult_data.name and actual_adult_data.name.isdigit():
    #                     raise Warning(_('Name should not be numeric: %s in Adult Education Training!')%(actual_adult_data.name))
    #                 if actual_adult_data.surname and actual_adult_data.surname.isdigit():
    #                     raise Warning(_('Surname should not be numeric: %s! in Adult Education Training')%(actual_adult_data.surname))
                # Validations passed
                    actual_data.write({'line_validated':True})
                # Validations for Employee ID
                for total_employment_data in self.total_employment_profile_ids:
                    if total_employment_data.citizen_resident_status_code in ['sa', 'dual']:
                        id_number = total_employment_data.employee_id
                        if id_number:
                            if total_employment_data.citizen_resident_status_code in ['sa', 'dual'] and total_employment_data.id_type == 'passport':
                                pass
                            else:
                                if not id_number.isdigit():
                                    msg = ''
                                    for x in id_number:
                                        if not x.isdigit():
                                            msg += x + str(type(x)) + '\n'
                                    #raise Warning(_(x + str(type(x))))
                                    dbg(type(id_number))
                                    id_number = id_number.encode('ascii', 'ignore')
                                    id_number = str(id_number)
                                if not id_number.isdigit():
                                    raise Warning(
                                        _('Employee ID should be numeric %s in Total Employment Profile!-' + str(type(id_number))) % (str(id_number)))
                                else:
                                    year = id_number[:2]
                                    id_number = id_number[2:]
                                    month = id_number[:2]
                                    id_number = id_number[2:]
                                    day = id_number[:2]
                                    if int(month) > 12 or int(month) < 1 or int(day) > 31 or int(day) < 1:
                                        raise Warning(
                                            _('Incorrect Identification Number %s in Total Employment Profile!') % (str(id_number)))
                                    else:
                                        # # Calculating last day of month.
                                        x_year = int(year)
                                        if x_year == 00:
                                            x_year = 2000
                                        last_day = calendar.monthrange(
                                            int(x_year), int(month))[1]
                                        if int(day) > last_day:
                                            raise Warning(
                                                _('Incorrect Identification Number %s in Total Employment Profile!') % (str(id_number)))
                    # Validations on Name and Surname
                    if total_employment_data.name and total_employment_data.name.isdigit():
                        raise Warning(_('Name should not be numeric: %s in Total Employment Profile!') % (
                            total_employment_data.name))
                    if total_employment_data.surname and total_employment_data.surname.isdigit():
                        raise Warning(_('Surname should not be numeric: %s! in Total Employment Profile!') % (
                            total_employment_data.surname))

                    # Validations on occupation and code
                    if total_employment_data.occupation:
                        dbg(total_employment_data.occupation)
                        ofo_code_data = self.env['ofo.code'].search(
                            [('occupation', '=', total_employment_data.occupation.name)])
                        ofo_list = [x.id for x in ofo_code_data]
                        if total_employment_data.ofo_code:
                            # if total_employment_data.ofo_code.id != ofo_code_data.id:
                            if total_employment_data.ofo_code.id not in ofo_list:
                                raise Warning(_('Wrong occupation for OFO code %s in Total Employment Profile!') % (
                                    self.env['ofo.code'].browse(total_employment_data.ofo_code.id).name or ''))

                    total_employment_data.write({'line_validated':True})
                for planned_training_data in self.training_planned_ids:
                    if planned_training_data.line_validated:
                        continue
                    if planned_training_data.start_date and planned_training_data.end_date:
                        msg = 'Planned Training ' + \
                            (planned_training_data.name and planned_training_data.surname and 'for ' +
                             planned_training_data.name + ' ' + planned_training_data.surname or '')
#                         self.validate_date(planned_training_data.start_date , planned_training_data.end_date ,msg)
                    # Validations on Name and Surname
                    if planned_training_data.name and planned_training_data.name.isdigit():
                        raise Warning(_('Name should not be numeric: %s in Planned Training!') % (
                            planned_training_data.name))
                    if planned_training_data.surname and planned_training_data.surname.isdigit():
                        raise Warning(_('Surname should not be numeric: %s! in Planned Training!') % (
                            planned_training_data.surname))
                    # Validating whether employee present in TEP.
                    if planned_training_data.employee_id:
                        #                     count = 0
                        #                     emp_status=planned_training_data.socio_economic_status == 'employed'
                        emp_ids = [
                            tep_data.employee_id for tep_data in self.total_employment_profile_ids]
                        if emp_ids and (not planned_training_data.employee_id in emp_ids) and planned_training_data.socio_economic_status == 'employed':
                            raise Warning(_('Employee with ID %s not present TEP. Invalid Employee!') % (
                                planned_training_data.employee_id))
    #                     if emp_ids :
    #                         if planned_training_data.employee_id in emp_ids and planned_training_data.socio_economic_status == 'employed':
    #                             count += 1
    #                         if count == 0:
    #                             raise Warning(_('Employee with ID %s not present TEP. Invalid Employee!')%(planned_training_data.employee_id))
    #                     else :
    #                         if emp_status:
    #                             raise Warning(_('Please enter Total Employment Profile data!'))

                    # Validations on occupation and code
                    if planned_training_data.occupation:
                        ofo_code_data = self.env['ofo.code'].search(
                            [('occupation', '=', planned_training_data.occupation.name)])
                        ofo_list = [x.id for x in ofo_code_data]
                        if planned_training_data.code:
                            # if planned_training_data.code.id != ofo_code_data.id:
                            if planned_training_data.code.id not in ofo_list:
                                raise Warning(_('Wrong occupation for OFO code %s in Planned Training') % (
                                    self.env['ofo.code'].browse(planned_training_data.code.id).name or ''))

                    # Validations on socio economic status
                    if planned_training_data.training_type == 'non-pivotal' and planned_training_data.socio_economic_status == 'unemployed':
                        raise Warning(_('Socio Economic Status should not be Unemployed %s in Planned Training, If Non pivotal is selected!!') % (
                            planned_training_data.code and 'for OFO ' + str(planned_training_data.code.name) or ''))
                    #passed validations, set it to true
                    planned_training_data.write({'line_validated': True})

                for planned_adult_data in self.planned_adult_education_ids:
                    if planned_adult_data.line_validated:
                        continue
                    if planned_adult_data.start_date and planned_adult_data.end_date:
                        msg = 'Planned Adult Education and Training ' + \
                            (planned_adult_data.name and planned_adult_data.surname and 'for ' +
                             planned_adult_data.name + ' ' + planned_adult_data.surname or '')
#                         self.validate_date(planned_adult_data.start_date , planned_adult_data.end_date ,msg)
                    # Validations on Name and Surname
                    if planned_adult_data.name and planned_adult_data.name.isdigit():
                        raise Warning(_('Name should not be numeric: %s in Planned Adult Education Training!') % (
                            planned_adult_data.name))
                    if planned_adult_data.surname and planned_adult_data.surname.isdigit():
                        raise Warning(_('Surname should not be numeric: %s! in Planned Adult Education Training!') % (
                            planned_adult_data.surname))
                    planned_adult_data.write({'line_validated':True})
                for actual_current_data in self.training_actual_planned_ids:
                    if actual_current_data.line_validated:
                        continue
                    if actual_current_data.start_date and actual_current_data.end_date:
                        msg = 'Planned Training for 2017-2018 ' + \
                            (actual_current_data.code and 'for OFO ' +
                             actual_current_data.code.name or '')
                        self.validate_date(
                            actual_current_data.start_date, actual_current_data.end_date, msg)
                    actual_current_data.write({'line_validated':True})
                for actual_current_adult in self.actual_planned_adult_education_ids:
                    if actual_current_adult.line_validated:
                        continue
                    if actual_current_adult.start_date and actual_current_adult.end_date:
                        msg = 'Planned Adult Education and Training for 2017-2018 ' + \
                            (actual_current_adult.name and actual_current_adult.surname and 'for ' +
                             actual_current_adult.name + ' ' + actual_current_adult.surname or '')
                        self.validate_date(
                            actual_current_adult.start_date, actual_current_adult.end_date, msg)
                    actual_current_adult.write({'line_validated':True})
                for scarce_data in self.scarce_and_critical_skills_ids:
                    if scarce_data.no_of_months == 0:
                        raise Warning(
                            '"No of months position has been vacant" should not be 0 in Scarce and Critical Skills!')
                    scarce_data.write({'line_validated':True})
        return res

    
    def generate_wsp_error_log(self):
        '''This method is used to generate error log for WSP'''
        msg = ''
        if not self.save_some_record:
            raise Warning("Please check Partial Save!")
        ##########Total Employment Profile##############
        total_employment_list = []
        self._cr.execute("select id, name, surname, citizen_resident_status_code, employee_id, id_type,\
               ofo_code, occupation from total_employment_profile_fields where total_employment_wsp_id = %s",(self.id,))
        total_employment_profile_ids = self._cr.fetchall()
        for total_employment_data in total_employment_profile_ids:
            # Validations for Employee ID
            if total_employment_data[3] in ['sa', 'dual']:
                id_number = total_employment_data[4]
                if id_number:
                    if total_employment_data[3] in ['sa', 'dual'] and total_employment_data[5] == 'passport':
                        pass
                    else:
                        if not id_number.isdigit():
                            msg = ''
                            for x in id_number:
                                if not x.isdigit():
                                    msg += x + str(type(x)) + '\n'
                            #raise Warning(_(x + str(type(x))))
                            dbg(type(id_number))
                            id_number = id_number.encode('ascii', 'ignore')
                            id_number = str(id_number)
                        if not id_number.isdigit():
                            msg += 'Employee ID should be numeric <b>' + \
                                str(id_number) + \
                                '</b> Total Employment Profile!='
                            total_employment_list.append(total_employment_data[0])
                        else:
                            year = id_number[:2]
                            id_number = id_number[2:]
                            month = id_number[:2]
                            id_number = id_number[2:]
                            day = id_number[:2]
                            if int(month) > 12 or int(month) < 1 or int(day) > 31 or int(day) < 1:
                                msg += 'Incorrect Identification Number <b>' + \
                                    str(id_number) + \
                                    '</b> in Total Employment Profile!='
                                total_employment_list.append(total_employment_data[0])
                            else:
                                # # Calculating last day of month.
                                x_year = int(year)
                                if x_year == 00:
                                    x_year = 2000
                                last_day = calendar.monthrange(
                                    int(x_year), int(month))[1]
                                if int(day) > last_day:
                                    msg += 'Incorrect Identification Number <b>' + \
                                        str(id_number) + \
                                        '</b> in Total Employment Profile!='
                                    total_employment_list.append(total_employment_data[0])
            # Validations on Name and Surname
            if total_employment_data[1] and total_employment_data[1].isdigit():
                msg += 'Name should not be numeric: <b>' + \
                    str(total_employment_data[1]) + \
                    '</b> in Total Employment Profile!='
                total_employment_list.append(total_employment_data[1])
            if total_employment_data[2] and total_employment_data[2].isdigit():
                msg += 'Surname should not be numeric: <b>' + \
                    str(total_employment_data[2]) + \
                    '</b> in Total Employment Profile!='
                total_employment_list.append(total_employment_data[2])
            # Validations on occupation and code
            if total_employment_data[6] and total_employment_data[7]:
                self._cr.execute('select id from ofo_code where occupation = %s',(total_employment_data[7],))
                ofo_occupation_id = self._cr.fetchone()
                if ofo_occupation_id:
                    if total_employment_data[6] != ofo_occupation_id[0]:
                        msg += 'Wrong occupation for OFO code <b>' + str(self.env['ofo.code'].browse(
                            total_employment_data[6]).name or '') + '</b> in Total Employment Profile!='
                        total_employment_list.append(total_employment_data[0])

        #########Planned Training###########
        planned_training_list = []
        self._cr.execute("select id, training_type, name, surname, employee_id,\
            socio_economic_status, code, occupation from planned_training_fields where planned_training_non_wsp_id = %s",(self.id,))
        training_planned_ids = self._cr.fetchall()
        for planned_training_data in training_planned_ids:
            # Validations on Name and Surname
            if planned_training_data[2] and planned_training_data[2].isdigit():
                msg += 'Name should not be numeric: <b>' + \
                    str(planned_training_data[2]) + \
                    '</b> in Planned Training!='
                planned_training_list.append(planned_training_data[0])
            if planned_training_data[3] and planned_training_data[3].isdigit():
                msg += 'Surname should not be numeric: <b>' + \
                    str(planned_training_data[3]) + \
                    '</b> in Planned Training!='
                planned_training_list.append(planned_training_data[0])
            # Validating whether employee present in TEP.
            if planned_training_data[4]:
                self._cr.execute("select employee_id from total_employment_profile_fields where total_employment_wsp_id = %s",(self.id,))
                emp_ids = map(lambda x:x[0], self._cr.fetchall())
                if emp_ids and (not planned_training_data[4] in emp_ids) and planned_training_data[5] == 'employed':
                    msg += 'Employee with ID : <b>' + \
                        str(planned_training_data[4]) + \
                        '</b> not present TEP. Invalid Employee!='
                    planned_training_list.append(planned_training_data[0])
            # Validations on occupation and code
            if planned_training_data[6] and planned_training_data[7]:
                self._cr.execute('select id from ofo_code where occupation = %s',(planned_training_data[7],))
                ofo_occupation_id = self._cr.fetchone()
                if ofo_occupation_id:
                    if planned_training_data[6] != ofo_occupation_id[0]:
                        msg += 'Wrong occupation for OFO code <b>' + str(self.env['ofo.code'].browse(
                            planned_training_data[6]).name or '') + '</b> in Planned Training!='
                        planned_training_list.append(planned_training_data[0])
            # Validations on socio economic status
            if planned_training_data[1] == 'non-pivotal' and planned_training_data[5] == 'unemployed':
                msg += 'Socio Economic Status should not be Unemployed ' + str(planned_training_data[6] and 'for OFO <b>' + str(self.env['ofo.code'].browse(
                            planned_training_data[6]).name or '')) + '</b> in Planned Training, If Non pivotal is selected!='
                planned_training_list.append(planned_training_data[0])
        ############Adult Education and Training###############
        adult_education_training_list = []
        self._cr.execute("select id, name, surname from planned_adult_education_training_fields where planned_adult_education_wsp_id = %s",(self.id,))
        planned_adult_education_ids = self._cr.fetchall()
        for planned_adult_data in planned_adult_education_ids:
            # Validations on Name and Surname
            if planned_adult_data[1] and planned_adult_data[1].isdigit():
                msg += 'Name should not be numeric: <b>' + \
                    str(planned_adult_data[1]) + \
                    '</b> Planned Adult Education Training!='
                adult_education_training_list.append(planned_adult_data[0])
            if planned_adult_data[2] and planned_adult_data[2].isdigit():
                msg += 'Surname should not be numeric: <b>' + \
                    str(planned_adult_data[2]) + \
                    '</b> Planned Adult Education Training!='
                adult_education_training_list.append(planned_adult_data[0])
        ###########Variance Hard To Fill############
        scarce_and_critical_list = []
        self._cr.execute("select id, no_of_months from scarce_and_critical_skills_fields where scarce_and_critical_wsp_id = %s", (self.id,))
        scarce_and_critical_skills_ids = self._cr.fetchall()
        for scarce_data in scarce_and_critical_skills_ids:
            if scarce_data[1] == 0:
                msg += '"No of months position has been vacant" should not be 0 in Scarce and Critical Skills!='
                scarce_and_critical_list.append(scarce_data[0])
        if msg:
            # This code will write incorrect data into newly created xls file
            wsp_buffered = cStringIO.StringIO()
            workbook = xlsxwriter.Workbook(wsp_buffered)
            worksheet1 = workbook.add_worksheet('Total Employment Profile')
            worksheet2 = workbook.add_worksheet('Planned Training')
            worksheet3 = workbook.add_worksheet('Adult Education And Training')
            worksheet4 = workbook.add_worksheet('Vacancies Hard to fill')

            worksheet1.set_column(0, 40, 16)
            worksheet2.set_column(0, 40, 16)
            worksheet3.set_column(0, 40, 16)
            worksheet4.set_column(0, 40, 16)

            merge_format = workbook.add_format({
                'bold': 1,
                'border': 1,
                'align': 'center',
                'valign': 'vcenter'})

            ##########Total Employment Profile##############
            worksheet1.write('A1', 'SDL Number', merge_format)
            worksheet1.write('B1', 'First Name', merge_format)
            worksheet1.write('C1', 'Last Name', merge_format)
            worksheet1.write('D1', 'Citizen Status', merge_format)
            worksheet1.write('E1', 'Employee ID', merge_format)
            worksheet1.write('F1', 'ID Type', merge_format)
            worksheet1.write('G1', 'Date Of Birth', merge_format)
            worksheet1.write('H1', 'OFO Code', merge_format)
            worksheet1.write('I1', 'Occupation', merge_format)
            worksheet1.write('J1', 'Specialisation', merge_format)
            worksheet1.write('K1', 'Province', merge_format)
            worksheet1.write('L1', 'City', merge_format)
            worksheet1.write('M1', 'Urban/Rural', merge_format)
            worksheet1.write('N1', 'Highest Education Level', merge_format)
            worksheet1.write('O1', 'Race', merge_format)
            worksheet1.write('P1', 'Gender', merge_format)
            worksheet1.write('Q1', 'Disability', merge_format)
            tep_id_list = []
            [tep_id_list.append(x)
             for x in total_employment_list if x not in tep_id_list]
            if tep_id_list:
                self._cr.execute("select sdl_number, name, surname, citizen_resident_status_code, employee_id, id_type, dob,\
                ofo_code, occupation, specialization, province, city_id, urban, highest_education_level, population_group,\
                gender, dissability from total_employment_profile_fields where id in %s",([tuple(tep_id_list)]))
                total_employement_ids = self._cr.fetchall()
                total_employment_data_list = []
                for record in total_employement_ids:
                    record_list = []
                    self._cr.execute("select name from employer_sdl_no where id = %s",(record[0],))
                    sdl_no = self._cr.fetchone()
                    if sdl_no:
                        record_list.append(sdl_no[0])
                    
                    record_list.append(record[1])
                    record_list.append(record[2])
                    record_list.append(record[3])
                    record_list.append(record[4])
                    record_list.append(record[5])
                    record_list.append(record[6])
                    
                    self._cr.execute('select name from ofo_code where id = %s',(record[7],))
                    ofo_code = self._cr.fetchone()
                    if ofo_code:
                        record_list.append(ofo_code[0])
                    
                    self._cr.execute('select name from occupation_ofo where id = %s',(record[8],))
                    occupation_name = self._cr.fetchone()
                    if occupation_name:
                        record_list.append(occupation_name[0])
                    
                    self._cr.execute('select name from specialize_subject where id = %s',(record[9],))
                    specialisation_name = self._cr.fetchone()
                    if specialisation_name:
                        record_list.append(specialisation_name[0])
                    
                    self._cr.execute('select name from res_country_state where id = %s',(record[10],))
                    province_name = self._cr.fetchone()
                    if province_name:
                        record_list.append(province_name[0])
                    
                    self._cr.execute('select name from res_city where id = %s',(record[11],))
                    city_name = self._cr.fetchone()
                    if city_name:
                        record_list.append(city_name[0])
                    
                    record_list.append(record[12])
                    record_list.append(record[13])
                    record_list.append(record[14])
                    record_list.append(record[15])
                    record_list.append(record[16])
                    total_employment_data_list.append(record_list)
                row = 1
                for list in total_employment_data_list:
                    col = 0
                    for e in list:
                        if e:
                            worksheet1.write(row, col, e)
                        col += 1
                    row += 1

            #########Planned Training###########
            worksheet2.write('A1', 'Type Of Training', merge_format)
            worksheet2.write('B1', 'Name', merge_format)
            worksheet2.write('C1', 'Surname', merge_format)
            worksheet2.write('D1', 'Employee ID', merge_format)
            worksheet2.write('E1', 'OFO Code', merge_format)
            worksheet2.write('F1', 'Occupation', merge_format)
            worksheet2.write('G1', 'Specialisation', merge_format)
            worksheet2.write('H1', 'Province', merge_format)
            worksheet2.write('I1', 'City', merge_format)
            worksheet2.write('J1', 'Urban/Rural', merge_format)
            worksheet2.write('K1', 'Employed/UnEmployed', merge_format)
            worksheet2.write(
                'L1', 'Type of Training Intervention', merge_format)
            worksheet2.write(
                'M1', 'Other Type Of Training Intervention', merge_format)
            worksheet2.write(
                'N1', 'Name of training Intervention', merge_format)
            worksheet2.write('O1', 'Pivotal Programme Type', merge_format)
            worksheet2.write(
                'P1', 'Pivotal Programme Qualification', merge_format)
            worksheet2.write(
                'Q1', 'Pivotal Programme institution', merge_format)
            worksheet2.write('R1', 'Cost Per Learner', merge_format)
            worksheet2.write('S1', 'Start Date', merge_format)
            worksheet2.write('T1', 'End Date', merge_format)
            worksheet2.write('U1', 'NQF Aligned', merge_format)
            worksheet2.write('V1', 'NQF Level', merge_format)
            worksheet2.write('W1', 'Race', merge_format)
            worksheet2.write('X1', 'Gender', merge_format)
            worksheet2.write('Y1', 'Disability', merge_format)

            pt_id_list = []
            [pt_id_list.append(x)
             for x in planned_training_list if x not in pt_id_list]
            if pt_id_list:
                self._cr.execute("select training_type, name, surname, employee_id, code, occupation, specialization,\
                learner_province, city_id, urban, socio_economic_status, type_training, other_type_of_intervention,\
                name_training, pivotal_programme_institution, pivotal_programme_qualification, training_cost,\
                start_date, end_date, nqf_aligned, nqf_level, population_group, gender, dissability from planned_training_fields where id in %s",([tuple(pt_id_list)]))
                planned_training_ids = self._cr.fetchall()
                planned_training_data_list = []
                for record in planned_training_ids:
                    record_list = []
                    record_list.append(record[0])
                    record_list.append(record[1])
                    record_list.append(record[2])
                    record_list.append(record[3])
                    
                    self._cr.execute('select name from ofo_code where id = %s',(record[4],))
                    ofo_code = self._cr.fetchone()
                    if ofo_code:
                        record_list.append(ofo_code[0])
                    
                    self._cr.execute('select name from occupation_ofo where id = %s',(record[5],))
                    occupation_name = self._cr.fetchone()
                    if occupation_name:
                        record_list.append(occupation_name[0])
                    
                    self._cr.execute('select name from specialize_subject where id = %s',(record[6],))
                    specialisation_name = self._cr.fetchone()
                    if specialisation_name:
                        record_list.append(specialisation_name[0])
                    
                    self._cr.execute('select name from res_country_state where id = %s',(record[7],))
                    province_name = self._cr.fetchone()
                    if province_name:
                        record_list.append(province_name[0])
                    
                    self._cr.execute('select name from res_city where id = %s',(record[8],))
                    city_name = self._cr.fetchone()
                    if city_name:
                        record_list.append(city_name[0])
                    
                    record_list.append(record[9])
                    record_list.append(record[10])
                    self._cr.execute("select name from training_intervention where id = %s",(record[11],))
                    type_training = self._cr.fetchone()
                    if type_training:
                        record_list.append(type_training[0])
                    
                    record_list.append(record[12])
                    record_list.append(record[13])
                    
                    if type_training:
                        record_list.append(type_training[0])
                    
                    record_list.append(record[14])
                    record_list.append(record[15])
                    record_list.append(record[16])
                    record_list.append(record[17])
                    record_list.append(record[18])
                    record_list.append(record[19])
                    record_list.append(record[20])
                    record_list.append(record[21])
                    record_list.append(record[22])
                    record_list.append(record[23])
                    planned_training_data_list.append(record_list)
                row = 1
                for list in planned_training_data_list:
                    col = 0
                    for e in list:
                        if e:
                            worksheet2.write(row, col, e)
                        col += 1
                    row += 1

            ###########Planned Adult Education and Training###########
            worksheet3.write('A1', 'First Name', merge_format)
            worksheet3.write('B1', 'Surname', merge_format)
            worksheet3.write('C1', 'ID Number', merge_format)
            worksheet3.write('D1', 'Population Group', merge_format)
            worksheet3.write('E1', 'Gender', merge_format)
            worksheet3.write('F1', 'Disability Status And Type', merge_format)
            worksheet3.write('G1', 'Learner Province', merge_format)
            worksheet3.write('H1', 'City', merge_format)
            worksheet3.write('I1', 'Urban/Rural', merge_format)
            worksheet3.write('J1', 'AET Start Date', merge_format)
            worksheet3.write('K1', 'AET End Date', merge_format)
            worksheet3.write('L1', 'Provider', merge_format)
            worksheet3.write('M1', 'AET Level', merge_format)
            worksheet3.write('N1', 'AET Subject', merge_format)

            aet_id_list = []
            [aet_id_list.append(
                x) for x in adult_education_training_list if x not in aet_id_list]
            if aet_id_list:
                self._cr.execute("select name, surname, id_number, population_group, gender, dissability_status_and_type,\
                province, city_id, urban, start_date, end_date, provider, aet_level from \
                planned_adult_education_training_fields where id in %s",([tuple(aet_id_list)]))
                planned_adult_education_training_ids = self._cr.fetchall()
                adult_education_training_data_list = []
                for record in planned_adult_education_training_ids:
                    record_list = []
                    record_list.append(record[0])
                    record_list.append(record[1])
                    record_list.append(record[2])
                    record_list.append(record[3])
                    record_list.append(record[4])
                    record_list.append(record[5])
                    self._cr.execute('select name from res_country_state where id = %s',(record[6],))
                    province_name = self._cr.fetchone()
                    if province_name:
                        record_list.append(province_name[0])
                    
                    self._cr.execute('select name from res_city where id = %s',(record[7],))
                    city_name = self._cr.fetchone()
                    if city_name:
                        record_list.append(city_name[0])
                    
                    record_list.append(record[8])
                    record_list.append(record[9])
                    record_list.append(record[10])
                    record_list.append(record[11])
                    record_list.append(record[12])
                    
                    aet_subject_list = []
                    self._cr.execute("select aet_subject_id from aet_subject_planned_rel where planed_adult_education_training_id = %s",(record[0],))
                    aet_subject_ids = self._cr.fetchall()
                    for aet_subject_id in aet_subject_ids :
                        self._cr.execute("select name from aet_subject where id=%s",(aet_subject_id[0],))
                        fetch_aet_subject = self._cr.fetchone()
                        aet_subject_list.append(str(fetch_aet_subject[0])) 
                    aet_subject = ', '.join(aet_subject_list)        
                    record_list.append(aet_subject)
                    adult_education_training_data_list.append(record_list)
                row = 1
                for list in adult_education_training_data_list:
                    col = 0
                    for e in list:
                        if e:
                            worksheet3.write(row, col, e)
                        col += 1
                    row += 1
            ###########Variance Hard To Fill############
            worksheet4.write('A1', 'OFO Code', merge_format)
            worksheet4.write('B1', 'Occupation', merge_format)
            worksheet4.write('C1', 'Specialisation', merge_format)
            worksheet4.write('D1', 'Province', merge_format)
            worksheet4.write('E1', 'Number of Vacancies', merge_format)
            worksheet4.write('F1', 'Gender', merge_format)
            worksheet4.write('G1', 'Race', merge_format)
            worksheet4.write(
                'H1', 'Number of months position has', merge_format)
            worksheet4.write('I1', 'Comments', merge_format)

            sc_id_list = []
            [sc_id_list.append(x)
             for x in scarce_and_critical_list if x not in sc_id_list]
            if sc_id_list:
                self._cr.execute("select ofo_code, occupation, specialization, province, number_of_vacancies, \
                gender, population_group, no_of_months, comments from scarce_and_critical_skills_fields where id in %s",([tuple(sc_id_list)]))
                scarce_and_critical_ids = self._cr.fetchall()
                scarce_and_critical_data_list = []
                for record in scarce_and_critical_ids:
                    record_list = []
                    self._cr.execute('select name from ofo_code where id = %s',(record[0],))
                    ofo_code = self._cr.fetchone()
                    if ofo_code:
                        record_list.append(ofo_code[0])
                    
                    self._cr.execute('select name from occupation_ofo where id = %s',(record[1],))
                    occupation_name = self._cr.fetchone()
                    if occupation_name:
                        record_list.append(occupation_name[0])
                    
                    self._cr.execute('select name from specialize_subject where id = %s',(record[2],))
                    specialisation_name = self._cr.fetchone()
                    if specialisation_name:
                        record_list.append(specialisation_name[0])
                    
                    self._cr.execute('select name from res_country_state where id = %s',(record[3],))
                    province_name = self._cr.fetchone()
                    if province_name:
                        record_list.append(province_name[0])
                        
                    record_list.append(record[4])
                    record_list.append(record[5])
                    record_list.append(record[6])
                    record_list.append(record[7])
                    record_list.append(record[8])
                    scarce_and_critical_data_list.append(record_list)
                row = 1
                for list in scarce_and_critical_data_list:
                    col = 0
                    for e in list:
                        if e:
                            worksheet4.write(row, col, e)
                        col += 1
                    row += 1

            workbook.close()
            wsp_xlsx_data = wsp_buffered.getvalue()
            wsp_out_data = base64.encodestring(wsp_xlsx_data)
            attachment_obj = self.env['ir.attachment']
            wsp_new_attach = attachment_obj.create({
                'name': 'Incorrect WSP.xlsx',
                'res_name': 'wsp_import',
                'type': 'binary',
                'res_model': 'wsp.plan',
                'datas': wsp_out_data,
            })
            self = self.with_context({'error_log_msg': msg, 'incorrect_id': wsp_new_attach.id,
                                      'total_employement_list': tep_id_list,
                                      'planned_training_data_list': pt_id_list,
                                      'adult_education_training_data_list': aet_id_list,
                                      'scarce_and_critical_data_list': sc_id_list,
                                      })
            return {
                'name': 'WSP Error Log',
                'type': 'ir.actions.act_window',
                'view_type': 'form',
                'view_mode': 'form',
                'res_model': 'wsp.error.log',
                'target': 'new',
                'context': self._context,
            }
        return True

    
    def generate_atr_error_log(self):
        '''This method is used generate error log for ATR'''
        msg = ''
        if not self.save_some_record:
            raise Warning("Please check Partial Save!")
        #########Actual Training###################
        actual_training_list = []
        self._cr.execute("select id, name, surname, occupation, code, training_type, socio_economic_status,\
        african_dissabled, african_male, african_female, coloured_dissabled, coloured_male, coloured_female,\
        white_dissabled, white_male, white_female, indian_dissabled, indian_male, indian_female\
        from actual_training_fields where actual_wsp_id = %s", (self.id,))
        training_actual_ids = self._cr.fetchall()
        for actual_data in training_actual_ids:
              # Validations on Name and Surname
            if actual_data[1] and actual_data[1].isdigit():
                msg += 'Name should not be numeric: <b>' + \
                    str(actual_data[1]) + '</b> in Actual Training!='
                actual_training_list.append(actual_data[0])
            if actual_data[2] and actual_data[2].isdigit():
                msg += 'Surname should not be numeric: <b>' + \
                    str(actual_data[2]) + '</b> in Actual Training!='
                actual_training_list.append(actual_data[2])
            # Validations on occupation and code
            if actual_data[3] and actual_data[4]:
                self._cr.execute('select id from ofo_code where occupation = %s',(actual_data[3],))
                ofo_occupation_id = self._cr.fetchone()
                if ofo_occupation_id:
                    if actual_data[4] != ofo_occupation_id[0]:
                        msg += 'Wrong occupation for OFO code <b>' + str(self.env['ofo.code'].browse(
                            actual_data[4]).name or '') + '</b> in Actual Training!='
                        actual_training_list.append(actual_data[0])
            # Validations on socio economic status
            if actual_data[5] == 'non-pivotal' and actual_data[6] == 'unemployed':
                msg += 'Socio Economic Status should not be Unemployed ' + str(actual_data[4] and 'for OFO <b>' + str(self.env['ofo.code'].browse(
                            actual_data[4]).name or '')) + '</b> in Actual Training, If Non pivotal is selected!='
                actual_training_list.append(actual_data[0])
            # Validating total male, total female should not grater than total
            # dissabled
            if actual_data[7] is not None and actual_data[8] is not None and actual_data[9] is not None:
                if actual_data[7] > actual_data[8] + actual_data[9]:
                    msg += 'In actual training AD should not be greater than total of AM + AF <b> ' + \
                    str(actual_data[4] and 'for OFO ' + str(self.env['ofo.code'].browse(actual_data[4]).name or '')) + '</b>!='
                    actual_training_list.append(actual_data[0])
            if actual_data[10] is not None and actual_data[11] is not None and actual_data[12] is not None:
                if actual_data[10] > actual_data[11] + actual_data[12]:
                    msg += 'In actual training CD should not be greater than total of CM + CF <b> ' + \
                        str(actual_data[4] and 'for OFO ' +
                            str(self.env['ofo.code'].browse(actual_data[4]).name or '')) + '</b>!='
                    actual_training_list.append(actual_data[0])
            if actual_data[13] is not None and actual_data[14] is not None and actual_data[15] is not None:
                if actual_data[13] > actual_data[14] + actual_data[15]:
                    msg += 'In actual training WD should not be greater than total of WM + WF <b> ' + \
                        str(actual_data[4] and 'for OFO ' +
                            str(self.env['ofo.code'].browse(actual_data[4]).name or '')) + '</b>!='
                    actual_training_list.append(actual_data[0])
            if actual_data[16] is not None and actual_data[17] is not None and actual_data[18] is not None:
                if actual_data[16] > actual_data[17] + actual_data[18]:
                    msg += 'In actual training ID should not be greater than total of IM + IF <b> ' + \
                        str(actual_data[4] and 'for OFO ' +
                            str(self.env['ofo.code'].browse(actual_data[4]).name or '')) + '</b>!='
                    actual_training_list.append(actual_data[0])
        ##############Actual Adult Education and Training##########
        actual_adult_education_training_list = []
        self._cr.execute("select id, name, surname from actual_adult_education_fields where actual_adult_wsp_id=%s", (self.id,))
        actual_adult_education_ids = self._cr.fetchall()
        for actual_adult_data in actual_adult_education_ids:
            # Validations on Name and Surname
            if actual_adult_data[1] and actual_adult_data[1].isdigit():
                msg += 'Name should not be numeric: <b>' + \
                    str(actual_adult_data[1]) + \
                    '</b> in Adult Education and Training!='
                actual_adult_education_training_list.append(
                    actual_adult_data[0])
            if actual_adult_data[2] and actual_adult_data[2].isdigit():
                msg += 'Surname should not be numeric: <b>' + \
                    str(actual_adult_data[2]) + \
                    '</b> in Adult Education and Training!='
                actual_adult_education_training_list.append(
                    actual_adult_data[0])
        if msg:
            # This code will write incorrect data into newly created xls file
            buffered = cStringIO.StringIO()
            workbook = xlsxwriter.Workbook(buffered)
            worksheet1 = workbook.add_worksheet('Actual Training')
            worksheet1.set_column(0, 40, 16)
            merge_format = workbook.add_format({
                'bold': 1,
                'border': 1,
                'align': 'center',
                'valign': 'vcenter'})

            worksheet1.write('A1', 'Type Of Training', merge_format)
            worksheet1.write('B1', 'Name', merge_format)
            worksheet1.write('C1', 'Surname', merge_format)
            worksheet1.write('D1', 'Employee ID', merge_format)
            worksheet1.write('E1', 'OFO Code', merge_format)
            worksheet1.write('F1', 'Occupation', merge_format)
            worksheet1.write('G1', 'Specialisation', merge_format)
            worksheet1.write('H1', 'Province', merge_format)
            worksheet1.write('I1', 'City', merge_format)
            worksheet1.write('J1', 'Urban/Rural', merge_format)
            worksheet1.write('K1', 'Employed/UnEmployed', merge_format)
            worksheet1.write(
                'L1', 'Type of Training Intervention', merge_format)
            worksheet1.write(
                'M1', 'Other Type Of Training Intervention', merge_format)
            worksheet1.write(
                'N1', 'Name of Training Intervention', merge_format)
            worksheet1.write('O1', 'Pivotal Programme Type', merge_format)
            worksheet1.write(
                'P1', 'Pivotal Programme Qualification', merge_format)
            worksheet1.write(
                'Q1', 'Pivotal Programme Institution', merge_format)
            worksheet1.write('R1', 'Cost Per Learner', merge_format)
            worksheet1.write('S1', 'Start Date', merge_format)
            worksheet1.write('T1', 'End Date', merge_format)
            worksheet1.write('U1', 'NQF Aligned', merge_format)
            worksheet1.write('V1', 'NQF Level', merge_format)
            worksheet1.write('W1', 'Race', merge_format)
            worksheet1.write('X1', 'Gender', merge_format)
            worksheet1.write('Y1', 'Disability', merge_format)
            at_id_list = []
            [at_id_list.append(x)
             for x in actual_training_list if x not in at_id_list]
            if at_id_list:
                self._cr.execute("select training_type, name, surname, employee_id, code, occupation, specialization,\
                learner_province, city_id, urban, socio_economic_status, type_training, other_type_of_intervention,\
                name_training, pivotal_programme_institution, pivotal_programme_qualification, training_cost,\
                start_date, end_date, nqf_aligned, nqf_level, population_group, gender, dissability from actual_training_fields where id in %s",([tuple(at_id_list)]))
                actual_training_ids = self._cr.fetchall()
                actual_training_data_list = []
                for record in actual_training_ids:
                    record_list = []
                    record_list.append(record[0])
                    record_list.append(record[1])
                    record_list.append(record[2])
                    record_list.append(record[3])
                    
                    self._cr.execute('select name from ofo_code where id = %s',(record[4],))
                    ofo_code = self._cr.fetchone()
                    if ofo_code:
                        record_list.append(ofo_code[0])
                    
                    self._cr.execute('select name from occupation_ofo where id = %s',(record[5],))
                    occupation_name = self._cr.fetchone()
                    if occupation_name:
                        record_list.append(occupation_name[0])
                    
                    self._cr.execute('select name from specialize_subject where id = %s',(record[6],))
                    specialisation_name = self._cr.fetchone()
                    if specialisation_name:
                        record_list.append(specialisation_name[0])
                    
                    self._cr.execute('select name from res_country_state where id = %s',(record[7],))
                    province_name = self._cr.fetchone()
                    if province_name:
                        record_list.append(province_name[0])
                    
                    self._cr.execute('select name from res_city where id = %s',(record[8],))
                    city_name = self._cr.fetchone()
                    if city_name:
                        record_list.append(city_name[0])
                    
                    record_list.append(record[9])
                    record_list.append(record[10])
                    self._cr.execute("select name from training_intervention where id = %s",(record[11],))
                    type_training = self._cr.fetchone()
                    if type_training:
                        record_list.append(type_training[0])
                    
                    record_list.append(record[12])
                    record_list.append(record[13])
                    
                    if type_training:
                        record_list.append(type_training[0])
                    
                    record_list.append(record[14])
                    record_list.append(record[15])
                    record_list.append(record[16])
                    record_list.append(record[17])
                    record_list.append(record[18])
                    record_list.append(record[19])
                    record_list.append(record[20])
                    record_list.append(record[21])
                    record_list.append(record[22])
                    record_list.append(record[23])
                    actual_training_data_list.append(record_list)
                row = 1
                for list in actual_training_data_list:
                    col = 0
                    for e in list:
                        if e:
                            worksheet1.write(row, col, e)
                        col += 1
                    row += 1

            worksheet2 = workbook.add_worksheet('Adult Education And Training')
            worksheet2.set_column(0, 40, 16)
            worksheet2.write('A1', 'First Name', merge_format)
            worksheet2.write('B1', 'SurName', merge_format)
            worksheet2.write('C1', 'Id Number', merge_format)
            worksheet2.write('D1', 'Population Group', merge_format)
            worksheet2.write('E1', 'Gender', merge_format)
            worksheet2.write('F1', 'Disability Status And Type', merge_format)
            worksheet2.write('G1', 'Learner Province', merge_format)
            worksheet2.write('H1', 'City', merge_format)
            worksheet2.write('I1', 'Urban/Rural', merge_format)
            worksheet2.write('J1', 'AET Start Date', merge_format)
            worksheet2.write('K1', 'AET End Date', merge_format)
            worksheet2.write('L1', 'Provider', merge_format)
            worksheet2.write('M1', 'AET Level', merge_format)
            worksheet2.write('N1', 'AET Subject', merge_format)
            aaet_id_list = []
            [aaet_id_list.append(
                x) for x in actual_adult_education_training_list if x not in aaet_id_list]
            if aaet_id_list:
                self._cr.execute("select name, surname, id_number, population_group, gender, dissability_status_and_type,\
                province, city_id, urban, start_date, end_date, provider, aet_level from \
                actual_adult_education_training_fields where id in %s",([tuple(aaet_id_list)]))
                actual_adult_education_ids = self._cr.fetchall()
                actual_adult_education_training_data_list = []
                for record in actual_adult_education_ids:
                    record_list = []
                    record_list.append(record[0])
                    record_list.append(record[1])
                    record_list.append(record[2])
                    record_list.append(record[3])
                    record_list.append(record[4])
                    record_list.append(record[5])
                    self._cr.execute('select name from res_country_state where id = %s',(record[6],))
                    province_name = self._cr.fetchone()
                    if province_name:
                        record_list.append(province_name[0])
                    
                    self._cr.execute('select name from res_city where id = %s',(record[7],))
                    city_name = self._cr.fetchone()
                    if city_name:
                        record_list.append(city_name[0])
                    
                    record_list.append(record[8])
                    record_list.append(record[9])
                    record_list.append(record[10])
                    record_list.append(record[11])
                    record_list.append(record[12])
                    
                    aet_subject_list = []
                    self._cr.execute("select aet_subject_id from aet_subject_planned_rel where planed_adult_education_training_id = %s",(record[0],))
                    aet_subject_ids = self._cr.fetchall()
                    for aet_subject_id in aet_subject_ids :
                        self._cr.execute("select name from aet_subject where id=%s",(aet_subject_id[0],))
                        fetch_aet_subject = self._cr.fetchone()
                        aet_subject_list.append(str(fetch_aet_subject[0])) 
                    aet_subject = ', '.join(aet_subject_list)        
                    record_list.append(aet_subject)
                    actual_adult_education_training_data_list.append(record_list)
                row = 1
                for list in actual_adult_education_training_data_list:
                    col = 0
                    for e in list:
                        if e:
                            worksheet2.write(row, col, e)
                        col += 1
                    row += 1
                
            workbook.close()
            xlsx_data = buffered.getvalue()
            out_data = base64.encodestring(xlsx_data)
            attachment_obj = self.env['ir.attachment']
            new_attach = attachment_obj.create({
                'name': 'Incorrect ATR.xlsx',
                'res_name': 'wsp_import',
                'type': 'binary',
                'res_model': 'wsp.plan',
                'datas': out_data,
            })
            self = self.with_context({'error_log_msg': msg, 'incorrect_id': new_attach.id,
                                      'actual_training_data_list': at_id_list,
                                      'actual_adult_education_training_list': aaet_id_list, })
            return {
                'name': 'ATR Error Log',
                'type': 'ir.actions.act_window',
                        'view_type': 'form',
                        'view_mode': 'form',
                        'res_model': 'atr.error.log',
                        'target': 'new',
                        'context': self._context,
            }
        return True

    
    def validate_date(self, start_date, end_date, wsp_section):
        ''' This method will validate start date and end date '''
        current_year = datetime.now().date().year
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        if start_date > end_date:
            raise Warning(
                _('Start Date should be less than End Date in %s !') % (wsp_section))
        if start_date.year < current_year or end_date.year < current_year:
            raise Warning(
                _('Start Date or End Date should be greater than or equals to current year in %s !') % (wsp_section))

    def clear_all_atr(self):
        if self.training_actual_ids:
            self._cr.execute(
                "delete from actual_training_fields where actual_wsp_id=%s", (self.id,))
            self._cr.commit()
#             self.write({'training_actual_ids':[(2,actual_data.id) for actual_data in self.training_actual_ids]})
        self.write({'is_prev_wsp_loaded': False})
        return True

    def clear_all_adult_education_training(self):
        if self.actual_adult_education_ids:
            self._cr.execute(
                "delete from actual_adult_education_fields where actual_adult_wsp_id=%s", (self.id,))
            self._cr.commit()
#             self.write({'actual_adult_education_ids':[(2,actual_data.id) for actual_data in self.actual_adult_education_ids]})
        return True

    def clear_all_tep(self):
        if self.total_employment_profile_ids:
            self._cr.execute(
                "delete from total_employment_profile_fields where total_employment_wsp_id=%s", (self.id,))
            self._cr.commit()
#             self.write({'total_employment_profile_ids':[(2,planned_data.id) for planned_data in self.total_employment_profile_ids]})
        self.write({'is_tep_loaded': False})
        return True

    def clear_all_planned_training(self):
        if self.training_planned_ids:
            self._cr.execute(
                "delete from planned_training_fields where planned_training_non_wsp_id=%s", (self.id,))
            self._cr.commit()
#             self.write({'training_planned_ids':[(2,planned_data.id) for planned_data in self.training_planned_ids]})
        self.write({'is_tep_to_planned_loaded': False})
        return True

    def clear_all_planned_adult_training(self):
        if self.planned_adult_education_ids:
            self._cr.execute(
                "delete from planned_adult_education_training_fields where planned_adult_education_wsp_id=%s", (self.id,))
            self._cr.commit()
#             self.write({'planned_adult_education_ids':[(2,planned_data.id) for planned_data in self.planned_adult_education_ids]})
        return True

    def clear_all_scarce_and_critical_skills(self):
        if self.scarce_and_critical_skills_ids:
            self._cr.execute(
                "delete from scarce_and_critical_skills_fields where scarce_and_critical_wsp_id=%s", (self.id,))
            self._cr.commit()
#             self.write({'scarce_and_critical_skills_ids':[(2,planned_data.id) for planned_data in self.scarce_and_critical_skills_ids]})
        return True


# Class for variance calculations.


class VarianceCalculate(models.Model):
    _name = 'variance.calculate'
    _description = 'WSP vs ATR Variance Calculation'

    line_validated = fields.Boolean(
        string='Validated',
        default=False,
        tracking=True
    )

    type_training = fields.Many2one(
        'training.intervention',
        string='Type of Training Intervention / Pivotal Programme Type',
        domain=[('pivotal', '=', False)],
        ondelete='restrict'
    )

    wsp_planned = fields.Integer(
        string='WSP (Planned Learners)'
    )

    total_cost_planned = fields.Float(
        string='Total Training Cost (Planned)'
    )

    atr_actual = fields.Integer(
        string='ATR (Actual Learners)'
    )

    total_cost_actual = fields.Float(
        string='Total Cost (Actual)'
    )

    variance_percentage = fields.Float(
        string='Variance Percentage',
        compute='_compute_variance_percentage',
        store=True
    )

    comments = fields.Text(
        string='Comments / Reasons'
    )

    wsp_plan_id = fields.Many2one(
        'wsp.plan',
        string='Related WSP',
        ondelete='cascade'
    )

    @api.depends('total_cost_planned', 'total_cost_actual')
    def _compute_variance_percentage(self):
        for rec in self:
            if rec.total_cost_planned:
                rec.variance_percentage = (
                    (rec.total_cost_actual - rec.total_cost_planned)
                    / rec.total_cost_planned
                ) * 100
            else:
                rec.variance_percentage = 0.0


class WspSubmissionStatus(models.Model):
    _name = 'wsp.submission.status'
    _description = 'WSP Submission Status'
    _order = 'date_evaluation desc'

    evaluator = fields.Many2one(
        'res.users',
        string='Evaluator'
    )

    date_evaluation = fields.Date(
        string='Date of Evaluation'
    )

    status = fields.Selection([
        ('recommended', 'Recommended'),
        ('not_recommended', 'Not Recommended'),
        ('query', 'Query')
    ], string='Evaluation Status')

    state = fields.Selection([
        ('draft', 'Draft'),
        ('submitted', 'Submitted'),
        ('evaluated', 'Assessment'),
        ('evaluated2', 'Evaluated'),
        ('approved', 'Accepted'),
        ('query', 'Query'),
        ('rejected', 'Rejected')
    ], string='Workflow State', default='draft', tracking=True)

    comments = fields.Text(
        string='Comments'
    )

    wsp_plan_id = fields.Many2one(
        'wsp.plan',
        string='Related WSP',
        ondelete='cascade'
    )



class WspExtensionStatus(models.Model):
    _name = 'wsp.extension.status'
    _description = 'WSP Extension Status'
    _order = 'date_evaluation desc'

    evaluator = fields.Many2one(
        'res.users',
        string='Evaluator'
    )

    date_evaluation = fields.Date(
        string='Date of Evaluation'
    )

    status = fields.Selection([
        ('requested', 'Requested'),
        ('resubmit', 'Re-submit'),
        ('verified', 'Verified'),
        ('approved', 'Approved'),
        ('rejected', 'Rejected')
    ], string='Extension Status', tracking=True)

    state = fields.Selection([
        ('requested', 'Requested'),
        ('resubmit', 'Re-submit'),
        ('verified', 'Verified'),
        ('approved', 'Approved'),
        ('rejected', 'Rejected')
    ], string='Extension State', default='requested')

    wsp_state = fields.Selection([
        ('draft', 'Draft'),
        ('submitted', 'Submitted'),
        ('evaluated', 'Assessment'),
        ('evaluated2', 'Evaluated'),
        ('approved', 'Accepted'),
        ('query', 'Query'),
        ('rejected', 'Rejected')
    ], string='WSP Submission State', default='draft')

    comments = fields.Text(
        string='Comments'
    )

    wsp_plan_id = fields.Many2one(
        'wsp.plan',
        string='Related WSP',
        ondelete='cascade'
    )


class WspQueryDocs(models.Model):
    _name = 'wsp.query.docs'
    _description = 'WSP Query Supporting Documents'

    query_docs_statement = fields.Many2one(
        'ir.attachment',
        string="Extension of WSP Reason Statement / Letter",
        help="Upload supporting document"
    )
    comments = fields.Text(string='Comments')
    wsp_query_id = fields.Integer(string='WSP Query ID')
    wsp_plan_id = fields.Many2one(
        'wsp.plan',
        string='Related WSP',
        ondelete='cascade'
    )


class WspQueryStatus(models.Model):
    _name = 'wsp.query.status'
    _description = 'WSP Query Status'

    evaluator = fields.Many2one(
        'res.users',
        string='Evaluator'
    )
    wsp_query_id = fields.Integer(string='WSP Query ID')
    date_evaluation = fields.Date(string='Date of Evaluation')

    status = fields.Selection(
        [
            ('requested', 'Requested'),
            ('resubmit', 'Re-submit'),
            ('verified', 'Verified'),
            ('approved', 'Approved'),
            ('rejected', 'Rejected'),
        ],
        string="Query Status"
    )

    state = fields.Selection(
        [
            ('requested', 'Requested'),
            ('resubmit', 'Re-submit'),
            ('verified', 'Verified'),
            ('approved', 'Approved'),
            ('rejected', 'Rejected'),
        ],
        string="Query State"
    )

    wsp_state = fields.Selection(
        [
            ('draft', 'Draft'),
            ('submitted', 'Submitted'),
            ('evaluated', 'Assessment'),
            ('evaluated2', 'Evaluated'),
            ('approved', 'Accepted'),
            ('query', 'Query'),
            ('rejected', 'Rejected'),
        ],
        string="WSP Submission State",
        default='draft'
    )

    comments = fields.Text(string='Comments')
    wsp_plan_id = fields.Many2one(
        'wsp.plan',
        string='Related WSP',
        ondelete='cascade'
    )

class WspExtension(models.Model):
    _name = 'wsp.extension'
    _description = 'WSP Extension'

    # --------------------------------------------------
    # Actions
    # --------------------------------------------------

    def action_verify_extension(self):
        return {
            'name': _('WSP Extension Verification'),
            'type': 'ir.actions.act_window',
            'res_model': 'wsp.extension.verify',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'active_model': self._name,
                'active_ids': self.ids,
            },
        }

    def action_approve_extension(self):
        return {
            'name': _('WSP Extension Approval'),
            'type': 'ir.actions.act_window',
            'res_model': 'wsp.extension.approval',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'active_model': self._name,
                'active_ids': self.ids,
            },
        }

    def action_reject_extension(self):
        return {
            'name': _('WSP Extension Rejection'),
            'type': 'ir.actions.act_window',
            'res_model': 'wsp.extension.reject',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'active_model': self._name,
                'active_ids': self.ids,
            },
        }

    def action_request_extension(self):
        return {
            'name': _('Request Extension'),
            'type': 'ir.actions.act_window',
            'res_model': 'wsp.extension.request',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'active_model': self._name,
                'active_ids': self.ids,
            },
        }

    def action_resubmit_extension(self):
        return {
            'name': _('Re-Submit Extension'),
            'type': 'ir.actions.act_window',
            'res_model': 'wsp.extension.resubmit',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'active_model': self._name,
                'active_ids': self.ids,
            },
        }

    # --------------------------------------------------
    # Compute
    # --------------------------------------------------

    @api.depends('allow_resubmit_extension')
    def _compute_approval_check(self):
        config = self.env['leavy.income.config'].search([], limit=1)
        today = fields.Date.today()

        for rec in self:
            rec.approval_check = False
            if not config:
                continue

            if (
                config.wsp_extension_request_start_date
                and config.wsp_extension_approval_end_date
                and config.wsp_extension_request_start_date <= today <= config.wsp_extension_approval_end_date
            ):
                rec.approval_check = True

    # --------------------------------------------------
    # Core Relations
    # --------------------------------------------------

    wsp_plan_id = fields.Many2one(
        'wsp.plan',
        string='Related WSP',
        required=True,
        ondelete='cascade'
    )

    # --------------------------------------------------
    # Related Fields (from wsp.plan)
    # --------------------------------------------------

    name = fields.Char(related='wsp_plan_id.name', store=True)
    sdf_id = fields.Many2one(
        related='wsp_plan_id.sdf_id',
        string='SDF',
        store=True
    )
    employer_id = fields.Many2one(
        related='wsp_plan_id.employer_id',
        string='Employer',
        store=True
    )
    emp_province = fields.Many2one(
        related='wsp_plan_id.emp_province',
        store=True
    )
    state = fields.Selection(
        related='wsp_plan_id.state',
        store=True
    )
    fiscal_year = fields.Many2one(
        related='wsp_plan_id.fiscal_year',
        store=True
    )

    extension_doc_statement = fields.Many2one(
        related='wsp_plan_id.extension_doc_statement',
        string='Extension Reason Document',
        store=True
    )

    submittable = fields.Boolean(related='wsp_plan_id.submittable', store=True)
    extension_allowed = fields.Boolean(related='wsp_plan_id.extension_allowed', store=True)
    allow_extension = fields.Boolean(related='wsp_plan_id.allow_extension', store=True)
    show_extension_date = fields.Boolean(related='wsp_plan_id.show_extension_date', store=True)

    is_tep_loaded = fields.Boolean(related='wsp_plan_id.is_tep_loaded', store=True)
    is_tep_to_planned_loaded = fields.Boolean(related='wsp_plan_id.is_tep_to_planned_loaded', store=True)
    is_prev_wsp_loaded = fields.Boolean(related='wsp_plan_id.is_prev_wsp_loaded', store=True)

    current_user_forbidden = fields.Boolean(
        related='wsp_plan_id.current_user_forbidden',
        store=True
    )

    wsp_start_period = fields.Date(related='wsp_plan_id.wsp_start_period', store=True)
    wsp_end_period = fields.Date(related='wsp_plan_id.wsp_end_period', store=True)
    start_period = fields.Date(related='wsp_plan_id.start_period', store=True)
    end_period = fields.Date(related='wsp_plan_id.end_period', store=True)

    request_extension_date = fields.Date(
        related='wsp_plan_id.request_extension_date',
        store=True
    )
    approve_extension_date = fields.Date(
        related='wsp_plan_id.approve_extension_date',
        store=True
    )
    extension_date = fields.Date(
        related='wsp_plan_id.extension_date',
        store=True
    )

    sdl_no = fields.Char(related='wsp_plan_id.sdl_no', store=True)
    wsp_submission_date = fields.Date(
        related='wsp_plan_id.wsp_submission_date',
        store=True
    )

    duplicate_total_mf_tep = fields.Integer(
        related='wsp_plan_id.duplicate_total_mf_tep',
        store=True
    )

    scheme_year_id = fields.Many2one(
        related='wsp_plan_id.scheme_year_id',
        store=True
    )

    # --------------------------------------------------
    # Extension State
    # --------------------------------------------------

    extension_state = fields.Selection(
        [
            ('requested', 'Requested'),
            ('resubmit', 'Re-submit'),
            ('verified', 'Verified'),
            ('approved', 'Approved'),
            ('rejected', 'Rejected'),
        ],
        string='Extension State'
    )

    extension_reason_comment = fields.Text()
    extension_verify_reason_comment = fields.Text()
    extension_approve_reason_comment = fields.Text()
    extension_reject_reason_comment = fields.Text()

    extension_verified_by = fields.Many2one('res.users')
    extension_approved_by = fields.Many2one('res.users')
    extension_rejected_by = fields.Many2one('res.users')

    allow_resubmit_extension = fields.Boolean(
        string='Allow Re-Submit',
        default=False
    )

    approval_check = fields.Boolean(
        string='Can Approve',
        compute='_compute_approval_check',
        store=False
    )

    user_id = fields.Many2one(
        'res.users',
        default=lambda self: self.env.user,
        string='Current User'
    )


class WspQuery(models.Model):
    _name = 'wsp.query'
    _description = 'WSP Query'

    # --------------------------------------------------
    # Actions
    # --------------------------------------------------

    def action_verify_query(self):
        return {
            'name': _('WSP Query Verification'),
            'type': 'ir.actions.act_window',
            'res_model': 'wsp.query.verify',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'active_model': self._name,
                'active_ids': self.ids,
            },
        }

    def action_update_query(self):
        return {
            'name': _('WSP Query Update'),
            'type': 'ir.actions.act_window',
            'res_model': 'wsp.query.update',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'active_model': self._name,
                'active_ids': self.ids,
            },
        }

    def action_approve_query(self):
        return {
            'name': _('WSP Query Approval'),
            'type': 'ir.actions.act_window',
            'res_model': 'wsp.query.approval',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'active_model': self._name,
                'active_ids': self.ids,
            },
        }

    def action_reject_query(self):
        return {
            'name': _('WSP Query Rejection'),
            'type': 'ir.actions.act_window',
            'res_model': 'wsp.query.reject',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'active_model': self._name,
                'active_ids': self.ids,
            },
        }

    def action_request_query(self):
        return {
            'name': _('Request Query'),
            'type': 'ir.actions.act_window',
            'res_model': 'wsp.query.request',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'active_model': self._name,
                'active_ids': self.ids,
            },
        }

    def action_resubmit_query(self):
        return {
            'name': _('Re-Submit Query'),
            'type': 'ir.actions.act_window',
            'res_model': 'wsp.query.resubmit',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'active_model': self._name,
                'active_ids': self.ids,
            },
        }

    # --------------------------------------------------
    # Core Relation
    # --------------------------------------------------

    wsp_plan_id = fields.Many2one(
        'wsp.plan',
        string='Related WSP',
        required=True,
        ondelete='cascade'
    )

    # --------------------------------------------------
    # Related Fields (from wsp.plan)
    # --------------------------------------------------

    name = fields.Char(related='wsp_plan_id.name', store=True)
    sdf_id = fields.Many2one(
        related='wsp_plan_id.sdf_id',
        string='SDF',
        store=True
    )
    employer_id = fields.Many2one(
        related='wsp_plan_id.employer_id',
        store=True
    )
    emp_province = fields.Many2one(
        related='wsp_plan_id.emp_province',
        store=True
    )
    state = fields.Selection(
        related='wsp_plan_id.state',
        store=True
    )
    fiscal_year = fields.Many2one(
        related='wsp_plan_id.fiscal_year',
        store=True
    )

    # --------------------------------------------------
    # Query Documents
    # --------------------------------------------------

    query_doc_statement = fields.Many2one(
        related='wsp_plan_id.query_doc_statement',
        store=True
    )

    query_docs_statement = fields.Many2one(
        'ir.attachment',
        string='Query Supporting Document'
    )

    wsp_query_docs_ids = fields.One2many(
        'wsp.query.docs',
        'wsp_plan_id',
        compute='_compute_query_docs',
        string='Query Documents'
    )

    @api.depends('wsp_plan_id')
    def _compute_query_docs(self):
        for rec in self:
            rec.wsp_query_docs_ids = rec.wsp_plan_id.wsp_query_docs_ids

    # --------------------------------------------------
    # Query State & Comments
    # --------------------------------------------------

    query_state = fields.Selection(
        [
            ('requested', 'Requested'),
            ('resubmit', 'Re-submit'),
            ('verified', 'Verified'),
            ('approved', 'Approved'),
            ('rejected', 'Rejected'),
        ],
        string='Query State'
    )

    query_reason_comment = fields.Text()
    query_update_reason_comment = fields.Text()
    query_verify_reason_comment = fields.Text()
    query_approve_reason_comment = fields.Text()
    query_reject_reason_comment = fields.Text()

    before_query_state = fields.Selection(
        [
            ('draft', 'Draft'),
            ('submitted', 'Submitted'),
            ('evaluated', 'Assessment'),
            ('evaluated2', 'Evaluated'),
            ('approved', 'Accepted'),
            ('query', 'Query'),
            ('rejected', 'Rejected'),
        ],
        string='Previous WSP State'
    )

    # --------------------------------------------------
    # Permissions / Flags
    # --------------------------------------------------

    submittable = fields.Boolean(related='wsp_plan_id.submittable', store=True)
    query_allowed = fields.Boolean(related='wsp_plan_id.query_allowed', store=True)
    allow_query = fields.Boolean(related='wsp_plan_id.allow_query', store=True)
    show_query_date = fields.Boolean(related='wsp_plan_id.show_query_date', store=True)

    is_tep_loaded = fields.Boolean(related='wsp_plan_id.is_tep_loaded', store=True)
    is_tep_to_planned_loaded = fields.Boolean(related='wsp_plan_id.is_tep_to_planned_loaded', store=True)
    is_prev_wsp_loaded = fields.Boolean(related='wsp_plan_id.is_prev_wsp_loaded', store=True)

    current_user_forbidden = fields.Boolean(
        related='wsp_plan_id.current_user_forbidden',
        store=True
    )

    # --------------------------------------------------
    # Dates
    # --------------------------------------------------

    wsp_start_period = fields.Date(related='wsp_plan_id.wsp_start_period', store=True)
    wsp_end_period = fields.Date(related='wsp_plan_id.wsp_end_period', store=True)
    start_period = fields.Date(related='wsp_plan_id.start_period', store=True)
    end_period = fields.Date(related='wsp_plan_id.end_period', store=True)

    request_query_date = fields.Date(
        related='wsp_plan_id.request_query_date',
        store=True
    )
    approve_query_date = fields.Date(
        related='wsp_plan_id.approve_query_date',
        store=True
    )
    query_date = fields.Date(
        related='wsp_plan_id.query_date',
        store=True
    )

    wsp_submission_date = fields.Date(
        related='wsp_plan_id.wsp_submission_date',
        store=True
    )

    # --------------------------------------------------
    # Misc
    # --------------------------------------------------

    sdl_no = fields.Char(related='wsp_plan_id.sdl_no', store=True)
    duplicate_total_mf_tep = fields.Integer(
        related='wsp_plan_id.duplicate_total_mf_tep',
        store=True
    )

    scheme_year_id = fields.Many2one(
        related='wsp_plan_id.scheme_year_id',
        store=True
    )

    user_id = fields.Many2one(
        'res.users',
        default=lambda self: self.env.user,
        string='Current User'
    )

    query_verified_by = fields.Many2one('res.users')
    query_approved_by = fields.Many2one('res.users')
    query_rejected_by = fields.Many2one('res.users')

    allow_resubmit_query = fields.Boolean(
        string='Allow Re-Submit Query',
        default=False
    )

# class wsp_submission_status_wspofficer(models.Model):
#     _name = 'wsp.submission.status.wspofficer'
#
#     evaluator = fields.Many2one('res.users', string='WSP officer Evaluation')
#     date_evaluation = fields.Date(string='Date of Evaluation')
#     status = fields.Selection([('recommended','Recommended'),('not_recommended','Not Recommended')],string='Status')
#     comments = fields.Text(string='Comments')
#     wsp_plan_id = fields.Many2one('wsp.plan', string='Related WSP')
#
# wsp_submission_status_wspofficer()
#
# class wsp_submission_status_wspmanager(models.Model):
#     _name = 'wsp.submission.status.wspmanager'
#
#     evaluator = fields.Many2one('res.users', string='WSP Manager Evaluation')
#     date_evaluation = fields.Date(string='Date of Evaluation')
#     status = fields.Selection([('recommended','Recommended'),('not_recommended','Not Recommended')],string='Status')
#     comments = fields.Text(string='Comments')
#     wsp_plan_id = fields.Many2one('wsp.plan', string='Related WSP')
#
# wsp_submission_status_wspmanager()


class VarianceCalculatePivotal(models.Model):
    _name = 'variance.calculate.pivotal'
    _description = 'Variance Calculation Pivotal'

    line_validated = fields.Boolean(default=False)

    type_training = fields.Many2one(
        'training.intervention',
        string='Type of Training Intervention / Pivotal Programme Type',
        domain=[('pivotal', '=', True)]
    )

    wsp_planned = fields.Integer(string='WSP')
    total_cost_planned = fields.Float(string='Total Training Cost for Planned Training')

    atr_actual = fields.Integer(string='ATR')
    total_cost_actual = fields.Float(string='Total Cost Actual Training')

    variance_percentage = fields.Float(string="Percentage")
    comments = fields.Text(string='Comments / Reasons')

    wsp_plan_id = fields.Many2one('wsp.plan', string='Related WSP')


# Class for Actual Training Data (For both Pivotal and Non-Pivotal
# Training Sections).
class ActualTrainingFields(models.Model):
    _name = 'actual.training.fields'
    _description = 'Actual Training Fields'

    line_validated = fields.Boolean(default=False)

    # -------------------------------------------------
    # COMPUTED FIELDS
    # -------------------------------------------------
    total_male = fields.Integer(
        string='TM', compute='_compute_totals', store=True)
    total_female = fields.Integer(
        string='TF', compute='_compute_totals', store=True)
    total_dissabled = fields.Integer(
        string='TD', compute='_compute_totals', store=True)
    total_cost = fields.Float(
        string='Total Cost', compute='_compute_total_cost', store=True)

    @api.depends('gender',
                 'african_dissabled', 'coloured_dissabled',
                 'indian_dissabled', 'white_dissabled')
    def _compute_totals(self):
        for rec in self:
            rec.total_male = 1 if rec.gender == 'male' else 0
            rec.total_female = 1 if rec.gender == 'female' else 0
            rec.total_dissabled = (
                rec.african_dissabled +
                rec.coloured_dissabled +
                rec.indian_dissabled +
                rec.white_dissabled
            )

    @api.depends('total_male', 'total_female', 'training_cost')
    def _compute_total_cost(self):
        for rec in self:
            rec.total_cost = (rec.total_male + rec.total_female) * rec.training_cost

    # -------------------------------------------------
    # BASIC DETAILS
    # -------------------------------------------------
    training_type = fields.Selection(
        [('pivotal', 'Pivotal'), ('non-pivotal', 'Non Pivotal')],
        string='Type of Training'
    )

    name = fields.Char(string='First Name')
    surname = fields.Char(string='Surname')
    employee_id = fields.Char(string='Employee ID', size=20)

    code = fields.Many2one('ofo.code', string='OFO Code')
    occupation = fields.Many2one('occupation.ofo', string='Occupation')
    specialization = fields.Many2one('specialize.subject', string='Specialisation')
    specialisation_exists = fields.Boolean()

    # -------------------------------------------------
    # LOCATION
    # -------------------------------------------------
    municipality_id = fields.Many2one('res.municipality', string='Municipality')
    city_id = fields.Many2one('res.city', string='City')
    suburb_id = fields.Many2one('res.suburb', string='Suburb')
    learner_province = fields.Many2one(
        'res.country.state',
        string='Province',
        domain=[('country_id.code', '=', 'ZA')]
    )

    urban = fields.Selection(
        [('urban', 'Urban'), ('rural', 'Rural'), ('unknown', 'Unknown')],
        string='Urban/Rural'
    )

    # -------------------------------------------------
    # POPULATION (OLD FORMAT)
    # -------------------------------------------------
    african_male = fields.Integer(string='AM')
    african_female = fields.Integer(string='AF')
    african_dissabled = fields.Integer(string='AD')

    coloured_male = fields.Integer(string='CM')
    coloured_female = fields.Integer(string='CF')
    coloured_dissabled = fields.Integer(string='CD')

    indian_male = fields.Integer(string='IM')
    indian_female = fields.Integer(string='IF')
    indian_dissabled = fields.Integer(string='ID')

    white_male = fields.Integer(string='WM')
    white_female = fields.Integer(string='WF')
    white_dissabled = fields.Integer(string='WD')

    # -------------------------------------------------
    # NEW FORMAT
    # -------------------------------------------------
    population_group = fields.Selection([
        ('african', 'African'),
        ('white', 'White'),
        ('coloured', 'Coloured'),
        ('indian', 'Indian')
    ])

    gender = fields.Selection([
        ('female', 'Female'),
        ('male', 'Male')
    ])

    age_group_less = fields.Integer(string='<35')
    age_group_upto = fields.Integer(string='35-55')
    age_group_greater = fields.Integer(string='>55')

    # -------------------------------------------------
    # TRAINING DETAILS
    # -------------------------------------------------
    nqf_aligned = fields.Selection([('yes', 'Yes'), ('no', 'No')])
    nqf_level = fields.Selection([
        ('abet', 'ABET'),
        ('level1', 'Level 1'),
        ('level2', 'Level 2'),
        ('level3', 'Level 3'),
        ('level4', 'Level 4'),
        ('level5', 'Level 5'),
        ('level6', 'Level 6'),
        ('level7', 'Level 7'),
        ('level8', 'Level 8'),
        ('level9', 'Level 9'),
        ('level10', 'Level 10')
    ])

    type_training = fields.Many2one(
        'training.intervention',
        string='Training Intervention / Pivotal Programme'
    )

    name_training = fields.Char(string='Training Name')
    socio_economic_status = fields.Selection([
        ('employed', 'Employed'),
        ('unemployed', 'Unemployed')
    ])

    pivotal_programme_institution = fields.Char()
    pivotal_programme_qualification = fields.Char()
    other_type_of_intervention = fields.Char()

    start_date = fields.Date()
    end_date = fields.Date()
    training_cost = fields.Float()

    dissability = fields.Selection([('yes', 'Yes'), ('no', 'No')], default='no')

    # -------------------------------------------------
    # RELATIONS
    # -------------------------------------------------
    actual_wsp_id = fields.Many2one('wsp.plan', string='Actual WSP')
    actual_planned_wsp_id = fields.Many2one('wsp.plan', string='Actual Planned WSP')
    save_actual_planned_record = fields.Boolean(default=False)

    # -------------------------------------------------
    # ONCHANGE METHODS
    # -------------------------------------------------
    @api.onchange('learner_province')
    def _onchange_learner_province(self):
        if self.learner_province:
            self.municipality_id = False
            self.city_id = False

    @api.onchange('suburb_id')
    def _onchange_suburb(self):
        if self.suburb_id:
            self.municipality_id = self.suburb_id.municipality_id
            self.city_id = self.suburb_id.city_id
            self.urban = self.suburb_id.urban_rural

    @api.onchange('city_id')
    def _onchange_city(self):
        if self.city_id:
            self.learner_province = self.city_id.province_id
            self.urban = self.city_id.urban_rural

    @api.onchange('municipality_id')
    def _onchange_municipality(self):
        if self.municipality_id:
            self.urban = self.municipality_id.urban_rural

    @api.onchange('training_type')
    def _onchange_training_type(self):
        if self.training_type == 'non-pivotal':
            self.socio_economic_status = 'employed'
        if self.training_type == 'pivotal':
            self.nqf_aligned = 'yes'

class ActualAdultEducationFields(models.Model):
    _name = 'actual.adult.education.fields'
    _description = 'Actual Adult Education Fields'

    # --------------------------------------------------
    # COMPUTED TOTALS
    # --------------------------------------------------

    @api.depends(
        'african_male', 'coloured_male', 'indian_male', 'white_male'
    )
    def _compute_total_male(self):
        for rec in self:
            rec.total_male = (
                rec.african_male + rec.coloured_male +
                rec.indian_male + rec.white_male
            )

    @api.depends(
        'african_female', 'coloured_female', 'indian_female', 'white_female'
    )
    def _compute_total_female(self):
        for rec in self:
            rec.total_female = (
                rec.african_female + rec.coloured_female +
                rec.indian_female + rec.white_female
            )

    @api.depends(
        'african_dissabled', 'coloured_dissabled',
        'indian_dissabled', 'white_dissabled'
    )
    def _compute_total_disabled(self):
        for rec in self:
            rec.total_dissabled = (
                rec.african_dissabled + rec.coloured_dissabled +
                rec.indian_dissabled + rec.white_dissabled
            )

    # --------------------------------------------------
    # VALIDATION
    # --------------------------------------------------

    @api.constrains('id_number')
    def _check_id_number(self):
        for rec in self:
            if rec.id_number and rec.id_number.isdigit() and int(rec.id_number) == 0:
                raise ValidationError(_('Please enter a valid ID Number.'))

    # --------------------------------------------------
    # FIELDS
    # --------------------------------------------------

    name = fields.Char(string='First Name')
    surname = fields.Char(string='Surname')
    ofo_code = fields.Many2one('ofo.code', string='OFO Code')
    id_number = fields.Char(string='ID Number', size=20)

    african_male = fields.Integer(string='AM')
    african_female = fields.Integer(string='AF')
    african_dissabled = fields.Integer(string='AD')
    coloured_male = fields.Integer(string='CM')
    coloured_female = fields.Integer(string='CF')
    coloured_dissabled = fields.Integer(string='CD')
    indian_male = fields.Integer(string='IM')
    indian_female = fields.Integer(string='IF')
    indian_dissabled = fields.Integer(string='ID')
    white_male = fields.Integer(string='WM')
    white_female = fields.Integer(string='WF')
    white_dissabled = fields.Integer(string='WD')

    total_male = fields.Integer(
        string='TM', compute='_compute_total_male', store=True)
    total_female = fields.Integer(
        string='TF', compute='_compute_total_female', store=True)
    total_dissabled = fields.Integer(
        string='TD', compute='_compute_total_disabled', store=True)

    population_group = fields.Selection([
        ('african', 'African'),
        ('white', 'White'),
        ('coloured', 'Coloured'),
        ('indian', 'Indian')
    ])

    gender = fields.Selection([
        ('female', 'F - Female'),
        ('male', 'M - Male')
    ])

    dissability_status_and_type = fields.Selection([
        ('site', '01 - Sight'),
        ('hearing', '02 - Hearing'),
        ('communication', '03 - Communication'),
        ('physical', '04 - Physical'),
        ('intellectual', '05 - Intellectual'),
        ('emotional', '06 - Emotional'),
        ('multiple', '07 - Multiple'),
        ('disabled', '09 - Disabled'),
        ('none', 'N - None')
    ])

    province = fields.Many2one(
        'res.country.state',
        domain=[('country_id.code', '=', 'ZA')]
    )
    municipality_id = fields.Many2one('res.municipality')
    city_id = fields.Many2one('res.city')
    suburb_id = fields.Many2one('res.suburb')

    urban = fields.Selection([
        ('urban', 'Urban'),
        ('rural', 'Rural'),
        ('unknown', 'Unknown')
    ])

    start_date = fields.Date(string='AET Start Date')
    end_date = fields.Date(string='AET End Date')
    provider = fields.Char(string='Provider')

    aet_level = fields.Selection([
        ('aet_level_1', 'AET Level 1'),
        ('aet_level_2', 'AET Level 2'),
        ('aet_level_3', 'AET Level 3'),
        ('aet_level_4', 'AET Level 4')
    ])

    aet_subject = fields.Selection([
        ('life_skills', 'Life Skills'),
        ('numeracy', 'Numeracy'),
        ('literacy', 'Literacy')
    ])

    aet_subject1 = fields.Many2many(
        'aet.subject',
        'aet_subject_rel',
        'actual_adult_education_ids',
        'aet_subject_id',
        string='AET Subjects'
    )

    actual_adult_wsp_id = fields.Many2one('wsp.plan')
    actual_planned_adult_wsp_id = fields.Many2one('wsp.plan')
    save_actual_adult_record = fields.Boolean(default=False)

    # --------------------------------------------------
    # ONCHANGE METHODS
    # --------------------------------------------------

    @api.onchange('city_id')
    def _onchange_city(self):
        for rec in self:
            if rec.city_id:
                rec.urban = rec.city_id.urban_rural

    @api.onchange('municipality_id')
    def _onchange_municipality(self):
        for rec in self:
            if rec.municipality_id:
                rec.urban = rec.municipality_id.urban_rural

    @api.onchange('province')
    def _onchange_province(self):
        for rec in self:
            if rec.province:
                return {
                    'domain': {
                        'municipality_id': [('province_id', '=', rec.province.id)],
                        'city_id': [('province_id', '=', rec.province.id)],
                    }
                }

# class employer_sdl_no(models.Model):
#     _name = 'employer.sdl.no'
#
#     name = fields.Char(string='Name')
#     employer_id = fields.Many2one('res.partner', string='Employer')
#
# employer_sdl_no()


class TotalEmploymentProfileFields(models.Model):
    _name = 'total.employment.profile.fields'
    _description = 'Total Employment Profile Fields'

    line_validated = fields.Boolean(default=False)

    # =========================
    # COMPUTED TOTALS
    # =========================

    @api.depends('african_male', 'coloured_male', 'indian_male', 'white_male')
    def _compute_total_male(self):
        for rec in self:
            rec.total_male = (
                rec.african_male +
                rec.coloured_male +
                rec.indian_male +
                rec.white_male
            )

    @api.depends('african_female', 'coloured_female', 'indian_female', 'white_female')
    def _compute_total_female(self):
        for rec in self:
            rec.total_female = (
                rec.african_female +
                rec.coloured_female +
                rec.indian_female +
                rec.white_female
            )

    @api.depends('african_dissabled', 'coloured_dissabled', 'indian_dissabled', 'white_dissabled')
    def _compute_total_disabled(self):
        for rec in self:
            rec.total_dissabled = (
                rec.african_dissabled +
                rec.coloured_dissabled +
                rec.indian_dissabled +
                rec.white_dissabled
            )

    employer_id = fields.Many2one(
        'res.partner',
        string='Employer',
        default=lambda self: self._context.get('employer')
    )

    sdl_number = fields.Many2one('employer.sdl.no', string='SDL Number')
    citizen_resident_status_code = fields.Selection(
        [('dual', 'D - Dual (SA plus other)'), ('other', 'O - Other'), ('sa', 'SA - South Africa'), ('unknown', 'U - Unknown')], string='Citizen Status')
    name = fields.Char(string='First Name')
    surname = fields.Char(string='Surname')
    employee_id = fields.Char(string='Employee ID', size=20)
    id_type = fields.Selection(
        [('id_document', 'ID Document'), ('passport', 'Passport')], string='ID Type')
    dob = fields.Date(string='Date of Birth(dd/mm/yyyy)')
    ofo_code = fields.Many2one('ofo.code', string='OFO Code ')
    occupation = fields.Many2one('occupation.ofo', string='Occupation')
    specialization = fields.Many2one(
        'specialize.subject', string='Specialisation')
    municipality_id = fields.Many2one(
        'res.municipality', string='Municipality')
    city_id = fields.Many2one('res.city', string='City')
    province = fields.Many2one(
        'res.country.state', string='Province', domain=[('country_id.code', '=', 'ZA')])
    suburb_id = fields.Many2one('res.suburb', string='Suburb')
    urban = fields.Selection(
        [('urban', 'Urban'), ('rural', 'Rural'), ('unknown', 'Unknown')], string='Urban/Rural')
    highest_education_level = fields.Selection([('abet_level_1', 'Abet Level 1'), ('abet_level_2', 'Abet Level 2'), ('abet_level_3', 'Abet Level 3'), (
        'abet_level_4', 'Abet Level 4'), ('nqf123', 'NQF 1,2,3'), ('nqf45', 'NQF 4,5'), ('nqf67', 'NQF 6,7'), ('nqf8910', 'NQF 8,9,10')], string='Highest Education Level')
    scarce_skill = fields.Char(string='Scarce Skill')
    # Needs to keep these fields for this year.
    african_male = fields.Integer(string='AM', help='African Male')
    african_female = fields.Integer(string='AF', help='African Female')
    african_dissabled = fields.Integer(string='AD', help='African Disabled')
    coloured_male = fields.Integer(string='CM', help='Coloured Male')
    coloured_female = fields.Integer(string='CF', help='Coloured Female')
    coloured_dissabled = fields.Integer(string='CD', help='Coloured Disabled')
    indian_male = fields.Integer(string='IM', help='Indian Male')
    indian_female = fields.Integer(string='IF', help='Indian Female')
    indian_dissabled = fields.Integer(string='ID', help='Indian Disabled')
    white_male = fields.Integer(string='WM', help='White Male')
    white_female = fields.Integer(string='WF', help='White Female')
    white_dissabled = fields.Integer(string='WD', help='White Disabled')
    total_male = fields.Integer(compute='_compute_total_male', store=True)
    total_female = fields.Integer(compute='_compute_total_female', store=True)
    total_dissabled = fields.Integer(compute='_compute_total_disabled', store=True)
    age_group_less = fields.Integer(
        string='<35', help='Age Group Less than 35')
    age_group_upto = fields.Integer(
        string='35-55', help='Age Group from 35 to 55')
    age_group_greater = fields.Integer(
        string='>55', help='Age Group greater than 55')
    ##
    population_group = fields.Selection(
        [('african', 'African'), ('white', 'White'), ('coloured', 'Coloured'), ('indian', 'Indian')], string='Race')
    gender = fields.Selection(
        [('female', 'F - Female'), ('male', 'M - Male')], string='Gender')
    age_group = fields.Selection([('less_than_thirty_five', '<35'),
                                  ('thirty_five_to_fifty_five', '35-55'),
                                  ('greater_than_fifty_five', '>55')], string='Age Group')
    dissability = fields.Selection(
        [('yes', 'Yes'), ('no', 'No')], string='Disability', default='no')
    total_employment_wsp_id = fields.Many2one(
        'wsp.plan', string='Total Employment Profile')
    specialisation_exists = fields.Boolean(string='Sepcialise Exists')
    selection = fields.Boolean(string='Select', default=True)
    save_total_employed_record = fields.Boolean("Partial TEP", default=False)

    
    @api.onchange('ofo_code')
    def onchange_code(self):
        for rec in self:
            if not rec.ofo_code:
                rec.occupation = False
                rec.specialisation_exists = False
                return

            values = get_occupation_and_specialization(rec.ofo_code, self.env)

            rec.occupation = values[0] or False

            if values[1]:
                rec.specialisation_exists = True
                return {
                    'domain': {'specialization': [('id', 'in', values[1])]}
                }
            else:
                rec.specialisation_exists = False
                return {
                    'domain': {'specialization': []}
                }

    @api.onchange('ofo_code')
    def onchange_code(self):
        for rec in self:
            if not rec.ofo_code:
                rec.occupation = False
                rec.specialisation_exists = False
                return

            values = get_occupation_and_specialization(rec.ofo_code, self.env)

            rec.occupation = values[0] or False

            if values[1]:
                rec.specialisation_exists = True
                return {
                    'domain': {'specialization': [('id', 'in', values[1])]}
                }
            else:
                rec.specialisation_exists = False
                return {
                    'domain': {'specialization': []}
                }


    @api.onchange('city_id')
    def onchange_city(self):
        for rec in self:
            rec.urban = rec.city_id.urban_rural if rec.city_id else False


    
    @api.onchange('province')
    def onchange_province_id(self):
        for rec in self:
            if not rec.province:
                rec.city_id = False
                return

            city_ids = self.env['res.city'].search([
                ('province_id', '=', rec.province.id)
            ]).ids

            return {
                'domain': {'city_id': [('id', 'in', city_ids)]}
            }

    @api.onchange('suburb_id')
    def onchange_suburb_id(self):
        for rec in self:
            if not rec.suburb_id:
                return

            rec.city_id = rec.suburb_id.city_id
            rec.urban = rec.suburb_id.urban_rural


    
    @api.onchange('citizen_resident_status_code')
    def onchange_citizen_code(self):
        for rec in self:
            if rec.citizen_resident_status_code == 'dual':
                rec.id_type = 'passport'


    
    @api.onchange('employer_id')
    def onchange_employer_id(self):
        for rec in self:
            if not rec.employer_id:
                rec.sdl_number = False
                return {'domain': {'sdl_number': []}}

            employer_ids = (
                rec.employer_id.child_employer_ids.mapped('employer_id').ids
                + [rec.employer_id.id]
            )

            sdl_ids = self.env['employer.sdl.no'].search([
                ('employer_id', 'in', employer_ids)
            ]).ids

            rec.sdl_number = sdl_ids[0] if sdl_ids else False

            return {
                'domain': {'sdl_number': [('id', 'in', sdl_ids)]}
            }


    @api.onchange('identification_id', 'crs_code', 'id_type')
    def onchange_id_no(self):
        for rec in self:
            if not rec.identification_id:
                return

            id_no = rec.identification_id.strip()

            # 🇿🇦 South African / Dual
            if rec.crs_code in ('sa', 'dual') and rec.id_type == 'id_document':

                if not id_no.isdigit():
                    raise UserError(_('Employee ID should be numeric!'))

                if len(id_no) < 6:
                    rec.identification_id = False
                    return {
                        'warning': {
                            'title': _('Invalid Identification Number'),
                            'message': _('Identification number is too short.')
                        }
                    }

                year = id_no[:2]
                month = id_no[2:4]
                day = id_no[4:6]

                try:
                    year_int = int(year)
                    month_int = int(month)
                    day_int = int(day)
                except ValueError:
                    rec.identification_id = False
                    return {
                        'warning': {
                            'title': _('Invalid Identification Number'),
                            'message': _('Identification number format is invalid.')
                        }
                    }

                if not (1 <= month_int <= 12):
                    rec.identification_id = False
                    return {
                        'warning': {
                            'title': _('Invalid Identification Number'),
                            'message': _('Invalid month in identification number.')
                        }
                    }

                full_year = 2000 + year_int if year_int <= 20 else 1900 + year_int
                last_day = calendar.monthrange(full_year, month_int)[1]

                if not (1 <= day_int <= last_day):
                    rec.identification_id = False
                    return {
                        'warning': {
                            'title': _('Invalid Identification Number'),
                            'message': _('Invalid day in identification number.')
                        }
                    }

                rec.dob = datetime.strptime(
                    f"{full_year}-{month}-{day}", "%Y-%m-%d"
                ).date()

            # 🌍 Passport / Foreign
            elif rec.crs_code not in ('sa', 'dual'):
                if not id_no.isalnum():
                    raise UserError(_('Employee ID should be alphanumeric!'))

            # Passport with SA / Dual → no DOB logic
            elif rec.id_type == 'passport':
                pass



class PlannedTrainingFields(models.Model):
    _name = 'planned.training.fields'
    _description = 'Planned Training Fields'

    line_validated = fields.Boolean(default=False)

    # =========================================================
    # COMPUTED FIELDS
    # =========================================================

    @api.depends('gender')
    def _compute_total_male(self):
        for rec in self:
            rec.total_male = 1 if rec.gender == 'male' else 0

    @api.depends('gender')
    def _compute_total_female(self):
        for rec in self:
            rec.total_female = 1 if rec.gender == 'female' else 0

    @api.depends(
        'african_dissabled',
        'coloured_dissabled',
        'indian_dissabled',
        'white_dissabled'
    )
    def _compute_total_disabled(self):
        for rec in self:
            rec.total_dissabled = (
                rec.african_dissabled +
                rec.coloured_dissabled +
                rec.indian_dissabled +
                rec.white_dissabled
            )

    @api.depends('total_male', 'total_female', 'training_cost')
    def _compute_total_cost(self):
        for rec in self:
            rec.total_cost = (rec.total_male + rec.total_female) * rec.training_cost

    training_type = fields.Selection(
        [('pivotal', 'Pivotal'), ('non-pivotal', 'Non Pivotal')],
        string='Type of Training'
    )

    name = fields.Char(string='First Name')
    surname = fields.Char(string='Surname')

    citizen_resident_status_code = fields.Selection(
        [
            ('dual', 'D - Dual (SA plus other)'),
            ('other', 'O - Other'),
            ('sa', 'SA - South Africa'),
            ('unknown', 'U - Unknown')
        ],
        string='Citizen Status'
    )

    id_type = fields.Selection(
        [('id_document', 'ID Document'), ('passport', 'Passport')],
        string='ID Type'
    )

    code = fields.Many2one('ofo.code', string='OFO Code')

    municipality_id = fields.Many2one('res.municipality', string='Municipality')
    city_id = fields.Many2one('res.city', string='City')

    learner_province = fields.Many2one(
        'res.country.state',
        string='Province',
        domain=[('country_id.code', '=', 'ZA')]
    )

    suburb_id = fields.Many2one('res.suburb', string='Suburb')

    urban = fields.Selection(
        [('urban', 'Urban'), ('rural', 'Rural'), ('unknown', 'Unknown')],
        string='Urban / Rural'
    )

    employee_id = fields.Char(string='Employee ID', size=20)
    # Needs to keep these fields for this year.
    african_male = fields.Integer(string='AM')
    african_female = fields.Integer(string='AF')
    african_dissabled = fields.Integer(string='AD')

    coloured_male = fields.Integer(string='CM')
    coloured_female = fields.Integer(string='CF')
    coloured_dissabled = fields.Integer(string='CD')

    indian_male = fields.Integer(string='IM')
    indian_female = fields.Integer(string='IF')
    indian_dissabled = fields.Integer(string='ID')

    white_male = fields.Integer(string='WM')
    white_female = fields.Integer(string='WF')
    white_dissabled = fields.Integer(string='WD')

    total_male = fields.Integer(
        string='TM',
        compute='_compute_total_male',
        store=True
    )

    total_female = fields.Integer(
        string='TF',
        compute='_compute_total_female',
        store=True
    )

    total_dissabled = fields.Integer(
        string='TD',
        compute='_compute_total_disabled',
        store=True
    )
    age_group = fields.Selection(
        [
            ('less_than_thirty_five', '<35'),
            ('thirty_five_to_fifty_five', '35-55'),
            ('greater_than_fifty_five', '>55')
        ],
        string='Age Group'
    )

    population_group = fields.Selection(
        [
            ('african', 'African'),
            ('coloured', 'Coloured'),
            ('indian', 'Indian'),
            ('white', 'White')
        ],
        string='Population Group'
    )

    gender = fields.Selection(
        [('female', 'Female'), ('male', 'Male')],
        string='Gender'
    )

    occupation = fields.Many2one('occupation.ofo', string='Occupation')
    specialization = fields.Many2one('specialize.subject', string='Specialisation')

    type_training = fields.Many2one(
        'training.intervention',
        string='Training Intervention / Pivotal Programme'
    )
    other_type_of_intervention = fields.Char(
        string='Other Type of Training Intervention')
    socio_economic_status = fields.Selection(
        [('employed', 'Employed'), ('unemployed', 'Unemployed')], string='Socio Economic Status')
    pivotal_programme_institution = fields.Char(
        string='Pivotal Programme Institution')
    pivotal_programme_qualification = fields.Char(
        string='Pivotal Programme Qualification')
    name_training = fields.Char(string='Training Name')

    start_date = fields.Date(string='Start Date')
    end_date = fields.Date(string='End Date')

    training_cost = fields.Float(string='Cost')

    total_cost = fields.Float(
        string='Total Cost',
        compute='_compute_total_cost',
        store=True
    )
    planned_training_non_wsp_id = fields.Many2one(
        'wsp.plan',
        string='Planned Training (Non WSP)'
    )

    specialisation_exists = fields.Boolean(string='Specialisation Exists')
    save_planned_record = fields.Boolean(string='Save Partial', default=False)
    nqf_aligned = fields.Selection(
        [('yes', 'Yes'), ('no', 'No')], string='NQF Aligned')
    nqf_level = fields.Selection(
        [('abet', 'ABET'), ('below_level_1', 'Below Level 1'), ('level1', 'Level 1'), ('level2', 'Level 2'),
         ('level3', 'Level 3'), ('level4', 'Level 4'), (
             'level5', 'Level 5'), ('level6', 'Level 6'), ('level7', 'Level 7'), ('level8', 'Level 8'),
         ('level9', 'Level 9'), ('level10', 'Level 10')], string='NQF Level')
    dissability = fields.Selection(
        [('yes', 'Yes'), ('no', 'No')], string='Disability', default='no')

    @api.onchange('code')
    def _onchange_code(self):
        if not self.code:
            self.occupation = False
            self.specialisation_exists = False
            return {'domain': {'specialization': []}}

        ofo_code_data = self.code

        values = get_occupation_and_specialization(ofo_code_data, self.env)

        # Set occupation
        self.occupation = values[0] or False

        # Handle specialisations
        if values[1]:
            self.specialisation_exists = True
            return {
                'domain': {
                    'specialization': [('id', 'in', values[1])]
                }
            }
        else:
            self.specialisation_exists = False
            self.specialization = False
            return {
                'domain': {
                    'specialization': []
                }
            }

    @api.onchange('city_id')
    def _onchange_city(self):
        if not self.city_id:
            self.urban = False
            return

        self.urban = self.city_id.urban_rural

    @api.onchange('learner_province')
    def _onchange_learner_province(self):
        if not self.learner_province:
            self.city_id = False
            return {
                'domain': {
                    'city_id': []
                }
            }

        city_ids = self.env['res.city'].search([
            ('province_id', '=', self.learner_province.id)
        ]).ids

        self.city_id = False

        return {
            'domain': {
                'city_id': [('id', 'in', city_ids)]
            }
        }

    
    def onchange_suburb_id(self, suburb_id):
        res = {}
        if not suburb_id:
            return res
        suburb_data = self.env['res.suburb'].browse(suburb_id)
        city_id = suburb_data.city_id and suburb_data.city_id.id or None
        res.update({'domain': {'city_id': [('id', '=', city_id)]}, 'value': {
                   'city_id': city_id, 'urban': suburb_data.urban_rural}})
        return res

    @api.onchange('training_type')
    def _onchange_training_type(self):
        if not self.training_type:
            self.type_training = False
            return {'domain': {'type_training': []}}

        training_intervention_obj = self.env['training.intervention']

        # NON-PIVOTAL
        if self.training_type == 'non-pivotal':
            intervention_ids = training_intervention_obj.search([
                ('pivotal', '=', False)
            ]).ids

            self.socio_economic_status = 'employed'
            self.nqf_aligned = False

            return {
                'domain': {
                    'type_training': [('id', 'in', intervention_ids)]
                }
            }

        # PIVOTAL
        if self.training_type == 'pivotal':
            pivotal_ids = training_intervention_obj.search([
                ('pivotal', '=', True)
            ]).ids

            self.nqf_aligned = 'yes'
            self.socio_economic_status = False

            return {
                'domain': {
                    'type_training': [('id', 'in', pivotal_ids)]
                }
            }


class PlannedAdultEducationTrainingFields(models.Model):
    _name = 'planned.adult.education.training.fields'
    _description = 'Planned Adult Education Training Fields'

    line_validated = fields.Boolean(default=False)

    # ----------------------------------------------------
    # COMPUTED TOTALS
    # ----------------------------------------------------

    total_male = fields.Integer(
        string='TM', compute='_compute_total_male', store=True)
    total_female = fields.Integer(
        string='TF', compute='_compute_total_female', store=True)
    total_dissabled = fields.Integer(
        string='TD', compute='_compute_total_disabled', store=True)

    @api.depends('african_male', 'coloured_male', 'indian_male', 'white_male')
    def _compute_total_male(self):
        for rec in self:
            rec.total_male = (
                rec.african_male +
                rec.coloured_male +
                rec.indian_male +
                rec.white_male
            )

    @api.depends('african_female', 'coloured_female', 'indian_female', 'white_female')
    def _compute_total_female(self):
        for rec in self:
            rec.total_female = (
                rec.african_female +
                rec.coloured_female +
                rec.indian_female +
                rec.white_female
            )

    @api.depends('african_dissabled', 'coloured_dissabled', 'indian_dissabled', 'white_dissabled')
    def _compute_total_disabled(self):
        for rec in self:
            rec.total_dissabled = (
                rec.african_dissabled +
                rec.coloured_dissabled +
                rec.indian_dissabled +
                rec.white_dissabled
            )

    # ----------------------------------------------------
    # ORM REPLACEMENT FOR RAW SQL
    # ----------------------------------------------------

    def get_aet_subject(self):
        """Return comma-separated AET subject names"""
        self.ensure_one()
        return ', '.join(self.aet_subject1.mapped('name'))

    # ----------------------------------------------------
    # FIELDS
    # ----------------------------------------------------

    name = fields.Char(string='First Name')
    surname = fields.Char(string='Surname')

    citizen_resident_status_code = fields.Selection([
        ('dual', 'D - Dual (SA plus other)'),
        ('other', 'O - Other'),
        ('sa', 'SA - South Africa'),
        ('unknown', 'U - Unknown')
    ], string='Citizen Status')

    id_type = fields.Selection([
        ('id_document', 'ID Document'),
        ('passport', 'Passport')
    ], string='ID Type')

    ofo_code = fields.Many2one('ofo.code', string='OFO Code')
    id_number = fields.Char(string='ID Number', size=20)

    african_male = fields.Integer('AM')
    african_female = fields.Integer('AF')
    african_dissabled = fields.Integer('AD')

    coloured_male = fields.Integer('CM')
    coloured_female = fields.Integer('CF')
    coloured_dissabled = fields.Integer('CD')

    indian_male = fields.Integer('IM')
    indian_female = fields.Integer('IF')
    indian_dissabled = fields.Integer('ID')

    white_male = fields.Integer('WM')
    white_female = fields.Integer('WF')
    white_dissabled = fields.Integer('WD')

    population_group = fields.Selection([
        ('african', 'African'),
        ('white', 'White'),
        ('coloured', 'Coloured'),
        ('indian', 'Indian')
    ])

    gender = fields.Selection([
        ('female', 'Female'),
        ('male', 'Male')
    ])

    dissability_status_and_type = fields.Selection([
        ('site', '01 - Sight'),
        ('hearing', '02 - Hearing'),
        ('communication', '03 - Communication'),
        ('physical', '04 - Physical'),
        ('intellectual', '05 - Intellectual'),
        ('emotional', '06 - Emotional'),
        ('multiple', '07 - Multiple'),
        ('disabled', '09 - Disabled'),
        ('none', 'N - None')
    ])

    municipality_id = fields.Many2one('res.municipality')
    city_id = fields.Many2one('res.city')
    province = fields.Many2one(
        'res.country.state',
        domain=[('country_id.code', '=', 'ZA')]
    )
    suburb_id = fields.Many2one('res.suburb')

    urban = fields.Selection([
        ('urban', 'Urban'),
        ('rural', 'Rural'),
        ('unknown', 'Unknown')
    ])

    start_date = fields.Date('AET Start Date')
    end_date = fields.Date('AET End Date')
    provider = fields.Char('Provider')

    aet_level = fields.Selection([
        ('aet_level_1', 'AET Level 1'),
        ('aet_level_2', 'AET Level 2'),
        ('aet_level_3', 'AET Level 3'),
        ('aet_level_4', 'AET Level 4')
    ])
    aet_subject = fields.Selection(
        [('life_skills', 'Life Skills'), ('numeracy', 'Numeracy'), ('literacy', 'Literacy')], string='AET Subject')
    aet_subject1 = fields.Many2many(
        'aet.subject',
        'aet_subject_planned_rel',
        'planed_adult_education_training_id',
        'aet_subject_id',
        string='AET Subjects'
    )

    reason = fields.Char()
    planned_adult_education_wsp_id = fields.Many2one('wsp.plan')
    save_planned_adult_record = fields.Boolean(default=False)

    # ----------------------------------------------------
    # ONCHANGES (ODOO 18 STYLE)
    # ----------------------------------------------------

    @api.onchange('ofo_code')
    def _onchange_ofo_code(self):
        if not self.ofo_code:
            return

        values = get_occupation_and_specialization(self.ofo_code, self.env)
        return {
            'domain': {
                'specialization': [('id', 'in', values[1])]
            }
        }

    @api.onchange('city_id')
    def _onchange_city(self):
        if self.city_id:
            self.urban = self.city_id.urban_rural

    @api.onchange('province')
    def _onchange_province(self):
        if not self.province:
            self.city_id = False
            return

        city_ids = self.env['res.city'].search([
            ('province_id', '=', self.province.id)
        ]).ids

        return {
            'domain': {
                'city_id': [('id', 'in', city_ids)]
            }
        }

    @api.onchange('suburb_id')
    def _onchange_suburb(self):
        if self.suburb_id:
            self.city_id = self.suburb_id.city_id
            self.urban = self.suburb_id.urban_rural


from odoo import models, fields, api


class ScarceAndCriticalSkillsFields(models.Model):
    _name = 'scarce.and.critical.skills.fields'
    _description = 'Scarce and Critical Skills Fields'

    # ----------------------------------------------------
    # COMPUTED TOTALS
    # ----------------------------------------------------

    total_male = fields.Integer(
        string='TM', compute='_compute_total_male', store=True)
    total_female = fields.Integer(
        string='TF', compute='_compute_total_female', store=True)
    total_dissabled = fields.Integer(
        string='TD', compute='_compute_total_disabled', store=True)

    @api.depends('african_male', 'coloured_male', 'indian_male', 'white_male')
    def _compute_total_male(self):
        for rec in self:
            rec.total_male = (
                rec.african_male +
                rec.coloured_male +
                rec.indian_male +
                rec.white_male
            )

    @api.depends('african_female', 'coloured_female', 'indian_female', 'white_female')
    def _compute_total_female(self):
        for rec in self:
            rec.total_female = (
                rec.african_female +
                rec.coloured_female +
                rec.indian_female +
                rec.white_female
            )

    @api.depends('african_dissabled', 'coloured_dissabled', 'indian_dissabled', 'white_dissabled')
    def _compute_total_disabled(self):
        for rec in self:
            rec.total_dissabled = (
                rec.african_dissabled +
                rec.coloured_dissabled +
                rec.indian_dissabled +
                rec.white_dissabled
            )

    # ----------------------------------------------------
    # FIELDS
    # ----------------------------------------------------

    name = fields.Char(string='First Name')
    surname = fields.Char(string='Surname')

    ofo_code = fields.Many2one('ofo.code', string='OFO Code')
    occupation = fields.Many2one('occupation.ofo', string='Occupation')
    specialization = fields.Many2one('specialize.subject', string='Specialisation')

    scarce_skill = fields.Char(string='Scarce Skills')
    critical_skill = fields.Char(string='Critical Skills')

    number_of_vacancies = fields.Integer(string='Number Of Vacancies')
    number_of_potential_vacancies = fields.Integer(string='Number Of Potential Vacancies')

    nqf_level = fields.Selection([
        ('abet', 'ABET'),
        ('below_level_1', 'Below Level 1'),
        ('level1', 'Level 1'),
        ('level2', 'Level 2'),
        ('level3', 'Level 3'),
        ('level4', 'Level 4'),
        ('level5', 'Level 5'),
        ('level6', 'Level 6'),
        ('level7', 'Level 7'),
        ('level8', 'Level 8'),
        ('level9', 'Level 9'),
        ('level10', 'Level 10')
    ], string='NQF Level')

    degree_of_scarcity = fields.Char(string='Degree of Scarcity')
    reason_for_scarcity = fields.Char(string='Reason for Scarcity/Critical')

    african_male = fields.Integer('AM')
    african_female = fields.Integer('AF')
    african_dissabled = fields.Integer('AD')

    coloured_male = fields.Integer('CM')
    coloured_female = fields.Integer('CF')
    coloured_dissabled = fields.Integer('CD')

    indian_male = fields.Integer('IM')
    indian_female = fields.Integer('IF')
    indian_dissabled = fields.Integer('ID')

    white_male = fields.Integer('WM')
    white_female = fields.Integer('WF')
    white_dissabled = fields.Integer('WD')

    population_group = fields.Selection([
        ('african', 'African'),
        ('white', 'White'),
        ('coloured', 'Coloured'),
        ('indian', 'Indian')
    ], string='Population Group')

    gender = fields.Selection([
        ('female', 'Female'),
        ('male', 'Male')
    ], string='Gender')

    planned_strategy_address = fields.Char(
        string='Planned Strategy to address the scarcity')

    province = fields.Many2one(
        'res.country.state',
        string='Province',
        domain=[('country_id.code', '=', 'ZA')]
    )

    is_reflected = fields.Boolean(
        string='Is this reflected to your EE Plan?')

    comments = fields.Char(string='Comments')

    scarce_and_critical_wsp_id = fields.Many2one(
        'wsp.plan',
        string='Scarce and Critical Skills'
    )

    no_of_months = fields.Integer(
        string='Number of Months Position has been vacant',
        default=1
    )

    specialisation_exists = fields.Boolean(
        string='Specialisation Exists')

    save_variance_record = fields.Boolean(
        string='Partial Variance',
        default=False
    )

    # ----------------------------------------------------
    # ONCHANGE (ODOO 18 STYLE)
    # ----------------------------------------------------

    @api.onchange('ofo_code')
    def _onchange_ofo_code(self):
        if not self.ofo_code:
            self.occupation = False
            self.specialization = False
            self.specialisation_exists = False
            return

        values = get_occupation_and_specialization(self.ofo_code, self.env)

        self.occupation = values[0] or False

        if values[1]:
            self.specialisation_exists = True
            return {
                'domain': {
                    'specialization': [('id', 'in', values[1])]
                }
            }
        else:
            self.specialisation_exists = False
            return {
                'domain': {
                    'specialization': []
                }
            }



# class res_partner(models.Model):
#     _inherit = 'res.partner'
#
#
#     def get_line(self, name, debit, credit, account_id, partner):
#         result = {
#             'analytic_account_id': False,
#             'tax_code_id': False,
#             'analytic_lines': [],
#             'tax_amount': False,
#             'name': name,
#             'ref': name,
#             'asset_id': False,
#             'currency_id': False,
#             'credit': credit,
#             'product_id': False,
#             'date_maturity': False,
#             'debit': debit,
#             'date': datetime.now().date(),
#             'amount_currency': 0,
#             'product_uom_id': False,
#             'quantity': 0,
#             'partner_id': partner,
#             'account_id': account_id,
#         }
#         return result
#
#
#     def get_move_line(self, name, amount, credit_account_id, debit_account_id, partner_id):
#         result = []
#         credit_line_dict = self.get_line(
#             name, 0, amount, credit_account_id, partner_id)
#         result.append((0, 0, credit_line_dict))
#         debit_line_dict = self.get_line(
#             name, amount, 0, debit_account_id, partner_id)
#         result.append((0, 0, debit_line_dict))
#         return result
#
#
#     def get_move_val(self, ref, line_id, journal_id, date, narration, company_id):
#         result = {
#             'ref': ref,
#             'line_id': line_id,
#             'journal_id': journal_id,
#             'date': date,
#             'narration': narration,
#             'company_id': company_id,
#         }
#         return result
#
#     @api.model
#     def make_dormant(self):
#         # Discretionary Reserve Account Configurations submission_end_date
#         # check
#         leavy_config_data = self.env['leavy.income.config'].search([])
#         move_obj = self.env['account.move']
# #         if leavy_config_data:
# #             invoice_data = self.env['account.invoice'].search([('employer','=',True),('state','!=','suspend'),('leavy_period','!=',None)])
# #             for invoice_obj in invoice_data:
# #                 if invoice_obj.date_invoice > leavy_config_data[0].submission_end_date:
# #                     if invoice_obj.invoice_line:
# #                         disc_val = self.get_move_line('Discretionary Reserve Amt',invoice_obj.invoice_line[0].price_subtotal, leavy_config_data[0].discretionary_reserve_acc.id, invoice_obj.account_id.id , invoice_obj.partner_id.id)
# #                         move_val = self.get_move_val('Discretionary Reserve Amt', disc_val, invoice_obj.journal_id.id, datetime.now().date(), 'Discretionary Reserve Amt', invoice_obj.partner_id.company_id.id)
# #                         move_obj.create(move_val)
# #                         invoice_obj.write({'state':'suspend'})
#         all_employers = self.env['res.partner'].search(
#             [('employer', '=', True)])
#         for employer in all_employers:
#             wsp_submission_years = []
#             for wsp in self.env['wsp.plan'].search([('employer_id', '=', employer.id), ('wsp_submission_date', '!=', None)]):
#                 wsp_submission_date = datetime.strptime(
#                     wsp.wsp_submission_date, "%Y-%m-%d").date()
#                 wsp_submission_years.append(wsp_submission_date)
#             if wsp_submission_years:
#                 max_date = max(wsp_submission_years)
#                 current_date = datetime.today().date()
#                 date_diff = abs((current_date - max_date).days)
#                 if date_diff > 365:
#                     employer.write({'dormant': True})
#         return True
#
# res_partner()
